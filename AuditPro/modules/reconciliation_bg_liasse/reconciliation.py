import logging
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

LOGGER = logging.getLogger(__name__)

WARNING_THRESHOLD = 100.0
CRITICAL_THRESHOLD = 10000.0
PERCENTAGE_UNDEFINED = None


RUBRIC_PATTERNS = {
    "TOTAL ACTIF": ["total actif"],
    "CAPITAUX PROPRES": ["capitaux propres", "capital propre"],
    "TOTAL PASSIF": ["total passif"],
    "CHIFFRE D'AFFAIRES": ["chiffre d'affaires", "chiffre affaire", "ca"],
    "RESULTAT NET": ["resultat net", "resultat de l'exercice", "benefice net"],
    "TRESORERIE ACTIF": ["banques", "banque", "caisse", "tresorerie actif"],
    "FOURNISSEURS": ["fournisseurs"],
    "CLIENTS": ["clients", "creances clients"],
}


ACCOUNT_RUBRIC_MAPPING = {
    "TOTAL ACTIF": ["2", "3", "4", "5"],
    "CAPITAUX PROPRES": ["11"],
    "TOTAL PASSIF": ["1", "4", "5"],
    "CHIFFRE D'AFFAIRES": ["71"],
    "RESULTAT NET": ["81", "83", "84", "86", "88"],
    "TRESORERIE ACTIF": ["51", "53"],
    "FOURNISSEURS": ["44"],
    "CLIENTS": ["34"],
}


@dataclass
class ReconciliationBundle:
    bg_accounts: pd.DataFrame
    liasse_rubrics: dict[str, float]
    reconciliation_df: pd.DataFrame


def _get_severity(ecart: float) -> str:
    abs_ecart = abs(float(ecart))
    if abs_ecart > CRITICAL_THRESHOLD:
        return "CRITIQUE"
    if abs_ecart > WARNING_THRESHOLD:
        return "ATTENTION"
    return "OK"


def _compute_percentage(ecart: float, liasse_total: float) -> float | None:
    if liasse_total:
        return ecart / liasse_total * 100.0
    return 0.0 if ecart == 0 else PERCENTAGE_UNDEFINED


def _excel_percentage_value(value: float | None):
    return float(value) if pd.notna(value) else PERCENTAGE_UNDEFINED


def _normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def _to_amount(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "-"}:
        return 0.0
    text = text.replace("\u202f", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-()]", "", text)
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    if text.count(",") > 0 and text.count(".") > 0:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") > 0:
        text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


def _detect_sheet(xls: pd.ExcelFile, keywords: list[str], default_index: int = 0) -> str:
    normalized = {name: _normalize_text(name) for name in xls.sheet_names}
    for name, normalized_name in normalized.items():
        if any(keyword in normalized_name for keyword in keywords):
            return name
    if not xls.sheet_names:
        raise ValueError("Aucun onglet trouvé dans le fichier Excel.")
    return xls.sheet_names[default_index]


def _find_bg_header_row(raw_df: pd.DataFrame) -> int:
    for row_idx in range(min(len(raw_df), 30)):
        row_values = [_normalize_text(v) for v in raw_df.iloc[row_idx].tolist()]
        has_compte = any("compte" in value for value in row_values)
        has_label = any("libelle" in value or "intitule" in value for value in row_values)
        if has_compte and has_label:
            return row_idx
    raise ValueError(
        "En-tête Balance Générale introuvable. Colonnes attendues : compte + libellé/intitulé."
    )


def _pick_column(columns: list[str], choices: list[str]) -> str | None:
    normalized_cols = {_normalize_text(col): col for col in columns}
    for col_norm, original in normalized_cols.items():
        if any(choice in col_norm for choice in choices):
            return original
    return None


def load_balance_generale(path: str | Path) -> pd.DataFrame:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Fichier BG introuvable : {file_path}")

    xls = pd.ExcelFile(file_path)
    sheet_name = _detect_sheet(xls, ["bg", "balance", "sheet1"])
    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    header_row = _find_bg_header_row(raw)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    if df.empty:
        raise ValueError("La Balance Générale est vide.")

    account_col = _pick_column(df.columns.tolist(), ["compte", "account", "numero"])
    label_col = _pick_column(df.columns.tolist(), ["libelle", "intitule", "label", "designation"])
    balance_col = _pick_column(df.columns.tolist(), ["solde", "balance", "net", "montant"])
    debit_col = _pick_column(df.columns.tolist(), ["debit"])
    credit_col = _pick_column(df.columns.tolist(), ["credit", "crédit"])

    if not account_col or not label_col:
        raise ValueError("Colonnes 'Compte' et 'Libellé/Intitulé' non détectées dans la BG.")

    cleaned = pd.DataFrame()
    cleaned["account"] = (
        df[account_col]
        .astype(str)
        .str.replace(r"[^0-9]", "", regex=True)
        .str.strip()
    )
    cleaned["label"] = df[label_col].fillna("").astype(str).str.strip()

    if balance_col:
        cleaned["balance"] = df[balance_col].apply(_to_amount)
    elif debit_col and credit_col:
        cleaned["balance"] = df[debit_col].apply(_to_amount) - df[credit_col].apply(_to_amount)
    else:
        raise ValueError("Colonne de solde introuvable (ou couple Débit/Crédit absent) dans la BG.")

    cleaned = cleaned[
        cleaned["account"].str.len().ge(3)
        & cleaned["account"].str.match(r"^\d+$", na=False)
        & cleaned["label"].ne("")
    ].copy()
    cleaned["balance"] = pd.to_numeric(cleaned["balance"], errors="coerce").fillna(0.0)

    if cleaned.empty:
        raise ValueError("Aucun compte valide détecté dans la Balance Générale.")

    LOGGER.info("BG chargée : %s comptes valides", len(cleaned))
    return cleaned


def _extract_rubrics_from_dataframe(df: pd.DataFrame) -> dict[str, float]:
    rubrics: dict[str, float] = {}
    if df.empty:
        return rubrics

    values = df.fillna("").astype(str).values.tolist()
    for row in values:
        if not row:
            continue
        text_cells = [str(cell).strip() for cell in row if str(cell).strip()]
        if not text_cells:
            continue
        row_text = " ".join(text_cells)
        row_normalized = _normalize_text(row_text)
        amount = 0.0
        for cell in reversed(row):
            parsed = _to_amount(cell)
            if parsed != 0.0:
                amount = parsed
                break
        for rubric, patterns in RUBRIC_PATTERNS.items():
            if any(pattern in row_normalized for pattern in patterns):
                rubrics[rubric] = amount
    return rubrics


def _load_liasse_excel(path: Path) -> dict[str, float]:
    xls = pd.ExcelFile(path)
    sheets_to_scan = []
    for keyword in ["actif", "passif", "cpc", "compte", "resultat", "sheet1"]:
        try:
            sheet_name = _detect_sheet(xls, [keyword])
            if sheet_name not in sheets_to_scan:
                sheets_to_scan.append(sheet_name)
        except Exception:
            continue
    if not sheets_to_scan and xls.sheet_names:
        sheets_to_scan = xls.sheet_names[:3]

    rubrics: dict[str, float] = {}
    for sheet in sheets_to_scan:
        df = pd.read_excel(path, sheet_name=sheet, header=None)
        rubrics.update(_extract_rubrics_from_dataframe(df))
    return rubrics


def _load_liasse_pdf(path: Path) -> dict[str, float]:
    rubrics: dict[str, float] = {}
    errors = []

    try:
        import pdfplumber

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                rows = [[line] for line in text.splitlines() if line.strip()]
                if rows:
                    rubrics.update(_extract_rubrics_from_dataframe(pd.DataFrame(rows)))
                tables = page.extract_tables() or []
                for table in tables:
                    rubrics.update(_extract_rubrics_from_dataframe(pd.DataFrame(table)))
    except ImportError as exc:  # pragma: no cover
        errors.append(f"pdfplumber indisponible ({exc}). Installez avec: pip install pdfplumber")
    except (OSError, ValueError, RuntimeError) as exc:  # pragma: no cover
        errors.append(f"Échec extraction pdfplumber ({exc}). Tentative du fallback tabula-py.")

    if rubrics:
        return rubrics

    try:
        import tabula

        tables = tabula.read_pdf(str(path), pages="all", multiple_tables=True)
        for table in tables or []:
            rubrics.update(_extract_rubrics_from_dataframe(table))
    except ImportError as exc:  # pragma: no cover
        errors.append(f"tabula-py indisponible ({exc}). Installez avec: pip install tabula-py")
    except (OSError, ValueError, RuntimeError) as exc:  # pragma: no cover
        errors.append(f"Échec extraction tabula-py ({exc}).")

    if not rubrics:
        raise ValueError(
            "Impossible d'extraire les rubriques de la liasse PDF. "
            + (" | ".join(errors) if errors else "")
        )
    return rubrics


def load_liasse(path: str | Path) -> dict[str, float]:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Fichier liasse introuvable : {file_path}")

    ext = file_path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        rubrics = _load_liasse_excel(file_path)
    elif ext == ".pdf":
        rubrics = _load_liasse_pdf(file_path)
    else:
        raise ValueError("Format liasse non supporté. Utilisez Excel (.xlsx/.xls) ou PDF.")

    for required in RUBRIC_PATTERNS:
        rubrics.setdefault(required, 0.0)
    LOGGER.info("Liasse chargée : %s rubriques détectées", len(rubrics))
    return rubrics


def reconcile(bg_accounts: pd.DataFrame, liasse_rubrics: dict[str, float]) -> pd.DataFrame:
    rows = []
    for rubric, account_prefixes in ACCOUNT_RUBRIC_MAPPING.items():
        mask = bg_accounts["account"].astype(str).str.startswith(tuple(account_prefixes))
        bg_total = float(bg_accounts.loc[mask, "balance"].sum())
        liasse_total = float(liasse_rubrics.get(rubric, 0.0))
        ecart = bg_total - liasse_total
        pct = _compute_percentage(ecart, liasse_total)
        severity = _get_severity(ecart)

        rows.append(
            {
                "Rubrique": rubric,
                "Total BG": bg_total,
                "Total Liasse": liasse_total,
                "Ecart": ecart,
                "Ecart %": pct,
                "Sévérité": severity,
            }
        )

    result = pd.DataFrame(rows)
    order = pd.CategoricalDtype(["CRITIQUE", "ATTENTION", "OK"], ordered=True)
    result["Sévérité"] = result["Sévérité"].astype(order)
    result = result.sort_values(by=["Sévérité", "Rubrique"], ascending=[True, True]).reset_index(drop=True)
    result["Sévérité"] = result["Sévérité"].astype(str)
    return result


def run_reconciliation(bg_path: str | Path, liasse_path: str | Path) -> ReconciliationBundle:
    bg_accounts = load_balance_generale(bg_path)
    liasse_rubrics = load_liasse(liasse_path)
    rec_df = reconcile(bg_accounts, liasse_rubrics)
    return ReconciliationBundle(
        bg_accounts=bg_accounts,
        liasse_rubrics=liasse_rubrics,
        reconciliation_df=rec_df,
    )


def export_reconciliation_report(
    reconciliation_df: pd.DataFrame,
    output_path: str | Path,
    bg_count: int,
    liasse_count: int,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapprochement"

    headers = ["Rubrique", "Total BG", "Total Liasse", "Ecart", "Ecart %", "Sévérité"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fill_ok = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    fill_warn = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    fill_crit = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

    for _, row in reconciliation_df.iterrows():
        ws.append(
            [
                row["Rubrique"],
                float(row["Total BG"]),
                float(row["Total Liasse"]),
                float(row["Ecart"]),
                _excel_percentage_value(row["Ecart %"]),
                row["Sévérité"],
            ]
        )
        current_row = ws.max_row
        severity = _get_severity(float(row["Ecart"]))
        if severity == "CRITIQUE":
            fill = fill_crit
        elif severity == "ATTENTION":
            fill = fill_warn
        else:
            fill = fill_ok
        for col_idx in range(1, 7):
            ws.cell(row=current_row, column=col_idx).fill = fill

    ws.column_dimensions["A"].width = 34
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16

    summary = wb.create_sheet("Résumé")
    summary["A1"] = "Statistiques"
    summary["A1"].font = Font(bold=True)
    summary_rows = [
        ("Rubriques analysées", len(reconciliation_df)),
        ("Comptes BG analysés", bg_count),
        ("Rubriques liasse détectées", liasse_count),
        ("Rubriques OK", int((reconciliation_df["Sévérité"] == "OK").sum())),
        ("Rubriques ATTENTION", int((reconciliation_df["Sévérité"] == "ATTENTION").sum())),
        ("Rubriques CRITIQUE", int((reconciliation_df["Sévérité"] == "CRITIQUE").sum())),
        ("Ecart absolu cumulé", float(reconciliation_df["Ecart"].abs().sum())),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=2):
        summary[f"A{idx}"] = label
        summary[f"B{idx}"] = value

    wb.save(output)
    LOGGER.info("Rapport exporté : %s", output)
    return output
