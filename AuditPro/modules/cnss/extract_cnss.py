#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXTRACT-CNSS V2 — Extraction des Bordereaux CNSS
Appelé par module.py (CentralisationCNSS).

Structure du bordereau CNSS (2 pages par fichier) :
  Page 1 — Régime Général :
    C1: Allocations Familiales  | Masse salariale | 6.40%  | Montant AF
    C2: Prestations Sociales    | Masse salariale | 13.46% | Montant PS
    C3: Total cotisations       |                          | AF + PS
    C4: Pénalités               |                          | 0.00
    C8: Taxe Formation Pro      | Masse salariale | 1.60%  | Montant TFP
    C9: Pénalités TFP           |                          | 0.00
    C10: Montant global         |                          | Total page 1

  Page 2 — AMO :
    C1: Participation AMO       | Masse salariale | 1.85%  | Montant
    C2: Cotisation AMO          | Masse salariale | 4.52%  | Montant
    C3: Total AMO               |                          | Part + Cot
    C4: Pénalités AMO           |                          | 0.00
    C10: Montant global AMO     |                          | Total page 2
"""

import os, re, sys, datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Optional

# ============================================================
# CONFIGURATION — surchargés par module.py avant appel
# ============================================================
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH   = r"C:\poppler\poppler-25.12.0\Library\bin"
OUTPUT_DIR     = Path(".")

def _install(p):
    os.system(f"{sys.executable} -m pip install {p} -q")

try:
    import pdfplumber
except ImportError:
    _install("pdfplumber"); import pdfplumber

try:
    import pytesseract
    from pdf2image import convert_from_path
except ImportError:
    _install("pytesseract"); _install("pdf2image")
    import pytesseract
    from pdf2image import convert_from_path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    _install("openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

MOIS = {
    1: 'Janvier',  2: 'Février',   3: 'Mars',      4: 'Avril',
    5: 'Mai',      6: 'Juin',      7: 'Juillet',   8: 'Août',
    9: 'Septembre',10: 'Octobre',  11: 'Novembre', 12: 'Décembre'
}

# ============================================================
# PARSE AMOUNT
# ============================================================
def parse_amount(s: str) -> Optional[float]:
    if not s:
        return None
    s = s.strip().replace('\u00A0', '').replace(' ', '')
    # Anglo-saxon: 86,111.10
    if re.match(r'^\d{1,3}(,\d{3})*\.\d{2}$', s):
        return float(s.replace(',', ''))
    # Européen: 86 111,10
    s2 = s.replace(',', '.')
    try:
        return float(s2)
    except Exception:
        pass
    s3 = re.sub(r'[^\d.\-]', '', s)
    try:
        return float(s3)
    except Exception:
        return None

AMOUNT_RE = re.compile(r'(\d{1,3}(?:,\d{3})*\.\d{2})')

def find_amounts(text: str) -> List[float]:
    return [parse_amount(m) for m in AMOUNT_RE.findall(text) if parse_amount(m) is not None]

# ============================================================
# DATACLASS
# ============================================================
@dataclass
class BordereauCNSS:
    fichier: str = ""
    mois_num: int = 0
    annee: int = 0
    raison_sociale: str = ""
    n_affilie: str = ""

    # Page 1 — Régime Général
    masse_salariale_af: float = 0.0
    masse_salariale_ps: float = 0.0
    af_taux: float = 0.0
    af_montant: float = 0.0
    ps_taux: float = 0.0
    ps_montant: float = 0.0
    total_cotisations: float = 0.0
    penalites_cotis: float = 0.0
    tfp_masse: float = 0.0
    tfp_taux: float = 0.0
    tfp_montant: float = 0.0
    penalites_tfp: float = 0.0
    montant_global_rg: float = 0.0

    # Page 2 — AMO
    masse_salariale_amo: float = 0.0
    participation_amo_taux: float = 0.0
    participation_amo: float = 0.0
    cotisation_amo_taux: float = 0.0
    cotisation_amo: float = 0.0
    total_amo: float = 0.0
    penalites_amo: float = 0.0
    montant_global_amo: float = 0.0

    # Totaux
    montant_total: float = 0.0

    anomalies: List[str] = field(default_factory=list)


# ============================================================
# EXTRACTION
# ============================================================
def extract_cnss_line(text: str, line_code: int) -> List[float]:
    """Extrait les montants d'une ligne codée C (1, 2, 3, 4, 8, 9, 10)."""
    lines = text.split('\n')
    for i, line in enumerate(lines):
        m = re.match(rf'^\s*{line_code}\b', line)
        if not m:
            continue
        combined = line
        if i + 1 < len(lines) and not re.match(r'^\s*\d{1,2}\b', lines[i + 1]):
            combined += " " + lines[i + 1]
        return find_amounts(combined)
    return []


def extract_page_rg(text: str, bord: BordereauCNSS):
    """Extrait les données de la page Régime Général."""
    # Période
    m = re.search(r'(\d{1,2})\s+(\d{4})\s*\n\s*VERSEMENT', text, re.I)
    if not m:
        m = re.search(r'VERSEMENT\s+DU\s+MOIS\s+DE\s*\n?\s*(\d{1,2})\s+(\d{4})', text, re.I)
    if not m:
        m = re.search(r'(\d{1,2})\s+(20\d{2})', text)
    if m:
        mois = int(m.group(1))
        annee = int(m.group(2))
        if 1 <= mois <= 12:
            bord.mois_num = mois
            bord.annee = annee

    # Raison sociale
    m = re.search(r'\b(\d{7})\s+([A-Z][A-Z\s]{3,}?)(?:\n|$)', text)
    if m:
        name = m.group(2).strip()
        if name and name not in ("REFERENCE STRUCTUREE", "N D AFFILIE"):
            bord.raison_sociale = name[:50]
    if not bord.raison_sociale:
        m = re.search(r'Raison\s+sociale\s*:\s*\n?\s*([A-Z][A-Z\s]{3,})', text, re.I)
        if m:
            bord.raison_sociale = m.group(1).strip()[:50]

    # N° affilié
    m = re.search(r'(\d{7})', text)
    if m:
        bord.n_affilie = m.group(1)

    # C1: Allocations Familiales
    amounts = extract_cnss_line(text, 1)
    if len(amounts) >= 2:
        bord.masse_salariale_af = amounts[0]
        bord.af_montant = amounts[-1]
        for line in text.split('\n'):
            if re.match(r'^\s*1\b', line) and '6' in line:
                tm = re.search(r'(\d+[.,]\d+)%', line)
                if tm:
                    bord.af_taux = float(tm.group(1).replace(',', '.'))
                break

    # C2: Prestations Sociales
    amounts = extract_cnss_line(text, 2)
    if len(amounts) >= 2:
        bord.masse_salariale_ps = amounts[0]
        bord.ps_montant = amounts[-1]

    # C3: Total cotisations
    amounts = extract_cnss_line(text, 3)
    if amounts:
        bord.total_cotisations = amounts[-1]

    # C4: Pénalités
    amounts = extract_cnss_line(text, 4)
    if amounts:
        bord.penalites_cotis = amounts[-1]

    # C8: TFP
    m = re.search(r'(\d[\d,]*\.\d{2})\s+1[.,]60%\s+(\d[\d,]*\.\d{2})', text)
    if m:
        bord.tfp_masse = parse_amount(m.group(1)) or 0
        bord.tfp_montant = parse_amount(m.group(2)) or 0
        bord.tfp_taux = 1.60

    # C9: Pénalités TFP
    amounts = extract_cnss_line(text, 9)
    if amounts:
        bord.penalites_tfp = amounts[-1]

    # C10: Montant global RG
    amounts = extract_cnss_line(text, 10)
    if amounts:
        bord.montant_global_rg = amounts[-1]


def extract_page_amo(text: str, bord: BordereauCNSS):
    """Extrait les données de la page AMO."""
    # C1: Participation AMO
    amounts = extract_cnss_line(text, 1)
    if len(amounts) >= 2:
        bord.masse_salariale_amo = amounts[0]
        bord.participation_amo = amounts[-1]
        for line in text.split('\n'):
            if re.match(r'^\s*1\b', line):
                tm = re.search(r'(\d+[.,]\d+)%', line)
                if tm:
                    bord.participation_amo_taux = float(tm.group(1).replace(',', '.'))
                break

    # C2: Cotisation AMO
    amounts = extract_cnss_line(text, 2)
    if len(amounts) >= 2:
        bord.cotisation_amo = amounts[-1]
        for line in text.split('\n'):
            if re.match(r'^\s*2\b', line):
                tm = re.search(r'(\d+[.,]\d+)%', line)
                if tm:
                    bord.cotisation_amo_taux = float(tm.group(1).replace(',', '.'))
                break

    # C3: Total AMO
    amounts = extract_cnss_line(text, 3)
    if amounts:
        bord.total_amo = amounts[-1]

    # C4: Pénalités AMO
    amounts = extract_cnss_line(text, 4)
    if amounts:
        bord.penalites_amo = amounts[-1]

    # C10: Montant global AMO
    amounts = extract_cnss_line(text, 10)
    if amounts:
        bord.montant_global_amo = amounts[-1]


def process_cnss_file(pdf_path: Path) -> Optional[BordereauCNSS]:
    """Traite un fichier PDF CNSS et retourne un BordereauCNSS."""
    bord = BordereauCNSS(fichier=pdf_path.name)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
    except Exception:
        # Fallback OCR
        try:
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
            images = convert_from_path(str(pdf_path), dpi=250, poppler_path=POPPLER_PATH)
            pages = [
                pytesseract.image_to_string(img, lang="fra+eng", config="--psm 6")
                for img in images
            ]
        except Exception as e:
            bord.anomalies.append(f"Erreur lecture PDF: {e}")
            return None

    if not pages:
        return None

    # Fichier debug
    try:
        debug_dir = Path(OUTPUT_DIR)
        debug_dir.mkdir(parents=True, exist_ok=True)
        debug = debug_dir / f"DEBUG_CNSS_{pdf_path.stem}.txt"
        with open(debug, "w", encoding="utf-8") as f:
            for i, t in enumerate(pages):
                f.write(f"\n{'='*60}\nPAGE {i+1}\n{'='*60}\n{t}\n")
    except Exception:
        pass

    for text in pages:
        if "Allocations" in text or "Régime Général" in text or "formation professionnelle" in text:
            extract_page_rg(text, bord)
        if "AMO" in text or "Assurance Maladie" in text:
            extract_page_amo(text, bord)

    bord.montant_total = bord.montant_global_rg + bord.montant_global_amo

    # Cross-checks
    if bord.af_montant > 0 and bord.ps_montant > 0:
        calc = bord.af_montant + bord.ps_montant
        if bord.total_cotisations > 0 and abs(calc - bord.total_cotisations) > 0.02:
            bord.anomalies.append(f"AF+PS={calc:.2f} ≠ C3={bord.total_cotisations:.2f}")

    if bord.masse_salariale_af > 0 and bord.af_montant > 0:
        taux_calc = bord.af_montant / bord.masse_salariale_af * 100
        if abs(taux_calc - 6.40) > 0.1:
            bord.anomalies.append(f"Taux AF calculé={taux_calc:.2f}% ≠ 6.40%")

    if not bord.mois_num:
        m = re.search(r'(\d{1,2})[-_](\d{2,4})', pdf_path.stem)
        if m:
            bord.mois_num = int(m.group(1))
            a = m.group(2)
            bord.annee = int(a) if len(a) == 4 else 2000 + int(a)
        if not bord.mois_num:
            bord.anomalies.append("Mois non trouvé")

    return bord


# ============================================================
# GÉNÉRATION EXCEL
# ============================================================
def generate_excel(declarations: List[BordereauCNSS], societe: str, output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "CNSS"

    headers = [
        "Mois", "Masse Salariale AF", "Masse Salariale PS",
        "Alloc. Familiales (6.40%)", "Prestations Sociales (13.46%)",
        "Total Cotisations", "Pénalités Cotis.",
        "TFP (1.60%)", "Pénalités TFP",
        "Total Régime Général",
        "Participation AMO (1.85%)", "Cotisation AMO (4.52%)",
        "Total AMO", "Pénalités AMO",
        "Total Global (RG+AMO)",
        "Anomalies"
    ]
    ws.append(headers)

    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name="Arial", color="FFFFFF", bold=True, size=9)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    data_by_month = {d.mois_num: d for d in declarations if d.mois_num}

    for m in range(1, 13):
        if m in data_by_month:
            d = data_by_month[m]
            ws.append([
                f"{MOIS[m]} {d.annee}",
                d.masse_salariale_af, d.masse_salariale_ps,
                d.af_montant, d.ps_montant,
                d.total_cotisations, d.penalites_cotis,
                d.tfp_montant, d.penalites_tfp,
                d.montant_global_rg,
                d.participation_amo, d.cotisation_amo,
                d.total_amo, d.penalites_amo,
                d.montant_total,
                "; ".join(d.anomalies) if d.anomalies else "OK"
            ])
        else:
            ws.append([MOIS[m]] + [""] * 15)
            rn = ws.max_row
            for c in range(2, 16):
                ws.cell(row=rn, column=c).fill = PatternFill("solid", fgColor="FFEB9C")

        rn = ws.max_row
        for c in range(2, 16):
            cell = ws.cell(row=rn, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            cell.font = Font(name="Arial", size=9)

    # Ligne TOTAL
    total_row = 14
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", size=9, bold=True)
    for c in range(2, 16):
        cl = get_column_letter(c)
        ws.cell(row=total_row, column=c).value = f"=SUM({cl}2:{cl}13)"
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(name="Arial", size=9, bold=True)
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")

    widths = {
        1: 16, 2: 18, 3: 18, 4: 22, 5: 24, 6: 18, 7: 16,
        8: 16, 9: 14, 10: 20, 11: 20, 12: 20, 13: 14, 14: 14, 15: 20, 16: 30
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    soc = re.sub(r'[\\/*?:"<>|]', '_', societe)
    out = output_dir / f"CNSS_{soc}_{ts}.xlsx"
    wb.save(out)
    return out
