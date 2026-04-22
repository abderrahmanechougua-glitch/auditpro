from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    import pdfplumber  # type: ignore
except Exception:  # pragma: no cover
    pdfplumber = None

try:
    import tabula  # type: ignore
except Exception:  # pragma: no cover
    tabula = None


ACCOUNT_RUBRIC_MAP = {
    "21": "Immobilisations incorporelles",
    "22": "Immobilisations corporelles",
    "23": "Immobilisations en cours",
    "24": "Immobilisations financières",
    "31": "Stocks",
    "32": "Stocks",
    "34": "Créances clients",
    "35": "Trésorerie actif",
    "37": "Stocks",
    "44": "Dettes fiscales et sociales",
    "45": "Autres dettes",
    "51": "Banques",
    "53": "Caisse",
    "61": "Achats",
    "62": "Charges externes",
    "63": "Charges de personnel",
    "64": "Impôts et taxes",
    "66": "Charges financières",
    "67": "Charges non courantes",
    "71": "Ventes",
    "73": "Produits d'exploitation",
    "75": "Produits financiers",
    "77": "Produits non courants",
}

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

DEFAULT_HEADER_SCAN_ROWS = 30
MIN_MEDIUM_THRESHOLD = 10.0
MEDIUM_THRESHOLD_PCT = 0.01
MIN_LARGE_THRESHOLD = 100.0
LARGE_THRESHOLD_PCT = 0.05
PDF_LINE_PATTERN = re.compile(r"^(.+?)\s+(-?\d[\d\s\.,]*)$")


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _normalize_amount(value: Any) -> float:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return 0.0
    if isinstance(value, (int, float, np.number)):
        return float(value)
    raw = str(value).strip().replace(" ", "")
    raw = raw.replace("\u202f", "").replace("\xa0", "")
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            comma_idx = raw.rfind(",")
            raw = raw[:comma_idx].replace(".", "").replace(",", "") + "." + raw[comma_idx + 1 :]
        else:
            dot_idx = raw.rfind(".")
            raw = raw[:dot_idx].replace(",", "") + "." + raw[dot_idx + 1 :].replace(",", "")
    elif raw.count(",") == 1 and raw.count(".") == 0:
        raw = raw.replace(",", ".")
    elif raw.count(",") > 1 and raw.count(".") == 0:
        raw = raw.replace(",", "")
    try:
        return float(raw)
    except Exception:
        return 0.0


def detect_header_row(
    raw_df: pd.DataFrame, expected_keywords: list[str], max_scan: int = DEFAULT_HEADER_SCAN_ROWS
) -> int:
    if raw_df.empty:
        return 0

    expected = [_normalize_text(k) for k in expected_keywords]
    best_index = 0
    best_score = -1
    for idx in range(min(max_scan, len(raw_df))):
        row_values = [_normalize_text(v) for v in raw_df.iloc[idx].tolist()]
        score = 0
        for cell in row_values:
            if not cell:
                continue
            for key in expected:
                if key and key in cell:
                    score += 1
        if score > best_score:
            best_score = score
            best_index = idx
    return best_index


def _extract_excel_dataframe(path: str, expected_keywords: list[str]) -> pd.DataFrame:
    raw_df = pd.read_excel(path, sheet_name=0, header=None, dtype=object)
    header_idx = detect_header_row(raw_df, expected_keywords)
    header = [str(v).strip() for v in raw_df.iloc[header_idx].tolist()]
    data = raw_df.iloc[header_idx + 1 :].copy().reset_index(drop=True)
    data.columns = header
    data = data.loc[:, [bool(str(c).strip()) for c in data.columns]]
    data = data.dropna(how="all")
    data.columns = [str(c).strip() for c in data.columns]
    return data


def _find_column(columns: list[str], patterns: list[str]) -> str | None:
    normalized = {_normalize_text(c): c for c in columns}
    for pattern in patterns:
        p = _normalize_text(pattern)
        for col_norm, col_real in normalized.items():
            if p in col_norm:
                return col_real
    return None


def map_account_to_rubric(account: Any) -> str:
    digits = re.sub(r"\D", "", str(account or ""))
    if not digits:
        return "Non classé"

    for prefix_len in (4, 3, 2):
        prefix = digits[:prefix_len]
        if prefix in ACCOUNT_RUBRIC_MAP:
            return ACCOUNT_RUBRIC_MAP[prefix]
    return "Non classé"


def _extract_bg(path: str) -> pd.DataFrame:
    df = _extract_excel_dataframe(path, ["compte", "intitul", "solde", "debit", "credit"])
    cols = list(df.columns)

    account_col = _find_column(cols, ["compte", "n°", "num"])
    if not account_col:
        raise ValueError("Colonne compte introuvable dans la BG")

    solde_col = _find_column(cols, ["solde"])
    sd_col = _find_column(cols, ["solde débit", "sd", "debit"])
    sc_col = _find_column(cols, ["solde crédit", "sc", "credit"])

    work = pd.DataFrame()
    work["Rubrique"] = df[account_col].map(map_account_to_rubric)

    def _series_from_col(col_name: str | None) -> pd.Series:
        if col_name and col_name in df.columns:
            return df[col_name]
        return pd.Series([0] * len(df))

    if solde_col:
        work["Montant_BG"] = df[solde_col].map(_normalize_amount)
    else:
        work["Montant_BG"] = _series_from_col(sd_col).map(_normalize_amount) - _series_from_col(sc_col).map(
            _normalize_amount
        )

    out = work.groupby("Rubrique", as_index=False)["Montant_BG"].sum()
    return out


def _extract_liasse_excel(path: str) -> pd.DataFrame:
    df = _extract_excel_dataframe(path, ["rubrique", "montant", "compte", "liasse"])
    cols = list(df.columns)

    rubric_col = _find_column(cols, ["rubrique", "poste", "libellé", "intitulé"])
    account_col = _find_column(cols, ["compte", "n°", "num"])
    amount_col = _find_column(cols, ["montant", "valeur", "solde"])

    if not amount_col:
        raise ValueError("Colonne montant introuvable dans la liasse Excel")

    work = pd.DataFrame()
    if rubric_col:
        work["Rubrique"] = df[rubric_col].astype(str).str.strip()
    elif account_col:
        work["Rubrique"] = df[account_col].map(map_account_to_rubric)
    else:
        raise ValueError("Aucune colonne Rubrique/Compte détectée dans la liasse Excel")

    work["Montant_Liasse"] = df[amount_col].map(_normalize_amount)
    work = work[work["Rubrique"].astype(str).str.strip() != ""]
    return work.groupby("Rubrique", as_index=False)["Montant_Liasse"].sum()


def _extract_liasse_pdf(path: str) -> pd.DataFrame:
    records: list[dict[str, Any]] = []

    if pdfplumber is not None:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    line_clean = re.sub(r"\s+", " ", line).strip()
                    m = PDF_LINE_PATTERN.search(line_clean)
                    if not m:
                        continue
                    rubrique = m.group(1).strip()
                    montant = _normalize_amount(m.group(2))
                    if rubrique:
                        records.append({"Rubrique": rubrique, "Montant_Liasse": montant})

    if not records and tabula is not None:
        tables = tabula.read_pdf(path, pages="all", multiple_tables=True)
        for table in tables:
            if table.empty:
                continue
            columns = [str(c) for c in table.columns]
            amount_col = _find_column(columns, ["montant", "valeur", "solde"]) or columns[-1]
            rubric_col = _find_column(columns, ["rubrique", "poste", "libellé", "intitulé"]) or columns[0]
            for _, row in table.iterrows():
                records.append(
                    {
                        "Rubrique": str(row.get(rubric_col, "")).strip(),
                        "Montant_Liasse": _normalize_amount(row.get(amount_col, 0)),
                    }
                )

    if not records:
        raise ValueError(
            "Impossible d'extraire la liasse PDF (installez pdfplumber/tabula-py ou vérifiez le format)."
        )

    df = pd.DataFrame(records)
    df = df[df["Rubrique"].astype(str).str.strip() != ""]
    return df.groupby("Rubrique", as_index=False)["Montant_Liasse"].sum()


def _extract_liasse(path: str) -> pd.DataFrame:
    ext = Path(path).suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return _extract_liasse_excel(path)
    if ext == ".pdf":
        return _extract_liasse_pdf(path)
    raise ValueError(f"Format liasse non supporté : {ext}")


def _build_report_dataframe(bg_df: pd.DataFrame, liasse_df: pd.DataFrame) -> pd.DataFrame:
    merged = pd.merge(bg_df, liasse_df, on="Rubrique", how="outer").fillna(0)
    merged["Ecart"] = merged["Montant_BG"] - merged["Montant_Liasse"]

    def _status(row: pd.Series) -> str:
        ecart = abs(float(row["Ecart"]))
        base = abs(float(row["Montant_Liasse"]))
        medium_threshold = max(MIN_MEDIUM_THRESHOLD, base * MEDIUM_THRESHOLD_PCT)
        large_threshold = max(MIN_LARGE_THRESHOLD, base * LARGE_THRESHOLD_PCT)
        if ecart <= medium_threshold:
            return "OK"
        if ecart <= large_threshold:
            return "ECART_MOYEN"
        return "ECART_SIGNIFICATIF"

    merged["Statut"] = merged.apply(_status, axis=1)
    merged = merged.sort_values(by=["Statut", "Rubrique"]).reset_index(drop=True)
    return merged


def _write_report(report_df: pd.DataFrame, output_path: Path, summary: dict[str, Any]) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="Reconciliation")
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Resume")

    wb = load_workbook(output_path)
    ws = wb["Reconciliation"]

    status_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip().lower() == "statut":
            status_col = col_idx
            break

    if status_col is not None:
        for row in range(2, ws.max_row + 1):
            value = str(ws.cell(row=row, column=status_col).value or "")
            fill = GREEN_FILL if value == "OK" else YELLOW_FILL if value == "ECART_MOYEN" else RED_FILL
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill

    wb.save(output_path)


def run_reconciliation(bg_file: str, liasse_file: str, output_dir: str | Path) -> tuple[str, dict[str, Any]]:
    bg_path = Path(bg_file)
    liasse_path = Path(liasse_file)
    if bg_path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError("Le fichier BG doit être au format Excel (.xlsx/.xls)")

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    bg_df = _extract_bg(str(bg_path))
    liasse_df = _extract_liasse(str(liasse_path))
    report_df = _build_report_dataframe(bg_df, liasse_df)

    summary = {
        "total_rubriques": int(len(report_df)),
        "ok": int((report_df["Statut"] == "OK").sum()),
        "medium": int((report_df["Statut"] == "ECART_MOYEN").sum()),
        "large": int((report_df["Statut"] == "ECART_SIGNIFICATIF").sum()),
        "total_abs_ecart": float(report_df["Ecart"].abs().sum()),
        "max_abs_ecart": float(report_df["Ecart"].abs().max()) if not report_df.empty else 0.0,
    }

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = output_path / f"reconciliation_bg_liasse_{ts}.xlsx"
    _write_report(report_df, report_file, summary)

    return str(report_file), summary
