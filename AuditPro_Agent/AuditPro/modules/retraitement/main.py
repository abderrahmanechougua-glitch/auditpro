"""
Retraitement Comptable Intelligent.

Objectif:
- Detecter automatiquement GL / BG / Balance Auxiliaire
- Nettoyer et standardiser les donnees
- Exporter un classeur .xlsx avec 3 feuilles:
  1) donnees_nettoyees
  2) resume
  3) anomalies
"""

import json
import re
import unicodedata
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd

warnings.filterwarnings("ignore")

DOC_GL = "GL"
DOC_BG = "BG"
DOC_AUX = "BALANCE_AUXILIAIRE"


def _norm_text(value: str) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _safe_str(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


BASE_VARIANTS = {
    "N°Compte": ["compte", "n°compte", "num compte", "cpte", "account", "n° cpt", "code compte"],
    "Date": ["date", "date ecriture", "date piece", "date operation", "dt"],
    "Libellé": ["libelle", "designation", "description", "intitule ecriture", "lib"],
    "Débit": ["debit", "mnt debit", "montant debit", "dt", "mvt debit"],
    "Crédit": ["credit", "mnt credit", "montant credit", "ct", "mvt credit"],
    "Intitulé": ["intitule", "libelle compte", "designation compte", "account name"],
    "Tiers": ["tiers", "fournisseur", "client", "nom", "raison sociale"],
    "Solde_Débit": ["solde debiteur", "solde debit", "sd", "solde d", "sf debit"],
    "Solde_Crédit": ["solde crediteur", "solde credit", "sc", "solde c", "sf credit"],
    "SI_Débit": ["solde initial debit", "si debit"],
    "SI_Crédit": ["solde initial credit", "si credit"],
    "Mvt_Débit": ["mouvement debit", "mvt debit"],
    "Mvt_Crédit": ["mouvement credit", "mvt credit"],
    "SF_Débit": ["solde final debit", "sf debit", "solde debit final"],
    "SF_Crédit": ["solde final credit", "sf credit", "solde credit final"],
}

SCHEMAS = {
    DOC_BG: {
        "required": ["N°Compte", "Intitulé", "Solde_Débit", "Solde_Crédit"],
        "optional": ["Date", "Libellé"],
    },
    DOC_GL: {
        "required": ["N°Compte", "Date", "Libellé", "Débit", "Crédit"],
        "optional": ["Intitulé"],
    },
    DOC_AUX: {
        "required": ["Tiers", "SF_Débit", "SF_Crédit"],
        "optional": ["SI_Débit", "SI_Crédit", "Mvt_Débit", "Mvt_Crédit"],
    },
}

TOTAL_LINE_RE = re.compile(r"\b(total|total general|report|a reporter)\b", re.IGNORECASE)
DATE_RE = re.compile(r"\b\d{1,2}[\-/]\d{1,2}[\-/]\d{2,4}\b")
COMPTE_RE = re.compile(r"^\d{4,}$")

MAPPING_MEMORY_FILE = Path(__file__).with_name("column_mapping_memory.json")


def _load_mapping_memory() -> dict:
    if not MAPPING_MEMORY_FILE.exists():
        return {}
    try:
        return json.loads(MAPPING_MEMORY_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_mapping_memory(memory: dict) -> None:
    try:
        MAPPING_MEMORY_FILE.write_text(
            json.dumps(memory, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        pass


def _header_score(cells: list[str]) -> int:
    row_norm = [_norm_text(c) for c in cells if _safe_str(c)]
    score = 0
    known = {
        _norm_text(alias)
        for aliases in BASE_VARIANTS.values()
        for alias in aliases
    }
    for val in row_norm:
        if val in known:
            score += 3
        else:
            for word in known:
                if word and word in val:
                    score += 1
                    break
    return score


def _detect_header_row(raw_df: pd.DataFrame) -> int:
    if raw_df.empty:
        return 0
    best_row = 0
    best_score = -1
    max_scan = min(30, len(raw_df))
    for idx in range(max_scan):
        values = [str(v) for v in raw_df.iloc[idx].tolist() if _safe_str(v)]
        if not values:
            continue
        score = _header_score(values)
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def _prepare_sheet(raw_df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict], int]:
    anomalies = []
    if raw_df.empty:
        return pd.DataFrame(), anomalies, 0

    # Nettoyage universel: supprimer systematiquement les 7 premieres lignes.
    start_idx = min(7, len(raw_df))
    for i in range(start_idx):
        anomalies.append({
            "Ligne_Source": i + 1,
            "Raison": "Supprimee_ligne_entete_initiale_1_7",
            "Contenu": " | ".join(_safe_str(v) for v in raw_df.iloc[i].tolist() if _safe_str(v))[:600],
        })

    work_df = raw_df.iloc[start_idx:].reset_index(drop=True)
    work_df = work_df.dropna(how="all")
    if work_df.empty:
        return pd.DataFrame(), anomalies, start_idx

    header_row = _detect_header_row(work_df)
    header_values = [_safe_str(v) for v in work_df.iloc[header_row].tolist()]
    body = work_df.iloc[header_row + 1 :].copy().reset_index(drop=True)
    body.columns = header_values

    # Nettoyer colonnes vides/inutiles.
    body = body.loc[:, [bool(_safe_str(c)) for c in body.columns]]
    body = body.dropna(axis=1, how="all")

    return body, anomalies, start_idx + header_row + 1


def _build_variants_with_memory() -> dict:
    merged = {k: list(v) for k, v in BASE_VARIANTS.items()}
    memory = _load_mapping_memory()
    for standard, aliases in memory.items():
        if standard not in merged:
            merged[standard] = []
        for alias in aliases:
            if alias not in merged[standard]:
                merged[standard].append(alias)
    return merged


def _match_standard_column(raw_col: str, variants: dict) -> tuple[str | None, float]:
    col_norm = _norm_text(raw_col)
    col_compact = col_norm.replace(" ", "")

    # 1) exact
    for standard, aliases in variants.items():
        for alias in aliases:
            alias_norm = _norm_text(alias)
            if col_norm == alias_norm or col_compact == alias_norm.replace(" ", ""):
                return standard, 1.0

    # 2) inclusion textuelle
    for standard, aliases in variants.items():
        for alias in aliases:
            alias_norm = _norm_text(alias)
            if alias_norm and alias_norm in col_norm:
                return standard, 0.7

    return None, 0.0


def _detect_doc_type(df: pd.DataFrame, filename: str) -> tuple[str | None, float, dict]:
    variants = _build_variants_with_memory()
    mapped = {}
    for col in df.columns:
        std, conf = _match_standard_column(col, variants)
        if std:
            mapped[col] = (std, conf)

    mapped_standards = {v[0] for v in mapped.values()}

    scores = {DOC_GL: 0.0, DOC_BG: 0.0, DOC_AUX: 0.0}

    # Signaux colonnes
    if "Date" in mapped_standards:
        scores[DOC_GL] += 2.0
    if "Libellé" in mapped_standards:
        scores[DOC_GL] += 2.0
    if "Débit" in mapped_standards:
        scores[DOC_GL] += 2.0
    if "Crédit" in mapped_standards:
        scores[DOC_GL] += 2.0
    if "N°Compte" in mapped_standards:
        scores[DOC_GL] += 1.0
        scores[DOC_BG] += 1.0

    if "Solde_Débit" in mapped_standards:
        scores[DOC_BG] += 2.5
    if "Solde_Crédit" in mapped_standards:
        scores[DOC_BG] += 2.5
    if "Intitulé" in mapped_standards:
        scores[DOC_BG] += 1.5

    if "Tiers" in mapped_standards:
        scores[DOC_AUX] += 3.0
    if "SF_Débit" in mapped_standards:
        scores[DOC_AUX] += 2.0
    if "SF_Crédit" in mapped_standards:
        scores[DOC_AUX] += 2.0
    if "SI_Débit" in mapped_standards or "SI_Crédit" in mapped_standards:
        scores[DOC_AUX] += 1.0
    if "Mvt_Débit" in mapped_standards or "Mvt_Crédit" in mapped_standards:
        scores[DOC_AUX] += 1.0

    # Signaux contenu (premieres lignes)
    sample = df.head(30)
    sample_text = " ".join(_safe_str(v) for row in sample.itertuples(index=False) for v in row)
    if DATE_RE.search(sample_text):
        scores[DOC_GL] += 0.8
    if re.search(r"\bfournisseur\b|\bclient\b|\btiers\b", _norm_text(sample_text)):
        scores[DOC_AUX] += 0.8

    # Nom de fichier
    fn = _norm_text(filename)
    if "grand livre" in fn or re.search(r"\bgl\b", fn):
        scores[DOC_GL] += 1.2
    if "balance generale" in fn or re.search(r"\bbg\b", fn):
        scores[DOC_BG] += 1.2
    if "aux" in fn or "auxiliaire" in fn or "fournisseur" in fn or "client" in fn:
        scores[DOC_AUX] += 1.2

    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    best_type, best_score = ranked[0]
    second_score = ranked[1][1]
    confidence = 0.0 if best_score <= 0 else min(1.0, best_score / (best_score + second_score + 0.1))

    # Si le score est trop faible ou ambigu, on considere le type non detecte.
    if best_score < 2.5 or (best_score - second_score) < 0.7:
        return None, round(confidence, 2), scores

    return best_type, round(confidence, 2), scores


def _map_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, dict, list[dict]]:
    variants = _build_variants_with_memory()
    rename_map = {}
    mapping_details = []
    used_targets = set()

    for col in df.columns:
        standard, conf = _match_standard_column(col, variants)
        if standard and standard not in used_targets:
            rename_map[col] = standard
            used_targets.add(standard)
            mapping_details.append({
                "Colonne_Source": col,
                "Colonne_Cible": standard,
                "Confiance": conf,
                "Statut": "ok" if conf >= 0.9 else "incertain",
            })
        else:
            mapping_details.append({
                "Colonne_Source": col,
                "Colonne_Cible": col,
                "Confiance": 0.0,
                "Statut": "non_mappee",
            })

    mapped = df.rename(columns=rename_map).copy()
    return mapped, rename_map, mapping_details


def _to_number(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(r"[\s\u00A0]", "", regex=True)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[€$MADdhDHS]", "", regex=True)
        .str.replace(r"[^0-9.\-]", "", regex=True)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def _log_anomaly(anomalies: list[dict], source_line: int | None, reason: str, row: pd.Series | None = None) -> None:
    content = ""
    if row is not None:
        content = " | ".join(_safe_str(v) for v in row.tolist() if _safe_str(v))[:600]
    anomalies.append({
        "Ligne_Source": source_line if source_line is not None else "",
        "Raison": reason,
        "Contenu": content,
    })


def _universal_clean(df: pd.DataFrame, source_row_offset: int, anomalies: list[dict]) -> pd.DataFrame:
    if df.empty:
        return df

    # Supprimer lignes totalement vides
    keep_mask = df.notna().any(axis=1)
    for idx in df.index[~keep_mask]:
        _log_anomaly(anomalies, source_row_offset + idx + 1, "Ligne_vide_supprimee", df.loc[idx])
    df = df[keep_mask].copy()

    # Supprimer TOTAL / TOTAL GENERAL / REPORT
    if not df.empty:
        text_series = df.astype(str).agg(" ".join, axis=1).str.lower()
        total_mask = text_series.str.contains(TOTAL_LINE_RE, regex=True, na=False)
        for idx in df.index[total_mask]:
            _log_anomaly(anomalies, source_row_offset + idx + 1, "Ligne_total_report_supprimee", df.loc[idx])
        df = df[~total_mask].copy()

    return df.reset_index(drop=True)


def _apply_schema(df: pd.DataFrame, doc_type: str, anomalies: list[dict]) -> tuple[pd.DataFrame, list[str]]:
    issues = []
    schema = SCHEMAS.get(doc_type, {"required": [], "optional": []})
    required = schema["required"]
    optional = schema["optional"]

    for col in required + optional:
        if col not in df.columns:
            df[col] = pd.NA
            if col in required:
                issues.append(f"Colonne obligatoire absente et creee vide: {col}")

    # Conversion montants
    amount_cols = [
        "Débit", "Crédit", "Solde_Débit", "Solde_Crédit",
        "SI_Débit", "SI_Crédit", "Mvt_Débit", "Mvt_Crédit", "SF_Débit", "SF_Crédit",
    ]
    for col in amount_cols:
        if col in df.columns:
            df[col] = _to_number(df[col]).fillna(0.0)

    # Date
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")

    # Regle stricte GL: N°Compte valide (>=4 chiffres) ET Libelle non vide
    if doc_type == DOC_GL:
        compte = df["N°Compte"].astype(str).str.replace(r"\s+", "", regex=True)
        libelle = df["Libellé"].astype(str).str.strip()
        valid = compte.str.match(COMPTE_RE, na=False) & libelle.ne("") & libelle.ne("nan")
        removed = df[~valid]
        for idx, row in removed.iterrows():
            _log_anomaly(anomalies, None, "GL_regle_stricte_compte_ou_libelle_invalide", row)
        df = df[valid].copy().reset_index(drop=True)
        df["N°Compte"] = df["N°Compte"].astype(str).str.replace(r"\s+", "", regex=True)

    # Validation minimale obligatoire
    for col in required:
        if col not in df.columns:
            issues.append(f"Colonne obligatoire manquante: {col}")

    return df, issues


def _period_text(df: pd.DataFrame) -> str:
    if "Date" not in df.columns:
        return "Non disponible"
    dates = pd.to_datetime(df["Date"], errors="coerce").dropna()
    if dates.empty:
        return "Non disponible"
    return f"{dates.min().date()} -> {dates.max().date()}"


def _build_resume(
    file_path: Path,
    detected_type: str | None,
    applied_type: str,
    confidence: float,
    scores: dict,
    df_in_rows: int,
    df_out: pd.DataFrame,
    mapping_details: list[dict],
    issues: list[str],
) -> pd.DataFrame:
    totals = []
    for col in ["Débit", "Crédit", "Solde_Débit", "Solde_Crédit", "SF_Débit", "SF_Crédit"]:
        if col in df_out.columns:
            totals.append((f"Total_{col}", round(float(df_out[col].sum()), 2)))

    rows = [
        ("Fichier_source", file_path.name),
        ("Type_detecte", detected_type or "NON_DETECTE"),
        ("Type_applique", applied_type),
        ("Confiance_detection", confidence),
        ("Score_GL", round(scores.get(DOC_GL, 0.0), 2)),
        ("Score_BG", round(scores.get(DOC_BG, 0.0), 2)),
        ("Score_BALANCE_AUX", round(scores.get(DOC_AUX, 0.0), 2)),
        ("Nb_lignes_entree", df_in_rows),
        ("Nb_lignes_sortie", len(df_out)),
        ("Periode", _period_text(df_out)),
        ("Colonnes_sortie", ", ".join(df_out.columns.tolist())),
        ("Mappings_incertain", sum(1 for m in mapping_details if m["Statut"] == "incertain")),
        ("Mappings_non_mappees", sum(1 for m in mapping_details if m["Statut"] == "non_mappee")),
        ("Nb_issues", len(issues)),
    ]
    rows.extend(totals)
    for i, issue in enumerate(issues, start=1):
        rows.append((f"Issue_{i}", issue))

    return pd.DataFrame(rows, columns=["Métrique", "Valeur"])


def _format_excel(ws):
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        fill = PatternFill("solid", fgColor="1F4E78")
        hdr_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(60, max(12, width + 2))

        ws.freeze_panes = "A2"
    except Exception:
        pass


def _enrich_memory(rename_map: dict) -> None:
    if not rename_map:
        return
    memory = _load_mapping_memory()
    for source, target in rename_map.items():
        memory.setdefault(target, [])
        src_norm = _norm_text(source)
        if src_norm and src_norm not in memory[target]:
            memory[target].append(src_norm)
    _save_mapping_memory(memory)


def _select_best_sheet(excel_file: pd.ExcelFile) -> tuple[pd.DataFrame, dict]:
    best_df = pd.DataFrame()
    best_meta = {"sheet": None, "confidence": -1.0, "detected": None, "scores": {}}

    for sheet in excel_file.sheet_names:
        raw = pd.read_excel(excel_file, sheet_name=sheet, header=None, dtype=str)
        prepared, _, _ = _prepare_sheet(raw)
        if prepared.empty:
            continue
        detected, confidence, scores = _detect_doc_type(prepared, sheet)
        if confidence > best_meta["confidence"]:
            best_df = prepared
            best_meta = {
                "sheet": sheet,
                "confidence": confidence,
                "detected": detected,
                "scores": scores,
            }

    return best_df, best_meta


def _process_one(file_path: Path, output_dir: Path, forced_type: str | None = None) -> dict:
    try:
        xl = pd.ExcelFile(str(file_path))
        best_df, meta = _select_best_sheet(xl)
        if best_df.empty:
            return {
                "success": False,
                "error": "Fichier vide ou aucune feuille exploitable.",
                "filename": file_path.name,
            }

        raw = pd.read_excel(str(file_path), sheet_name=meta["sheet"], header=None, dtype=str)
        prepared_df, anomalies, source_offset = _prepare_sheet(raw)

        cleaned_df = _universal_clean(prepared_df, source_offset, anomalies)
        mapped_df, rename_map, mapping_details = _map_columns(cleaned_df)

        detected_type, confidence, scores = _detect_doc_type(mapped_df, file_path.name)
        applied_type = forced_type or detected_type or DOC_GL
        issues = []

        if detected_type is None:
            issues.append(
                "Type non detecte avec certitude. Type applique par defaut: GL. "
                "Veuillez confirmer manuellement si necessaire."
            )

        mapped_uncertain = [m for m in mapping_details if m["Statut"] == "incertain"]
        if mapped_uncertain:
            issues.append(
                "Mapping incertain detecte. Mappings proposes modifiables dans la feuille resume."
            )

        final_df, schema_issues = _apply_schema(mapped_df, applied_type, anomalies)
        issues.extend(schema_issues)

        # Conserver uniquement les colonnes du schema cible + colonnes supplementaires utiles
        schema_cols = SCHEMAS[applied_type]["required"] + SCHEMAS[applied_type]["optional"]
        extra_cols = [c for c in final_df.columns if c not in schema_cols]
        ordered_cols = schema_cols + extra_cols
        final_df = final_df[ordered_cols]

        resume_df = _build_resume(
            file_path=file_path,
            detected_type=detected_type,
            applied_type=applied_type,
            confidence=confidence,
            scores=scores,
            df_in_rows=len(cleaned_df),
            df_out=final_df,
            mapping_details=mapping_details,
            issues=issues,
        )

        anomalies_df = pd.DataFrame(anomalies, columns=["Ligne_Source", "Raison", "Contenu"])
        if anomalies_df.empty:
            anomalies_df = pd.DataFrame(
                [{"Ligne_Source": "", "Raison": "Aucune_anomalie_detectee", "Contenu": ""}]
            )

        mapping_df = pd.DataFrame(mapping_details)
        if not mapping_df.empty:
            mapping_df.insert(0, "Section", "Mapping_Colonnes")
            mapping_df = mapping_df.rename(columns={
                "Colonne_Source": "Métrique",
                "Colonne_Cible": "Valeur",
            })
            mapping_notes = mapping_df[["Métrique", "Valeur"]].copy()
            notes_as_resume = pd.concat([resume_df, mapping_notes], ignore_index=True)
        else:
            notes_as_resume = resume_df

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"Retraitement_{applied_type}_{file_path.stem}_{ts}.xlsx"

        with pd.ExcelWriter(str(output_file), engine="openpyxl") as writer:
            final_df.to_excel(writer, sheet_name="données_nettoyées", index=False)
            notes_as_resume.to_excel(writer, sheet_name="résumé", index=False)
            anomalies_df.to_excel(writer, sheet_name="anomalies", index=False)

            wb = writer.book
            for ws_name in ["données_nettoyées", "résumé", "anomalies"]:
                _format_excel(wb[ws_name])

        _enrich_memory(rename_map)

        return {
            "success": True,
            "output_file": str(output_file),
            "filename": file_path.stem,
            "doc_type": applied_type,
            "detected_type": detected_type or "NON_DETECTE",
            "confidence": confidence,
            "total_lignes": len(final_df),
            "retraitements_appliques": max(0, len(cleaned_df) - len(final_df)),
            "issues": issues,
        }

    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
            "filename": file_path.name,
        }


def process_files(file_paths, output_dir, **kwargs) -> list[dict]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    forced_type = kwargs.get("forced_type")
    results = []
    for path in file_paths:
        results.append(_process_one(Path(path), out_dir, forced_type=forced_type))
    return results


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python main.py fichier1.xlsx [fichier2.xlsx ...]")
        raise SystemExit(1)

    outputs = process_files(sys.argv[1:], "output_retraitement")
    for item in outputs:
        if item.get("success"):
            print(
                f"OK {item['filename']} -> {item['doc_type']} "
                f"({item['total_lignes']} lignes) -> {item['output_file']}"
            )
        else:
            print(f"KO {item.get('filename', '?')} -> {item.get('error', 'Erreur inconnue')}")
