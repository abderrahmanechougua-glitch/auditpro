#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FACTEXT V2 — Extraction Précise de Factures Marocaines
Amélioré après analyse de 7 factures réelles :
  - AFRIQUIA (scan RICOH, TVA 10%)
  - BEST PROFIL x3 (CamScanner, format espace+point)
  - EDIC (scan, format marocain standard)
  - GestamProjects (scan)
  - FIBMED CONSULTING (scan)
  - Green Line International (format anglo-saxon MAD)
  - Engineering Conseil Solution (scan)

Corrections majeures vs V1:
  1. normalize_amount: gère 4 formats (EU, US, espace+point, espace+virgule)
  2. Patterns élargis pour couvrir toute la diversité
  3. Détection du taux TVA (10%, 14%, 20%)
  4. Extraction fournisseur améliorée
  5. Gestion des dates sur 2 chiffres (25 → 2025)
  6. Montant en lettres → cross-check
"""

import os
import re
import sys
import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Optional

# ============================================================
# CONFIGURATION — À ADAPTER À TON POSTE
# ============================================================
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH = r"C:\poppler\poppler-25.12.0\Library\bin"
BASE_DIR = Path.home() / "AuditPro_output" / "factX"

os.environ["PATH"] = os.environ["PATH"] + os.pathsep + r"C:\Program Files\Tesseract-OCR"

def install_package(package):
    os.system(f"{sys.executable} -m pip install {package}")

try:
    import pdfplumber
except ImportError:
    install_package("pdfplumber")
    import pdfplumber

try:
    import pytesseract
    from pdf2image import convert_from_path
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
except ImportError:
    install_package("pytesseract")
    install_package("pdf2image")
    import pytesseract
    from pdf2image import convert_from_path
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    install_package("openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
try:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass

# ============================================================
# DATA CLASSES
# ============================================================
@dataclass
class ExtractionResult:
    value: str = ""
    confidence: int = 0
    method: str = ""

@dataclass
class TVAControl:
    ht: Optional[float] = None
    tva_calculee: Optional[float] = None
    tva_document: Optional[float] = None
    ttc: Optional[float] = None
    taux: Optional[float] = None
    taux_detecte: Optional[float] = None  # Taux lu sur le document
    statut: str = ""
    message: str = ""

@dataclass
class Invoice:
    fichier_original: str = ""
    page_num: int = 0
    num_facture: ExtractionResult = field(default_factory=ExtractionResult)
    date_facture: ExtractionResult = field(default_factory=ExtractionResult)
    montant_ht: ExtractionResult = field(default_factory=ExtractionResult)
    montant_tva: ExtractionResult = field(default_factory=ExtractionResult)
    montant_ttc: ExtractionResult = field(default_factory=ExtractionResult)
    taux_tva: ExtractionResult = field(default_factory=ExtractionResult)
    description: ExtractionResult = field(default_factory=ExtractionResult)
    fournisseur: ExtractionResult = field(default_factory=ExtractionResult)
    ice: ExtractionResult = field(default_factory=ExtractionResult)
    if_fiscal: ExtractionResult = field(default_factory=ExtractionResult)
    tp_patente: ExtractionResult = field(default_factory=ExtractionResult)
    rc: ExtractionResult = field(default_factory=ExtractionResult)
    cnss: ExtractionResult = field(default_factory=ExtractionResult)
    devise: str = "MAD"
    montant_lettres: str = ""
    tva_control: TVAControl = field(default_factory=TVAControl)
    anomalies: List[str] = field(default_factory=list)
    score_global: int = 0

# ============================================================
# NORMALISATION DES MONTANTS — CORRIGÉE
# ============================================================

def normalize_amount(amount_str: str) -> Optional[float]:
    """
    Convertit un montant texte en float.
    Gère les 4 formats rencontrés dans les factures marocaines :
      1. Marocain standard : 175 000,00  (espace=milliers, virgule=décimal)
      2. Best Profil :       1 749 118.68 (espace=milliers, point=décimal)
      3. Anglo-saxon (MAD) : 25,000.00    (virgule=milliers, point=décimal)
      4. Simple :            24000         (pas de séparateur)
    """
    if not amount_str:
        return None

    s = amount_str.strip()
    # Nettoyer les caractères non numériques (DH, MAD, etc.)
    s = s.replace("\u00A0", " ")  # espace insécable
    s = re.sub(r"[A-Za-zÀ-ÿ()]", "", s)  # enlever lettres
    s = s.strip()

    if not s:
        return None

    # Compter les séparateurs
    nb_points = s.count(".")
    nb_virgules = s.count(",")
    nb_espaces = len(re.findall(r" ", s))

    # CAS 1: Espace + virgule décimale → 175 000,00 (Maroc standard)
    if nb_espaces >= 1 and nb_virgules == 1 and nb_points == 0:
        s = s.replace(" ", "").replace(",", ".")
    # CAS 2: Espace + point décimal → 1 749 118.68 (Best Profil)
    elif nb_espaces >= 1 and nb_points == 1 and nb_virgules == 0:
        s = s.replace(" ", "")
    # CAS 3: Espace + point milliers + virgule décimale → 1.749.118,68
    elif nb_espaces >= 0 and nb_virgules == 1 and nb_points >= 1 and s.rfind(",") > s.rfind("."):
        s = s.replace(" ", "").replace(".", "").replace(",", ".")
    # CAS 4: Virgule milliers + point décimal → 25,000.00 (anglo-saxon)
    elif nb_virgules >= 1 and nb_points == 1 and s.rfind(".") > s.rfind(","):
        s = s.replace(" ", "").replace(",", "")
    # CAS 5: Point milliers sans décimal → 1.241.572
    elif nb_points >= 2 and nb_virgules == 0:
        s = s.replace(" ", "").replace(".", "")
    # CAS 6: Virgule décimale simple → 24000,00
    elif nb_virgules == 1 and nb_points == 0 and nb_espaces == 0:
        s = s.replace(",", ".")
    # CAS 7: Tout le reste → juste nettoyer les espaces
    else:
        s = s.replace(" ", "")

    # Garder uniquement chiffres, point, moins
    s = re.sub(r"[^\d.\-]", "", s)

    try:
        val = float(s)
        return val if val >= 0 else None
    except (ValueError, TypeError):
        return None


# ============================================================
# NORMALISATION DES DATES — CORRIGÉE
# ============================================================

def normalize_date(date_str: str) -> str:
    """
    Normalise une date au format JJ/MM/AAAA.
    Gère les années sur 2 chiffres (25 → 2025).
    """
    if not date_str:
        return ""

    s = date_str.strip()

    # Format JJ/MM/AAAA ou JJ-MM-AAAA ou JJ.MM.AAAA
    m = re.match(r"(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})", s)
    if m:
        jj, mm, aaaa = m.groups()
        if len(aaaa) == 2:
            aaaa = "20" + aaaa
        return f"{int(jj):02d}/{int(mm):02d}/{aaaa}"

    # Format AAAA-MM-JJ
    m = re.match(r"(\d{4})[\/\-\.](\d{1,2})[\/\-\.](\d{1,2})", s)
    if m:
        aaaa, mm, jj = m.groups()
        return f"{int(jj):02d}/{int(mm):02d}/{aaaa}"

    return s


# ============================================================
# MONTANT EN LETTRES → NOMBRE
# ============================================================

def parse_french_words_to_number(text: str) -> Optional[float]:
    """
    Convertit un montant en lettres françaises en nombre.
    Gère les cas complexes comme :
    - "Un Million Deux Cent Quarante Et Un Mille Cinq Cent Soixante-Douze Dirham et Quatre-Vingt-Dix Centimes"
    - "DEUX MILLION QUATRE-VINGT DIX-HUIT MILLE NEUF CENT QUARANTE DEUX DIRHAMS QUARANTE DEUX CENTIMES"
    """
    if not text:
        return None

    text = text.lower().strip()

    # Séparer partie principale et centimes
    parts = re.split(r"\b(?:dirham|dirhams|dh)\b", text, maxsplit=1)
    main_text = parts[0].strip()
    centimes_text = parts[1].strip() if len(parts) > 1 else ""

    # Nettoyer les centimes de leur label
    centimes_text = re.sub(r"\b(?:centime|centimes|cts?)\b", "", centimes_text).strip()

    def words_to_int(txt: str) -> int:
        if not txt.strip():
            return 0

        UNITS = {
            'zéro': 0, 'zero': 0, 'un': 1, 'une': 1, 'deux': 2, 'trois': 3,
            'quatre': 4, 'cinq': 5, 'six': 6, 'sept': 7, 'huit': 8, 'neuf': 9,
            'dix': 10, 'onze': 11, 'douze': 12, 'treize': 13, 'quatorze': 14,
            'quinze': 15, 'seize': 16, 'dix-sept': 17, 'dix-huit': 18, 'dix-neuf': 19,
            'vingt': 20, 'trente': 30, 'quarante': 40, 'cinquante': 50,
            'soixante': 60,
        }

        # Pré-traitement des composés
        txt = txt.replace("et ", " ")
        # Normaliser les tirets
        txt = re.sub(r"[\-–]", " ", txt)

        # Traitement spécial des dizaines composées
        # "quatre vingt dix huit" → 98, "soixante douze" → 72
        tokens = txt.split()
        numbers = []
        i = 0
        while i < len(tokens):
            w = tokens[i]
            if w == "quatre" and i+1 < len(tokens) and tokens[i+1] in ("vingt", "vingts"):
                base = 80
                i += 2
                if i < len(tokens) and tokens[i] == "dix":
                    base = 90
                    i += 1
                    if i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 10:
                        base += UNITS[tokens[i]]
                        i += 1
                elif i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 20:
                    base += UNITS[tokens[i]]
                    i += 1
                numbers.append(("unit", base))
            elif w == "soixante":
                base = 60
                i += 1
                if i < len(tokens) and tokens[i] == "dix":
                    base = 70
                    i += 1
                    if i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 10:
                        base += UNITS[tokens[i]]
                        i += 1
                elif i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 20:
                    base += UNITS[tokens[i]]
                    i += 1
                numbers.append(("unit", base))
            elif w in ("cent", "cents"):
                numbers.append(("mult", 100))
                i += 1
            elif w == "mille":
                numbers.append(("mult", 1000))
                i += 1
            elif w in ("million", "millions"):
                numbers.append(("mult", 1_000_000))
                i += 1
            elif w in ("milliard", "milliards"):
                numbers.append(("mult", 1_000_000_000))
                i += 1
            elif w in UNITS:
                numbers.append(("unit", UNITS[w]))
                i += 1
            elif w.isdigit():
                numbers.append(("unit", int(w)))
                i += 1
            else:
                i += 1

        # Évaluation avec la logique standard
        total = 0
        current = 0
        for typ, val in numbers:
            if typ == "unit":
                current += val
            elif typ == "mult":
                if val == 100:
                    current = current * 100 if current > 0 else 100
                elif val == 1000:
                    current = current * 1000 if current > 0 else 1000
                    total += current
                    current = 0
                elif val >= 1_000_000:
                    current = current * val if current > 0 else val
                    total += current
                    current = 0
        total += current
        return total

    main_amount = words_to_int(main_text)
    cents_amount = words_to_int(centimes_text)

    result = main_amount + cents_amount / 100.0
    return result if result > 0 else None


# ============================================================
# PATTERNS D'EXTRACTION — V2 ÉLARGI
# ============================================================
# Chaque liste est ordonnée par spécificité décroissante.
# Le premier match gagne.

PATTERNS_NUM_FACTURE = [
    # Format exact avec label
    (r"NUM[ÉE]RO\s*:\s*([A-Z0-9][A-Z0-9\-_\/\s\.]{1,25})", 95),        # AFRIQUIA: NUMÉRO: FP10063800
    (r"N[°º]\s*FACTURE\s*:\s*([A-Z0-9][A-Z0-9\-_\/\s\.]{1,25})", 95),   # BEST PROFIL: N° FACTURE : FA253939
    (r"FACTURE\s+N[°º]\s*:?\s*([A-Z0-9][A-Z0-9\-_\/\s\.]{1,25})", 95),  # EDIC: FACTURE N° : 38
    (r"Facture\s+N[°º]\s*:?\s*([A-Z0-9][A-Z0-9\-_\/\s\.]{1,25})", 90),  # minuscule
    (r"Facture\s+n[°º]\s*:?\s*([A-Z0-9][A-Z0-9\-_\/\s\.]{1,25})", 90),  # Green Line: Facture n° ART_011225_004
    (r"Invoice\s+N[°º]\s*:?\s*([A-Z0-9][A-Z0-9\-_\/\s\.]{1,25})", 85),
    # Sans label explicite mais avec N°
    (r"N[°º]\s*:?\s*([A-Z]{0,3}\d{2,}[A-Z0-9\-_\/]*)", 80),
]

PATTERNS_DATE = [
    # Avec label explicite
    (r"DATE\s*:\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})", 95),                    # DATE : 31/03/2025
    (r"Date\s+d[''']?\s*[ée]mission\s*:\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})", 95),  # Date d'émission : 01-12-25
    (r"Date\s+d[''']?\s*[ée]ch[ée]ance\s*:\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})", 85),
    # Avec ville
    (r"(?:Rabat|Casablanca|Marrakech|Tanger|Fès|Agadir|Tétouan|Kénitra)[,\s]+le\s+(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})", 90),
    # Règlement le
    (r"R[èe]glement\s+le\s*:\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})", 80),
    # Date isolée après "le"
    (r"\ble\s+(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})", 70),
    # Date isolée (fallback)
    (r"(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{4})", 60),
]

PATTERNS_HT = [
    # Avec label explicite
    (r"Total\s+HT\s*[:\s]*(\d[\d\s\.,]+\d)", 95),                         # Total HT 175 000,00
    (r"Montant\s+(?:total\s+)?H\.?T\.?\s*[:\s]*(\d[\d\s\.,]+\d)", 95),    # Montant total H.T 25 000,00
    (r"Sous[- ]?total\s*[:\s]*(?:MAD\s*)?(\d[\d\s\.,]+\d)", 90),          # Sous-total MAD 25,000.00
    (r"H\.?T\.?\s*[:\s]+(\d[\d\s\.,]+\d)", 80),
    # BEST PROFIL: dans le tableau en bas
    (r"Montant\s+H\.T\.\s*\n?\s*(\d[\d\s\.,]+\d)", 85),
    # Avec MAD devant
    (r"(?:MAD|DH)\s+(\d[\d\s\.,]+\d)\s*$", 75),
]

PATTERNS_TVA_MONTANT = [
    # TVA montant avec label
    (r"T\.?V\.?A\.?\s*(?:\(\s*\d+\s*%?\s*\))?\s*[:\s]*(?:MAD\s*)?(\d[\d\s\.,]+\d)", 90),
    (r"TVA\s+\d+\s*%?\s*[:\s]*(\d[\d\s\.,]+\d)", 90),                    # TVA 20 % 35 000,00
    (r"T\.V\.A\.\s*[:\s]*(\d[\d\s\.,]+\d)", 85),
    (r"TVA\s*[:\s]+(\d[\d\s\.,]+\d)", 80),
]

PATTERNS_TVA_TAUX = [
    (r"TVA\s*\(?\s*(\d{1,2}(?:[.,]\d{1,2})?)\s*%\s*\)?", 95),            # TVA (20%) ou TVA 20%
    (r"T\.?V\.?A\.?\s+(\d{1,2}(?:[.,]\d{1,2})?)\s*%", 95),               # T.V.A 20%
    (r"Taux\s+T\.?V\.?A\.?\s*[:\s]*(\d{1,2}(?:[.,]\d{1,2})?)", 90),      # Taux T.V.A. 20.00
    (r"(\d{1,2})\s*%\s*$", 60),                                            # fallback: 20% en fin de ligne
]

PATTERNS_TTC = [
    # Montant en lettres (haute confiance)
    (r"Arr[êeé]t[ée]e?\s+(?:la\s+pr[ée]sente\s+facture\s+)?[àa]\s+la\s+(?:somme|valeur)\s+de\s*:?\s*(.+?)(?:\.|$)", 98),
    # Labels explicites
    (r"Total\s+G[ée]n[ée]ral\s*[:\s]*(\d[\d\s\.,]+\d)", 95),             # AFRIQUIA: Total Général
    (r"Total\s+T\.?T\.?C\.?\s*[:\s]*(\d[\d\s\.,]+\d)", 95),              # Total TTC
    (r"Montant\s+(?:total\s+)?T\.?T\.?C\.?\s*[:\s]*(\d[\d\s\.,]+\d)", 95),  # Montant total T.T.C 30 000,00
    (r"Net\s+[àa]\s+payer\s*[:\s]*(?:MAD\s*)?(\d[\d\s\.,]+\d)", 90),
    # Montant suivi de DH/MAD
    (r"(\d[\d\s\.,]+\d)\s*(?:DH|DIRHAMS?|MAD)\s*$", 75),
    # MAD devant le montant
    (r"MAD\s+(\d[\d\s\.,]+\d)\s*$", 75),
]

PATTERNS_FOURNISSEUR = [
    # Labels explicites
    (r"(?:Soci[ée]t[ée]|Sté|Entreprise|Raison\s+sociale)\s*:?\s*([A-ZÀ-ÿ][A-Za-zÀ-ÿ\s\-&\.]{3,60})", 90),
    # Nom en début de document (tout en majuscules, au moins 2 mots)
    (r"^([A-Z][A-Z\s\-&\.]{4,50})\s*$", 70),
]

PATTERNS_ICE = [
    (r"ICE\s*:?\s*(\d{15})", 95),
    (r"I\.?C\.?E\.?\s*:?\s*(\d{15})", 90),
    (r"Identifiant\s+commun\s+(?:de\s+l['''])?entreprise\s*(?:\(ICE\))?\s*:?\s*(\d{15})", 90),
]

PATTERNS_IF = [
    (r"I\.?F\.?\s*:?\s*(\d{6,10})", 90),
    (r"Identifiant\s+fiscal\s*:?\s*(\d{6,10})", 90),
]

PATTERNS_TP = [
    (r"(?:PATENTE|TP|Patente|T\.?P\.?)\s*:?\s*(\d{6,15})", 90),
]

PATTERNS_RC = [
    (r"R\.?C\.?\s*:?\s*(\d{3,10})", 90),
    (r"(?:Registre\s+de\s+Commerce|N[°o]\s+de\s+RC)\s*:?\s*(\d{3,10})", 90),
]

PATTERNS_CNSS = [
    (r"CNSS\s*:?\s*(\d{6,15})", 90),
    (r"N[°o]\s+CNSS\s*:?\s*(\d{6,15})", 90),
]

PATTERNS_DESCRIPTION = [
    (r"(?:Objet|D[ée]signation|Description)\s*:?\s*([^\n]{10,200})", 85),
    (r"Projet\s*:?\s*([^\n]{10,200})", 80),
    (r"Mission\s+([^\n]{10,150})", 75),
    (r"Prestation\s+(?:de\s+)?([^\n]{10,150})", 70),
]


# ============================================================
# MOTEUR D'EXTRACTION
# ============================================================

def extract_first_match(text: str, patterns: list) -> ExtractionResult:
    """Essaie chaque pattern dans l'ordre, retourne le premier match."""
    for pattern, confidence in patterns:
        m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if m:
            value = m.group(1).strip()
            # Nettoyer les espaces de fin, les pipes, etc.
            value = re.sub(r"[\|\s]+$", "", value).strip()
            if value:
                return ExtractionResult(value=value, confidence=confidence, method=pattern[:40])
    return ExtractionResult()


def extract_fournisseur(text: str) -> ExtractionResult:
    """Extraction du fournisseur avec logique spéciale."""
    # 1. Essayer les patterns explicites
    result = extract_first_match(text, PATTERNS_FOURNISSEUR)
    if result.value:
        return result

    # 2. Chercher un nom en majuscules dans les 10 premières lignes
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    STOP_WORDS = {"FACTURE", "INVOICE", "DATE", "TOTAL", "N°", "CLIENT", "DESTINATAIRE",
                  "LIVRÉ", "FACTURÉ", "TVA", "HT", "TTC", "COMPTABILISÉ", "COMPTABILISE",
                  "PAGE", "DESIGNATION", "MONTANT", "PROJET", "OBJET", "PAYÉ", "PAYE"}

    for line in lines[:10]:
        # Ligne principalement en majuscules, pas un label
        if len(line) < 4 or len(line) > 80:
            continue
        upper_ratio = sum(1 for c in line if c.isupper()) / max(len(line), 1)
        if upper_ratio < 0.5:
            continue
        words = line.upper().split()
        if any(w in STOP_WORDS for w in words):
            continue
        # Vérifier que ce n'est pas un numéro
        if re.match(r"^\d+$", line):
            continue
        return ExtractionResult(value=line[:60], confidence=50, method="header_line")

    return ExtractionResult()


def extract_ttc_with_words(text: str) -> Tuple[Optional[float], str]:
    """Extrait le TTC depuis le montant en lettres (cross-check)."""
    # Pattern : "Arrêté à la somme de : ..."
    m = re.search(
        r"Arr[êeé]t[ée]e?\s+(?:la\s+pr[ée]sente\s+facture\s+)?[àa]\s+la\s+(?:somme|valeur)\s+de\s*:?\s*(.+?)(?:\.|\n|$)",
        text, re.IGNORECASE
    )
    if m:
        words_text = m.group(1).strip()
        num = parse_french_words_to_number(words_text)
        if num and num > 0:
            return num, words_text

    # Variante : "Arrêté à la valeur de:" (sur la même ligne)
    m = re.search(
        r"Arr[êeé]t[ée]e?\s+[àa]\s+la\s+valeur\s+de\s*:?\s*(.+?)(?:DH|DIRHAMS?|MAD|\.|\n|$)",
        text, re.IGNORECASE
    )
    if m:
        words_text = m.group(1).strip()
        num = parse_french_words_to_number(words_text)
        if num and num > 0:
            return num, words_text

    return None, ""


def extract_amounts_context(text: str) -> dict:
    """
    Extraction contextuelle des montants.
    Cherche les montants dans un tableau structuré si possible.
    """
    amounts = {}

    # Chercher un bloc tableau avec HT / TVA / TTC sur des lignes consécutives
    # Pattern typique :
    #   Total HT        175 000,00
    #   TVA 20%          35 000,00
    #   Total TTC       210 000,00
    table_pattern = re.compile(
        r"(?:Total|Montant|Sous[- ]?total)\s+H\.?T\.?\s*[:\s]*(\d[\d\s\.,]+\d)"
        r".*?"
        r"T\.?V\.?A\.?\s*(?:\(?\s*\d+\s*%?\s*\)?)?\s*[:\s]*(\d[\d\s\.,]+\d)"
        r".*?"
        r"(?:Total|Montant)\s+(?:G[ée]n[ée]ral|T\.?T\.?C\.?)\s*[:\s]*(\d[\d\s\.,]+\d)",
        re.IGNORECASE | re.DOTALL
    )
    m = table_pattern.search(text)
    if m:
        amounts["ht_table"] = m.group(1).strip()
        amounts["tva_table"] = m.group(2).strip()
        amounts["ttc_table"] = m.group(3).strip()

    return amounts


def extract_all_fields(text: str, extract_juridical: bool = True) -> Invoice:
    """Extraction complète de tous les champs d'une facture."""
    inv = Invoice()

    # === NUMÉRO DE FACTURE ===
    inv.num_facture = extract_first_match(text, PATTERNS_NUM_FACTURE)

    # === DATE ===
    date_result = extract_first_match(text, PATTERNS_DATE)
    if date_result.value:
        date_result.value = normalize_date(date_result.value)
    inv.date_facture = date_result

    # === MONTANTS : essayer d'abord le bloc tableau ===
    table_amounts = extract_amounts_context(text)

    # HT
    if "ht_table" in table_amounts:
        inv.montant_ht = ExtractionResult(value=table_amounts["ht_table"], confidence=95, method="table_block")
    else:
        inv.montant_ht = extract_first_match(text, PATTERNS_HT)

    # TVA montant
    if "tva_table" in table_amounts:
        inv.montant_tva = ExtractionResult(value=table_amounts["tva_table"], confidence=95, method="table_block")
    else:
        inv.montant_tva = extract_first_match(text, PATTERNS_TVA_MONTANT)

    # TVA taux
    inv.taux_tva = extract_first_match(text, PATTERNS_TVA_TAUX)

    # TTC
    if "ttc_table" in table_amounts:
        inv.montant_ttc = ExtractionResult(value=table_amounts["ttc_table"], confidence=95, method="table_block")
    else:
        inv.montant_ttc = extract_first_match(text, PATTERNS_TTC)

    # Si le TTC match est le pattern "montant en lettres", traiter séparément
    if inv.montant_ttc.confidence == 98:
        # C'est le pattern en lettres, essayer de convertir
        words_text = inv.montant_ttc.value
        num = parse_french_words_to_number(words_text)
        if num:
            inv.montant_lettres = words_text
            inv.montant_ttc = ExtractionResult(value=f"{num:.2f}", confidence=95, method="words_to_number")
        else:
            # Le montant en lettres n'a pas pu être converti, chercher un pattern numérique
            inv.montant_lettres = words_text
            inv.montant_ttc = extract_first_match(text, PATTERNS_TTC[1:])  # Skip le pattern lettres

    # Cross-check avec montant en lettres
    ttc_words, words_text = extract_ttc_with_words(text)
    if ttc_words and not inv.montant_lettres:
        inv.montant_lettres = words_text

    # === DESCRIPTION ===
    inv.description = extract_first_match(text, PATTERNS_DESCRIPTION)

    # === FOURNISSEUR ===
    inv.fournisseur = extract_fournisseur(text)

    # === DONNÉES JURIDIQUES ===
    if extract_juridical:
        inv.ice = extract_first_match(text, PATTERNS_ICE)
        inv.if_fiscal = extract_first_match(text, PATTERNS_IF)
        inv.tp_patente = extract_first_match(text, PATTERNS_TP)
        inv.rc = extract_first_match(text, PATTERNS_RC)
        inv.cnss = extract_first_match(text, PATTERNS_CNSS)

    # === CONTRÔLE TVA ===
    ht_val = normalize_amount(inv.montant_ht.value)
    ttc_val = normalize_amount(inv.montant_ttc.value)
    tva_val = normalize_amount(inv.montant_tva.value)
    taux_str = inv.taux_tva.value.replace(",", ".") if inv.taux_tva.value else None
    taux_doc = float(taux_str) if taux_str else None

    inv.tva_control = TVAControl(ht=ht_val, ttc=ttc_val, tva_document=tva_val, taux_detecte=taux_doc)

    if ht_val and ttc_val and ht_val > 0:
        inv.tva_control.tva_calculee = round(ttc_val - ht_val, 2)
        inv.tva_control.taux = round((inv.tva_control.tva_calculee / ht_val) * 100, 2)

        # Vérifier la cohérence HT + TVA = TTC
        if tva_val is not None:
            ecart_tva = abs(inv.tva_control.tva_calculee - tva_val)
            somme_check = abs((ht_val + tva_val) - ttc_val) if ttc_val else 999
            if ecart_tva <= 1.0 and somme_check <= 1.0:
                inv.tva_control.statut = "OK"
                inv.tva_control.message = f"HT({ht_val:.2f}) + TVA({tva_val:.2f}) = TTC({ttc_val:.2f})"
            else:
                inv.tva_control.statut = "ANOMALIE"
                inv.tva_control.message = f"Écart TVA: {ecart_tva:.2f} | Écart somme: {somme_check:.2f}"
        else:
            # TVA non trouvée mais HT et TTC OK
            inv.tva_control.statut = "TVA_MANQUANTE"
            inv.tva_control.message = f"TVA calculée: {inv.tva_control.tva_calculee:.2f} (taux ~{inv.tva_control.taux:.1f}%)"

        # Vérifier cohérence du taux
        if taux_doc and inv.tva_control.taux:
            ecart_taux = abs(inv.tva_control.taux - taux_doc)
            if ecart_taux > 1.0:
                inv.tva_control.statut = "TAUX_ANOMALIE"
                inv.tva_control.message += f" | Taux doc: {taux_doc}%, Taux calculé: {inv.tva_control.taux:.1f}%"

    elif ht_val and tva_val and not ttc_val:
        inv.tva_control.ttc = round(ht_val + tva_val, 2)
        inv.tva_control.tva_calculee = tva_val
        if ht_val > 0:
            inv.tva_control.taux = round((tva_val / ht_val) * 100, 2)
        inv.tva_control.statut = "TTC_CALCULE"
        inv.tva_control.message = f"TTC calculé: {inv.tva_control.ttc:.2f}"
        inv.montant_ttc = ExtractionResult(value=f"{inv.tva_control.ttc:.2f}", confidence=70, method="calculated")
    else:
        inv.tva_control.statut = "INCOMPLET"
        missing = []
        if not ht_val: missing.append("HT")
        if not ttc_val: missing.append("TTC")
        if not tva_val: missing.append("TVA")
        inv.tva_control.message = f"Manquant: {', '.join(missing)}"

    # Cross-check montant en lettres vs TTC numérique
    if ttc_words and ttc_val:
        ecart_lettres = abs(ttc_words - ttc_val)
        if ecart_lettres > 1.0:
            inv.anomalies.append(f"Lettres ({ttc_words:.2f}) ≠ TTC ({ttc_val:.2f})")

    # === ANOMALIES ===
    if not inv.num_facture.value:
        inv.anomalies.append("N° facture manquant")
    if not inv.date_facture.value:
        inv.anomalies.append("Date manquante")
    if not inv.montant_ttc.value:
        inv.anomalies.append("TTC manquant")
    if not inv.montant_ht.value:
        inv.anomalies.append("HT manquant")
    if not inv.fournisseur.value:
        inv.anomalies.append("Fournisseur non identifié")
    if extract_juridical and not inv.ice.value:
        inv.anomalies.append("ICE manquant")

    # === SCORE GLOBAL ===
    fields_to_score = [inv.num_facture, inv.date_facture, inv.montant_ht, inv.montant_tva, inv.montant_ttc]
    scores = [f.confidence for f in fields_to_score if f.value]
    inv.score_global = sum(scores) // max(len(scores), 1) if scores else 0

    return inv


# ============================================================
# TRAITEMENT PDF
# ============================================================

def get_pdf_text(pdf_path: Path) -> Tuple[int, List[str]]:
    # Méthode 1 : extraction native pdfplumber (pas besoin de Poppler/Tesseract)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        if any(len(t.strip()) > 50 for t in pages_text):
            print(f"   ✅ {len(pages_text)} page(s) extraites (PDF natif)")
            return len(pages_text), pages_text
    except Exception as e:
        print(f"   ⚠️ pdfplumber: {e}")

    # Méthode 2 : OCR via Poppler + Tesseract (PDF scanné)
    try:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        images = convert_from_path(str(pdf_path), dpi=250, poppler_path=POPPLER_PATH)
        pages_text = []
        for i, img in enumerate(images):
            print(f"      OCR page {i+1}/{len(images)}...", end="\r")
            text = pytesseract.image_to_string(img, lang="fra+eng", config="--psm 6 --oem 3")
            pages_text.append(text)
        print(f"   ✅ {len(images)} page(s) traitées (OCR)" + " " * 20)
        return len(images), pages_text
    except Exception as e:
        print(f"   ❌ Erreur OCR: {e}")
        return 0, []


def detect_invoice_boundaries(pages_text: List[str]) -> List[Tuple[int, int]]:
    """
    Détecte les limites de chaque facture dans un PDF multi-pages.
    Retourne une liste de (page_debut, page_fin) en 0-indexed.
    """
    n = len(pages_text)
    if n == 0:
        return []
    if n == 1:
        return [(0, 0)]

    # Indicateurs de début de facture
    starts = []
    for i in range(n):
        text = pages_text[i]
        has_facture = bool(re.search(r"\b(FACTURE|Invoice)\b", text, re.I))
        has_num = bool(re.search(r"(?:N[°º]\s*(?:FACTURE|Client)|NUM[ÉE]RO|Facture\s+N)", text, re.I))
        has_new_total = bool(re.search(r"(?:Total\s+H\.?T|Montant\s+H\.?T)", text, re.I))

        if i == 0:
            starts.append(i)
        elif has_facture and has_num:
            starts.append(i)
        elif has_facture and has_new_total and i > 0:
            # Vérifier que la page précédente avait aussi un total (= fin de facture précédente)
            prev_has_ttc = bool(re.search(r"(?:T\.?T\.?C|Net\s+[àa]\s+payer|Total\s+G)", pages_text[i-1], re.I))
            if prev_has_ttc:
                starts.append(i)

    # Si aucune segmentation trouvée, chaque page = 1 facture
    if len(starts) <= 1:
        # Vérifier si chaque page a sa propre structure de facture
        all_have_structure = all(
            re.search(r"(?:N[°º]\s*FACTURE|Total|TTC|Montant)", pages_text[i], re.I)
            for i in range(n)
        )
        if all_have_structure and n > 1:
            return [(i, i) for i in range(n)]
        else:
            return [(0, n - 1)]

    # Construire les intervalles
    boundaries = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else n - 1
        boundaries.append((start, end))

    return boundaries


def process_pdf(pdf_path: Path, extract_juridical: bool) -> List[Invoice]:
    print(f"\n📄 {pdf_path.name}")
    pages_count, pages_text = get_pdf_text(pdf_path)
    if pages_count == 0:
        return []

    # Debug : sauvegarder le texte extrait
    try:
        Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
        debug_path = Path(OUTPUT_DIR) / f"DEBUG_OCR_{pdf_path.stem}.txt"
        with open(debug_path, "w", encoding="utf-8") as f:
            for i, t in enumerate(pages_text):
                f.write(f"\n{'='*60}\nPAGE {i+1}\n{'='*60}\n{t}\n")
    except Exception:
        pass

    boundaries = detect_invoice_boundaries(pages_text)
    print(f"   🔍 {len(boundaries)} facture(s) détectée(s)")

    invoices = []
    for idx, (start, end) in enumerate(boundaries):
        combined_text = "\n\n".join(pages_text[start:end + 1])

        inv = extract_all_fields(combined_text, extract_juridical)
        inv.fichier_original = pdf_path.name
        inv.page_num = start + 1

        num = inv.num_facture.value or "?"
        ttc = inv.montant_ttc.value or "?"
        date = inv.date_facture.value or "?"
        four = inv.fournisseur.value or "?"
        status = inv.tva_control.statut
        print(f"      #{idx+1} | P.{start+1} | {four[:20]} | N°{num} | TTC:{ttc} | {date} | TVA:{status}")

        if inv.anomalies:
            print(f"         ⚠️ {', '.join(inv.anomalies)}")

        invoices.append(inv)

    return invoices


# ============================================================
# GÉNÉRATION EXCEL — AMÉLIORÉE
# ============================================================

def generate_excel(invoices: List[Invoice], output_path: Path, extract_juridical: bool):
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"

    # En-têtes
    headers = [
        "Fichier", "Page", "Fournisseur", "N° Facture", "Date",
        "Montant HT", "Taux TVA", "Montant TVA", "Montant TTC",
        "TVA Calculée", "Taux Calculé", "Statut TVA", "Message TVA",
        "Montant Lettres", "Description",
        "Score %", "Anomalies"
    ]
    if extract_juridical:
        headers.extend(["ICE", "IF", "TP/Patente", "RC", "CNSS"])

    ws.append(headers)

    # Style en-tête
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Couleurs de statut
    FILL_OK = PatternFill("solid", fgColor="C6EFCE")
    FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
    FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")
    FILL_GREY = PatternFill("solid", fgColor="D9D9D9")
    FONT_DATA = Font(name="Arial", size=9)

    for inv in invoices:
        row_data = [
            inv.fichier_original,
            inv.page_num,
            inv.fournisseur.value,
            inv.num_facture.value,
            inv.date_facture.value,
            normalize_amount(inv.montant_ht.value) if inv.montant_ht.value else None,
            inv.taux_tva.value + "%" if inv.taux_tva.value else "",
            normalize_amount(inv.montant_tva.value) if inv.montant_tva.value else None,
            normalize_amount(inv.montant_ttc.value) if inv.montant_ttc.value else None,
            inv.tva_control.tva_calculee,
            f"{inv.tva_control.taux:.2f}%" if inv.tva_control.taux else "",
            inv.tva_control.statut,
            inv.tva_control.message,
            inv.montant_lettres[:80] if inv.montant_lettres else "",
            inv.description.value[:100] if inv.description.value else "",
            f"{inv.score_global}%",
            "; ".join(inv.anomalies) if inv.anomalies else "OK"
        ]
        if extract_juridical:
            row_data.extend([
                inv.ice.value, inv.if_fiscal.value, inv.tp_patente.value,
                inv.rc.value, inv.cnss.value
            ])

        ws.append(row_data)
        row_num = ws.max_row

        # Formater les montants
        for col_idx in [6, 8, 9, 10]:  # HT, TVA, TTC, TVA calculée
            cell = ws.cell(row=row_num, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

        # Couleur statut TVA
        status_cell = ws.cell(row=row_num, column=12)
        if inv.tva_control.statut == "OK":
            status_cell.fill = FILL_OK
        elif inv.tva_control.statut in ("TVA_MANQUANTE", "TTC_CALCULE"):
            status_cell.fill = FILL_WARN
        elif "ANOMALIE" in inv.tva_control.statut:
            status_cell.fill = FILL_ERROR
        else:
            status_cell.fill = FILL_GREY

        # Couleur anomalies
        anomaly_cell = ws.cell(row=row_num, column=17)
        if inv.anomalies:
            anomaly_cell.fill = FILL_WARN
            anomaly_cell.font = Font(name="Arial", size=9, color="9C5700")

        # Police de base
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            if not cell.font or cell.font == Font():
                cell.font = FONT_DATA
            cell.border = thin_border

    # Auto-largeur des colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Filtre automatique
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\n✅ EXCEL → {output_path}")


# ============================================================
# MENU INTERACTIF
# ============================================================

def show_menu():
    print("\n" + "=" * 70)
    print("  FACTEXT V2 — Extraction Précise de Factures Marocaines")
    print("  Supporte : AFRIQUIA, BEST PROFIL, EDIC, GestamProjects,")
    print("             FIBMED, Green Line, Engineering Conseil, etc.")
    print("=" * 70)

    print("\n📊 Données toujours extraites :")
    print("   ✓ N° facture, Date, Fournisseur, Description")
    print("   ✓ HT, TVA (taux + montant), TTC")
    print("   ✓ Contrôle TVA (HT+TVA=TTC, taux cohérent)")
    print("   ✓ Montant en lettres (cross-check)")

    print("\n⚖️ Données juridiques (optionnelles) :")
    print("   • ICE, IF, TP/Patente, RC, CNSS")

    while True:
        choix = input("\nExtraire les données juridiques ? (O/N) [O]: ").strip().upper()
        if choix in ('O', 'N', ''):
            return choix != 'N'
        print("❌ Répondez O ou N")


def main():
    extract_juridical = show_menu()

    pdf_files = sorted(INPUT_DIR.glob("*.pdf"))
    if not pdf_files:
        print(f"\n❌ Aucun PDF dans {INPUT_DIR}")
        print(f"   Placez vos factures PDF dans : {INPUT_DIR}")
        input("\nAppuyez sur Entrée...")
        return

    print(f"\n📁 {len(pdf_files)} fichier(s) PDF :")
    for i, pdf in enumerate(pdf_files, 1):
        size_mb = pdf.stat().st_size / 1024 / 1024
        print(f"   {i}. {pdf.name} ({size_mb:.1f} MB)")

    print(f"\nOptions : T=Tous, ou numéros séparés par virgule (ex: 1,3)")
    choix = input("Votre choix [T]: ").strip().upper()

    if choix and choix != 'T':
        selected = []
        for c in choix.split(','):
            c = c.strip()
            if c.isdigit() and 1 <= int(c) <= len(pdf_files):
                selected.append(pdf_files[int(c) - 1])
        if selected:
            pdf_files = selected

    all_invoices = []
    for pdf in pdf_files:
        invoices = process_pdf(pdf, extract_juridical)
        all_invoices.extend(invoices)

    if all_invoices:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = "Complet" if extract_juridical else "Base"
        output_path = OUTPUT_DIR / f"Factures_{suffix}_{timestamp}.xlsx"
        generate_excel(all_invoices, output_path, extract_juridical)

        # Résumé
        print(f"\n{'='*50}")
        print(f"📊 RÉSUMÉ")
        print(f"{'='*50}")
        print(f"   Factures traitées : {len(all_invoices)}")
        print(f"   Avec N° facture   : {sum(1 for i in all_invoices if i.num_facture.value)}/{len(all_invoices)}")
        print(f"   Avec Date         : {sum(1 for i in all_invoices if i.date_facture.value)}/{len(all_invoices)}")
        print(f"   Avec HT           : {sum(1 for i in all_invoices if i.montant_ht.value)}/{len(all_invoices)}")
        print(f"   Avec TTC          : {sum(1 for i in all_invoices if i.montant_ttc.value)}/{len(all_invoices)}")
        print(f"   Avec Fournisseur  : {sum(1 for i in all_invoices if i.fournisseur.value)}/{len(all_invoices)}")
        print(f"   TVA OK            : {sum(1 for i in all_invoices if i.tva_control.statut == 'OK')}/{len(all_invoices)}")
        print(f"   TVA ANOMALIE      : {sum(1 for i in all_invoices if 'ANOMALIE' in i.tva_control.statut)}/{len(all_invoices)}")
        print(f"   Score moyen       : {sum(i.score_global for i in all_invoices) // len(all_invoices)}%")
        print(f"\n   Fichier Excel     : {output_path.name}")
        print(f"   Debug OCR         : {OUTPUT_DIR}/DEBUG_OCR_*.txt")
    else:
        print("\n❌ Aucune facture extraite.")

    input("\n✅ Terminé ! Appuyez sur Entrée...")


if __name__ == "__main__":
    main()