"""
Détection intelligente du type de fichier.
Analyse les colonnes et noms d'onglets pour proposer le bon module.
"""
import pandas as pd
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


class FileDetector:
    """Analyse un fichier Excel et retourne le module le plus pertinent."""

    def __init__(self):
        self._profile_cache: dict[tuple[str, int, int], dict[str, Any]] = {}

    def _file_signature(self, path: Path) -> tuple[str, int, int]:
        stat = path.stat()
        return (str(path.resolve()), stat.st_size, stat.st_mtime_ns)

    def _estimate_excel_rows(self, file_path: str, suffix: str, sheet_names: list[str]) -> int:
        """Estime rapidement le nombre de lignes sans charger toutes les feuilles en DataFrame."""
        if load_workbook is None or suffix not in (".xlsx", ".xlsm"):
            return 0

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            total = 0
            for sheet in sheet_names[:3]:
                if sheet in wb.sheetnames:
                    total += max(0, int(wb[sheet].max_row or 0))
            wb.close()
            return total
        except Exception:
            return 0

    def _build_profile(self, file_path: str) -> dict[str, Any]:
        path = Path(file_path)
        suffix = path.suffix.lower()

        profile: dict[str, Any] = {
            "name": path.name,
            "size_kb": round(path.stat().st_size / 1024, 1),
            "extension": suffix,
            "sheets": [],
            "total_rows": 0,
            "all_columns": set(),
            "all_sheet_names": set(),
            "all_values": set(),
            "error": None,
        }

        if suffix == ".csv":
            try:
                df = pd.read_csv(file_path, nrows=80, encoding="utf-8", errors="replace")
                profile["sheets"] = ["CSV"]
                profile["all_sheet_names"] = {"csv"}
                profile["all_columns"] = {str(col).lower().strip() for col in df.columns}
                sample_values = set()
                for col in df.columns:
                    for val in df[col].dropna().head(12):
                        sample_values.add(str(val).lower().strip())
                profile["all_values"] = sample_values
            except Exception as exc:
                profile["error"] = str(exc)
            return profile

        try:
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
            profile["sheets"] = sheet_names
            profile["all_sheet_names"] = {s.lower().strip() for s in sheet_names}

            for sheet in sheet_names[:5]:
                try:
                    df = pd.read_excel(xls, sheet_name=sheet, nrows=80)
                except Exception:
                    continue

                for col in df.columns:
                    profile["all_columns"].add(str(col).lower().strip())

                for col in df.columns:
                    for val in df[col].dropna().head(12):
                        profile["all_values"].add(str(val).lower().strip())

            profile["total_rows"] = self._estimate_excel_rows(file_path, suffix, sheet_names)
        except Exception as exc:
            profile["error"] = str(exc)

        return profile

    def _get_profile(self, file_path: str) -> dict[str, Any]:
        path = Path(file_path)
        try:
            sig = self._file_signature(path)
        except Exception as exc:
            return {
                "name": path.name,
                "size_kb": 0,
                "extension": path.suffix.lower(),
                "sheets": [],
                "total_rows": 0,
                "all_columns": set(),
                "all_sheet_names": set(),
                "all_values": set(),
                "error": str(exc),
            }

        cached = self._profile_cache.get(sig)
        if cached is not None:
            return cached

        profile = self._build_profile(file_path)
        # Conserver un cache compact pour éviter une croissance indéfinie.
        if len(self._profile_cache) >= 32:
            self._profile_cache.clear()
        self._profile_cache[sig] = profile
        return profile

    def detect(self, file_path: str, modules: dict) -> list[dict]:
        """
        Analyse un fichier et score chaque module.
        
        Retourne une liste triée :
        [{"module": instance, "score": 0.85, "matched_keywords": [...], "columns": [...]}]
        """
        profile = self._get_profile(file_path)
        if profile.get("error"):
            return [{"module": None, "score": 0, "error": profile["error"]}]

        all_columns = profile["all_columns"]
        all_sheet_names = profile["all_sheet_names"]
        all_values = profile["all_values"]

        # Scorer chaque module
        results = []
        searchable = all_columns | all_sheet_names | all_values

        for mod in modules.values():
            keywords = getattr(mod, 'detection_keywords', [])
            threshold = getattr(mod, 'detection_threshold', 0.5)

            if not keywords:
                continue

            matched = []
            for kw in keywords:
                kw_lower = kw.lower()
                if any(kw_lower in item for item in searchable):
                    matched.append(kw)

            score = len(matched) / len(keywords) if keywords else 0

            if score >= threshold:
                results.append({
                    "module": mod,
                    "score": round(score, 2),
                    "matched_keywords": matched,
                    "columns": list(all_columns),
                    "sheets": list(all_sheet_names),
                })

        # Trier par score décroissant
        results.sort(key=lambda x: x["score"], reverse=True)
        return results

    def get_file_info(self, file_path: str) -> dict:
        """Retourne des métadonnées basiques sur un fichier."""
        profile = self._get_profile(file_path)
        return {
            "name": profile.get("name", Path(file_path).name),
            "size_kb": profile.get("size_kb", 0),
            "extension": profile.get("extension", Path(file_path).suffix.lower()),
            "sheets": profile.get("sheets", []),
            "total_rows": profile.get("total_rows", 0),
        }
