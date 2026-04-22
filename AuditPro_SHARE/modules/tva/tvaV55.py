#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXTRACT-TVA V3 — Extraction Déclarations Mensuelles TVA (DGI Maroc)

CORRECTIONS vs V11 :
  1. parse_amount universel : gère BOTH formats DGI
     - Anglo-saxon (Brouillon) : 1,834,791.03
     - Européen   (Déposée)   : 6 516 944,20
  2. Extraction par numéro de ligne (pas par regex textuel fragile)
  3. Filtre intelligent des taux (20.00, 10.00) vs montants réels
  4. Remplissage direct du canva Excel d'audit
  5. Cross-checks automatiques (132 - 190 = 200 ou 201)
  6. Log des fichiers problématiques

MAPPING DÉCLARATION DGI → CANVA EXCEL :
  Col A : Mois
  Col B : CA Hors champs (L20)
  Col C : CA taxable 20%    (somme bases L80-95,102)
  Col D : TVA à 20%         (somme taxes L80-95,102)
  Col E : CA taxable 14%    (L104)
  Col F : TVA à 14%
  Col G : CA taxable 10%    (somme bases L63,73-79,85-86,89,91-92,96-101,103,106,108-109,112,117-118)
  Col H : TVA à 10%
  Col I : CA taxable 7%     (L119)
  Col J : TVA à 7%
  Col K : TVA non résidents (L129)
  Col L : Retenue source    (L131)
  Col M : Total TVA Collectée (L132)
  Col N : TVA déductible    (L190)
  Col O : TVA due/Crédit calculé (=M-N)
  Col P : TVA due/Crédit déclaration (L200 ou -L201)
  Col Q : Écart KMAD        (=(O-P)/1000)
"""

import os
import re
import sys
import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict

# ============================================================
# CONFIGURATION — Détection automatique Tesseract / Poppler
# ============================================================
import importlib.util as _ilu

def _load_ocr_paths():
    """Charge core/ocr_paths.py depuis la racine de l'application (avec cache sys.modules)."""
    import sys as _sys
    _mod_name = "core.ocr_paths"
    if _mod_name in _sys.modules:
        return _sys.modules[_mod_name]
    _here = Path(__file__).resolve()
    for _parent in _here.parents:
        _candidate = _parent / "core" / "ocr_paths.py"
        if _candidate.exists():
            _spec = _ilu.spec_from_file_location(_mod_name, str(_candidate))
            _mod = _ilu.module_from_spec(_spec)
            _sys.modules[_mod_name] = _mod
            _spec.loader.exec_module(_mod)
            return _mod
    return None

_ocr = _load_ocr_paths()
if _ocr:
    TESSERACT_PATH = _ocr.TESSERACT_PATH
    POPPLER_PATH   = _ocr.POPPLER_PATH
else:
    TESSERACT_PATH = None
    POPPLER_PATH   = None

BASE_DIR = Path.home() / "AuditPro_output" / "tva_ir"
TVA_DIR = BASE_DIR / "TVA"
OUTPUT_DIR = BASE_DIR / "output"

try:
    TVA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass

import pdfplumber  # inclus dans requirements.txt

try:
    import pytesseract
    from pdf2image import convert_from_path
    if TESSERACT_PATH:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
except ImportError:
    pytesseract = None          # Mode dégradé : OCR désactivé
    convert_from_path = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MOIS = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
         7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}
MOIS_REV = {v.lower():k for k,v in MOIS.items()}
MOIS_REV.update({'aout':8,'decembre':12,'fevrier':2})

# Taux qui NE SONT PAS des montants (à exclure lors de l'extraction)
KNOWN_RATES = {"7.00","8.00","9.00","10.00","11.00","12.00","13.00",
               "14.00","15.00","16.00","18.00","20.00",
               "7,00","8,00","9,00","10,00","11,00","12,00","13,00",
               "14,00","15,00","16,00","18,00","20,00"}

# ============================================================
# PARSE_AMOUNT UNIVERSEL — LE FIX PRINCIPAL
# ============================================================
# Regex qui capture les deux formats DGI :
#   Anglo-saxon : 1,834,791.03  →  \d{1,3}(,\d{3})*\.\d{2}
#   Européen    : 6 516 944,20  →  \d{1,3}([\s\u00A0]\d{3})*,\d{2}
# Aussi : 0,00  0.00  153,71  539,131.48
AMOUNT_RE = re.compile(
    r'(\d{1,3}(?:[\s\u00A0]\d{3})+,\d{2})'   # Européen avec espaces : 6 516 944,20
    r'|(\d{1,3}(?:,\d{3})+\.\d{2})'           # Anglo-saxon avec virgules : 1,834,791.03
    r'|(\d+\.\d{2})'                            # Simple décimal : 539131.48
    r'|(\d+,\d{2})'                              # Simple virgule : 153,71  0,00
)

def parse_amount(s: str) -> Optional[float]:
    """Parse un montant DGI dans n'importe quel format."""
    if not s:
        return None
    s = s.strip()
    # Européen avec espaces : "6 516 944,20" → enlever espaces, virgule→point
    if re.match(r'^\d{1,3}(?:[\s\u00A0]\d{3})+,\d{2}$', s):
        return float(s.replace('\u00A0','').replace(' ','').replace(',','.'))
    # Anglo-saxon : "1,834,791.03" → enlever virgules
    if re.match(r'^\d{1,3}(?:,\d{3})+\.\d{2}$', s):
        return float(s.replace(',',''))
    # Simple avec point : "539131.48"
    if re.match(r'^\d+\.\d{2}$', s):
        return float(s)
    # Simple avec virgule : "153,71" ou "0,00"
    if re.match(r'^\d+,\d{2}$', s):
        return float(s.replace(',','.'))
    # Fallback
    try:
        cleaned = s.replace('\u00A0','').replace(' ','')
        if ',' in cleaned and '.' in cleaned:
            if cleaned.rfind(',') > cleaned.rfind('.'):
                return float(cleaned.replace('.','').replace(',','.'))
            else:
                return float(cleaned.replace(',',''))
        elif ',' in cleaned:
            return float(cleaned.replace(',','.'))
        return float(cleaned)
    except:
        return None

def find_amounts_in_text(text: str) -> List[str]:
    """Trouve tous les montants dans un texte (deux formats)."""
    results = []
    for m in AMOUNT_RE.finditer(text):
        val = m.group(0).strip()
        if val:
            results.append(val)
    return results

# ============================================================
# DATACLASS
# ============================================================
@dataclass
class DeclarationTVA:
    fichier: str = ""
    type_pdf: str = ""
    mois: str = ""
    mois_num: int = 0
    annee: int = 0
    raison_sociale: str = ""
    identifiant_fiscal: str = ""
    rc: str = ""
    etat: str = ""  # Brouillon, Déposée

    # Section A
    ligne_10: Optional[float] = None
    ligne_20: Optional[float] = None
    ligne_60: Optional[float] = None

    # Agrégats par taux
    ca_20: float = 0.0;  tva_20: float = 0.0
    ca_14: float = 0.0;  tva_14: float = 0.0
    ca_10: float = 0.0;  tva_10: float = 0.0
    ca_7: float = 0.0;   tva_7: float = 0.0

    # Lignes clés
    ligne_129: Optional[float] = None
    ligne_131: Optional[float] = None
    ligne_132: Optional[float] = None
    ligne_170: Optional[float] = None
    ligne_182: Optional[float] = None
    ligne_190: Optional[float] = None
    ligne_200: Optional[float] = None
    ligne_201: Optional[float] = None
    ligne_205: Optional[float] = None
    montant_total: Optional[float] = None

    anomalies: List[str] = field(default_factory=list)

# ============================================================
# EXTRACTION PAR NUMÉRO DE LIGNE — COMPATIBLE DEUX FORMATS
# ============================================================

def get_line_amounts(text: str, line_num: int, max_wrap_lines: int = 3) -> List[float]:
    """
    Trouve la ligne commençant par line_num et extrait tous les montants.
    Gère le wrapping sur plusieurs lignes du PDF.
    Retourne une liste de floats (montants réels, taux exclus).
    """
    lines = text.split('\n')
    for i, line in enumerate(lines):
        if not re.match(rf'^\s*{line_num}\b', line):
            continue
        # Combiner avec les lignes suivantes (wrapping)
        combined = line
        for j in range(1, max_wrap_lines + 1):
            if i + j >= len(lines):
                break
            next_l = lines[i + j]
            if re.match(r'^\s*\d{2,3}\b', next_l):
                break
            combined += " " + next_l

        raw_amounts = find_amounts_in_text(combined)
        # Filtrer les taux connus et les pourcentages
        real = []
        for a in raw_amounts:
            if a in KNOWN_RATES:
                continue
            # Exclure aussi "100.00" qui est le prorata
            v = parse_amount(a)
            if v is not None:
                real.append(v)
        return real
    return []


def get_single(text: str, line_num: int) -> Optional[float]:
    """Extrait la dernière valeur d'une ligne (typiquement le montant total)."""
    amounts = get_line_amounts(text, line_num)
    return amounts[-1] if amounts else None


def get_base_taxe(text: str, line_num: int) -> Tuple[float, float]:
    """Extrait (base, taxe) d'une ligne de la section B."""
    amounts = get_line_amounts(text, line_num)
    if len(amounts) >= 2:
        return amounts[0], amounts[1]
    elif len(amounts) == 1:
        return amounts[0], 0.0
    return 0.0, 0.0


# ============================================================
# EXTRACTION PRINCIPALE
# ============================================================

def extract_declaration(text: str, filename: str) -> DeclarationTVA:
    d = DeclarationTVA(fichier=filename)

    # Identité
    m = re.search(r"Raison\s+sociale\s*:\s*(.+?)(?:\n|$)", text)
    if m: d.raison_sociale = m.group(1).strip().split(':')[0].strip()

    m = re.search(r"Identifiant\s+fiscal\s*:\s*(\d+)", text)
    if m: d.identifiant_fiscal = m.group(1)

    m = re.search(r"R\.C\.N°\s*:\s*(\d+)", text)
    if m: d.rc = m.group(1)

    m = re.search(r"Etat\s+d[ée]claration\s*:\s*(\w+)", text, re.I)
    if m: d.etat = m.group(1)

    # Période
    m = re.search(r"P[ée]riode\s*:\s*(\d{1,2})\s+(\d{4})", text, re.I)
    if m:
        d.mois_num = int(m.group(1))
        d.annee = int(m.group(2))
        d.mois = MOIS.get(d.mois_num, "")

    # Section A
    d.ligne_10 = get_single(text, 10)
    d.ligne_20 = get_single(text, 20)
    d.ligne_60 = get_single(text, 60)

    # Section B — TVA 20% (lignes 80,81,82,83,87,93,94,95,102)
    for ln in [80,81,82,83,87,93,94,95,102]:
        b, t = get_base_taxe(text, ln)
        if b > 0: d.ca_20 += b
        if t > 0: d.tva_20 += t

    # TVA 14% (ligne 104)
    b, t = get_base_taxe(text, 104)
    d.ca_14 += b; d.tva_14 += t

    # TVA 10% (lignes 63,73,75-79,85-86,89,91-92,96-101,103,106,108-109,112,117-118)
    for ln in [63,73,75,76,77,78,79,85,86,89,91,92,96,97,98,99,100,101,103,106,108,109,112,117,118]:
        b, t = get_base_taxe(text, ln)
        if b > 0: d.ca_10 += b
        if t > 0: d.tva_10 += t

    # TVA 7% (ligne 119)
    b, t = get_base_taxe(text, 119)
    d.ca_7 += b; d.tva_7 += t

    # Lignes spéciales
    d.ligne_129 = get_single(text, 129)
    d.ligne_131 = get_single(text, 131)
    d.ligne_132 = get_single(text, 132)
    d.ligne_170 = get_single(text, 170)
    d.ligne_182 = get_single(text, 182)
    d.ligne_190 = get_single(text, 190)
    d.ligne_200 = get_single(text, 200)
    d.ligne_201 = get_single(text, 201)
    d.ligne_205 = get_single(text, 205)

    # Montant total
    m = re.search(r"Montant\s+Total\s+", text, re.I)
    if m:
        after = text[m.end():]
        amts = find_amounts_in_text(after[:100])
        real = [parse_amount(a) for a in amts if a not in KNOWN_RATES]
        real = [v for v in real if v is not None]
        if real:
            d.montant_total = real[-1]

    # ========== CROSS-CHECKS ==========
    # 1. TVA collectée vs somme par taux
    somme_taux = d.tva_20 + d.tva_14 + d.tva_10 + d.tva_7
    if d.ligne_132 and somme_taux > 0:
        ecart = abs(d.ligne_132 - somme_taux)
        if ecart > 1:
            d.anomalies.append(f"L132={d.ligne_132:.2f} ≠ somme taux={somme_taux:.2f} (Δ={ecart:.2f})")

    # 2. TVA due = collectée - déductible
    if d.ligne_132 is not None and d.ligne_190 is not None:
        calc = d.ligne_132 - d.ligne_190
        if calc > 0 and d.ligne_200:
            ecart = abs(calc - d.ligne_200)
            if ecart > 1:
                d.anomalies.append(f"TVA due calc={calc:.2f} ≠ L200={d.ligne_200:.2f}")
        elif calc < 0 and d.ligne_201:
            ecart = abs(abs(calc) - d.ligne_201)
            if ecart > 1:
                d.anomalies.append(f"Crédit calc={abs(calc):.2f} ≠ L201={d.ligne_201:.2f}")

    # 3. CA cohérence
    if d.ligne_10 and d.ligne_60:
        l20 = d.ligne_20 or 0
        reconst = d.ligne_60 + l20
        ecart = abs(d.ligne_10 - reconst)
        if ecart > 1:
            d.anomalies.append(f"L10={d.ligne_10:.2f} ≠ L60+L20={reconst:.2f}")

    # 4. Base 20% vs CA imposable
    if d.ligne_60 and d.ca_20 > 0:
        # Pour une société 100% services à 20%, CA imposable ≈ base 20%
        ecart_pct = abs(d.ligne_60 - d.ca_20) / d.ligne_60 * 100 if d.ligne_60 > 0 else 0
        if ecart_pct > 5 and d.ca_14 == 0 and d.ca_10 == 0 and d.ca_7 == 0:
            d.anomalies.append(f"L60={d.ligne_60:.2f} vs CA20={d.ca_20:.2f} (Δ={ecart_pct:.1f}%)")

    # Anomalies champs manquants
    if not d.mois: d.anomalies.append("Période manquante")
    if d.ligne_132 is None: d.anomalies.append("L132 manquante")
    if d.ligne_190 is None: d.anomalies.append("L190 manquante")

    return d


# ============================================================
# TRAITEMENT PDF
# ============================================================

def is_native(pdf_path: Path) -> bool:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if pdf.pages:
                t = pdf.pages[0].extract_text()
                return bool(t and len(t.strip()) > 100)
    except: pass
    return False

def extract_text(pdf_path: Path) -> Tuple[str, str]:
    """Retourne (texte_complet, type_pdf)."""
    if is_native(pdf_path):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                texts = [p.extract_text() or "" for p in pdf.pages]
                return "\n".join(texts), "natif"
        except Exception as e:
            print(f"   ❌ Erreur pdfplumber: {e}")

    # Fallback OCR
    if pytesseract is None or convert_from_path is None:
        print("   ⚠️ OCR non disponible (Tesseract absent). PDF scanné ignoré.")
        return "", "erreur"
    try:
        print(f"   🔄 OCR...")
        images = convert_from_path(str(pdf_path), dpi=250, poppler_path=POPPLER_PATH or None)
        texts = []
        for i, img in enumerate(images):
            texts.append(pytesseract.image_to_string(img, lang="fra+eng", config="--psm 6"))
        return "\n".join(texts), "scanné"
    except Exception as e:
        print(f"   ❌ Erreur OCR: {e}")
        return "", "erreur"


def process_file(pdf_path: Path) -> Optional[DeclarationTVA]:
    print(f"\n📄 {pdf_path.name}")
    text, pdf_type = extract_text(pdf_path)
    if not text or len(text) < 50:
        print(f"   ⚠️ Aucun texte extrait")
        return None

    print(f"   Type: {pdf_type} | {len(text)} chars | {text.count(chr(10))} lignes")

    # Debug
    debug = OUTPUT_DIR / f"DEBUG_{pdf_path.stem}.txt"
    with open(debug, "w", encoding="utf-8") as f:
        f.write(text)

    d = extract_declaration(text, pdf_path.name)
    d.type_pdf = pdf_type

    # Affichage
    print(f"   📅 {d.mois} {d.annee} | {d.raison_sociale} | État: {d.etat}")
    print(f"   CA total (L10): {d.ligne_10:,.2f}" if d.ligne_10 else "   CA total: ?")
    print(f"   TVA collectée (L132): {d.ligne_132:,.2f}" if d.ligne_132 else "   L132: ?")
    print(f"   TVA déductible (L190): {d.ligne_190:,.2f}" if d.ligne_190 else "   L190: ?")

    if d.ligne_200 and d.ligne_200 > 0:
        print(f"   💰 TVA DUE (L200): {d.ligne_200:,.2f}")
    elif d.ligne_201 and d.ligne_201 > 0:
        print(f"   💚 CRÉDIT (L201): {d.ligne_201:,.2f}")

    if d.montant_total and d.montant_total > 0:
        print(f"   🧾 Montant total à payer: {d.montant_total:,.2f}")

    if d.anomalies:
        for a in d.anomalies:
            print(f"   ⚠️ {a}")

    return d


# ============================================================
# REMPLISSAGE CANVA EXCEL
# ============================================================

CANVA_HEADERS = [
    "Mois", "CA Hors champs de TVA",
    "CA taxable à 20%", "TVA à 20%",
    "CA taxable à 14%", "TVA à 14%",
    "CA taxable à 10%", "TVA à 10%",
    "CA taxable à 7%", "TVA à 7%",
    "TVA sur opérations réalisées avec contribuables non résidents",
    "la retenue à la source opérée par les clients",
    "Total TVA Collectée", "TVA déductible",
    "TVA due / Crédit de TVA",
    "TVA due / Crédit TVA  selon déclaration",
    "Ecart en KMAD", "Retard"
]

def fill_canva(declarations: List[DeclarationTVA], canva_path: Optional[Path], output_path: Path):
    # Charger ou créer le canva
    if canva_path and canva_path.exists():
        print(f"\n📋 Canva chargé: {canva_path.name}")
        wb = load_workbook(canva_path)
        ws = wb.active
    else:
        print(f"\n📋 Nouveau canva créé")
        wb = Workbook(); ws = wb.active; ws.title = "TVA"
        ws.append(CANVA_HEADERS)
        hf = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(name="Arial", color="FFFFFF", bold=True, size=9)
        for c in range(1, len(CANVA_HEADERS)+1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.append(["Dec-23"])
        for m in range(1, 13): ws.append([MOIS[m]])
        ws.append(["Total"])

    # Mapper mois → row
    mois_row = {}
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip().lower()
        for name, num in MOIS_REV.items():
            if name in v:
                mois_row[num] = r
                break

    filled = 0
    for d in declarations:
        if d.mois_num not in mois_row:
            print(f"   ⚠️ {d.mois}: pas trouvé dans le canva")
            continue

        row = mois_row[d.mois_num]
        print(f"   ✏️ {d.mois} {d.annee} → Row {row}")

        ws.cell(row=row, column=2, value=d.ligne_20 or 0)           # B: Hors champ
        ws.cell(row=row, column=3, value=d.ca_20 or 0)              # C: CA 20%
        ws.cell(row=row, column=4, value=d.tva_20 or 0)             # D: TVA 20%
        ws.cell(row=row, column=5, value=d.ca_14 or 0)              # E: CA 14%
        ws.cell(row=row, column=6, value=d.tva_14 or 0)             # F: TVA 14%
        ws.cell(row=row, column=7, value=d.ca_10 or 0)              # G: CA 10%
        ws.cell(row=row, column=8, value=d.tva_10 or 0)             # H: TVA 10%
        ws.cell(row=row, column=9, value=d.ca_7 or 0)               # I: CA 7%
        ws.cell(row=row, column=10, value=d.tva_7 or 0)             # J: TVA 7%
        ws.cell(row=row, column=11, value=d.ligne_129 or 0)         # K: Non résidents
        ws.cell(row=row, column=12, value=d.ligne_131 or 0)         # L: Retenue
        ws.cell(row=row, column=13, value=d.ligne_132 or 0)         # M: TVA collectée
        ws.cell(row=row, column=14, value=d.ligne_190 or 0)         # N: TVA déductible
        ws.cell(row=row, column=15).value = f"=M{row}-N{row}"       # O: Due/Crédit calc
        # P: Due/Crédit déclaration
        if d.ligne_200 and d.ligne_200 > 0:
            ws.cell(row=row, column=16, value=d.ligne_200)
        elif d.ligne_201 and d.ligne_201 > 0:
            ws.cell(row=row, column=16, value=-d.ligne_201)
        else:
            ws.cell(row=row, column=16, value=0)
        ws.cell(row=row, column=17).value = f"=(O{row}-P{row})/1000"  # Q: Écart

        # Formatage
        for c in range(2, 17):
            cell = ws.cell(row=row, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            cell.font = Font(name="Arial", size=9)

        # Colorer crédit/dû
        p = ws.cell(row=row, column=16)
        if isinstance(p.value, (int, float)):
            if p.value < 0:
                p.font = Font(name="Arial", size=9, color="008000")
            elif p.value > 0:
                p.font = Font(name="Arial", size=9, color="FF0000")

        filled += 1

    # Formules Total
    total_row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip().lower() == "total":
            total_row = r; break

    if total_row and mois_row:
        fr = min(mois_row.values()); lr = max(mois_row.values())
        for c in range(2, 17):
            cl = get_column_letter(c)
            ws.cell(row=total_row, column=c).value = f"=SUM({cl}{fr}:{cl}{lr})"
            ws.cell(row=total_row, column=c).number_format = '#,##0.00'
            ws.cell(row=total_row, column=c).font = Font(name="Arial", size=9, bold=True)

    # Mise en forme
    for c, w in {1:14,2:18,3:18,4:16,5:18,6:16,7:18,8:16,9:16,10:14,
                 11:20,12:20,13:18,14:18,15:20,16:22,17:14,18:10}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\n✅ Canva → {output_path}")
    print(f"   {filled}/{len(declarations)} mois rempli(s)")
    return filled


# ============================================================
# FEUILLE DÉTAIL + LOG
# ============================================================

def generate_detail(declarations: List[DeclarationTVA], output_path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Détail"
    headers = [
        "Fichier","Type","Mois","Année","Société","IF","RC","État",
        "L10 CA","L20 Hors Champ","L60 CA Impos.",
        "CA 20%","TVA 20%","CA 14%","TVA 14%","CA 10%","TVA 10%","CA 7%","TVA 7%",
        "L129 Non Rés.","L131 Retenue","L132 TVA Coll.",
        "L170 Crédit Préc.","L182 Tot Déd.","L190 TVA Déd.",
        "L200 Due","L201 Crédit","L205 Période","Mt Total",
        "Anomalies"
    ]
    ws.append(headers)
    hf = PatternFill("solid", fgColor="1F4E79")
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hf; cell.font = Font(name="Arial", color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for d in sorted(declarations, key=lambda x: (x.annee, x.mois_num)):
        ws.append([
            d.fichier, d.type_pdf, d.mois, d.annee, d.raison_sociale, d.identifiant_fiscal, d.rc, d.etat,
            d.ligne_10, d.ligne_20, d.ligne_60,
            d.ca_20, d.tva_20, d.ca_14, d.tva_14, d.ca_10, d.tva_10, d.ca_7, d.tva_7,
            d.ligne_129, d.ligne_131, d.ligne_132,
            d.ligne_170, d.ligne_182, d.ligne_190,
            d.ligne_200, d.ligne_201, d.ligne_205, d.montant_total,
            "; ".join(d.anomalies) if d.anomalies else "OK"
        ])
        rn = ws.max_row
        for c in range(9, 29):
            cell = ws.cell(row=rn, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2, 22)

    # Feuille LOG
    ws_log = wb.create_sheet("Log")
    ws_log.append(["Fichier", "Mois", "Anomalies"])
    for d in declarations:
        if d.anomalies:
            ws_log.append([d.fichier, f"{d.mois} {d.annee}", "; ".join(d.anomalies)])
    if ws_log.max_row == 1:
        ws_log.append(["", "", "Aucune anomalie détectée"])

    wb.save(output_path)
    print(f"✅ Détail + Log → {output_path}")


# ============================================================
# MENU
# ============================================================

def main():
    print("\n" + "="*70)
    print("  EXTRACT-TVA V3 — Déclarations Mensuelles TVA (DGI Maroc)")
    print("  Gère les formats Anglo-saxon ET Européen")
    print("="*70)
    print(f"\n📁 TVA    : {TVA_DIR}")
    print(f"📁 Output : {OUTPUT_DIR}")

    # Chercher canva existant
    canva_candidates = list(OUTPUT_DIR.glob("canva*.xlsx")) + list(TVA_DIR.glob("canva*.xlsx"))
    canva_path = canva_candidates[0] if canva_candidates else None
    print(f"📋 Canva  : {canva_path.name if canva_path else 'Nouveau'}")

    pdf_files = sorted(TVA_DIR.glob("*.pdf"))
    if not pdf_files:
        print(f"\n❌ Aucun PDF dans {TVA_DIR}")
        input("Entrée..."); return

    print(f"\n📄 {len(pdf_files)} fichier(s):")
    for i, f in enumerate(pdf_files, 1):
        print(f"   {i}. {f.name}")

    choix = input(f"\nSélection (T=Tous, ou ex: 1,3,12) [T]: ").strip().upper()
    if choix and choix != 'T':
        sel = []
        for c in choix.split(','):
            c = c.strip()
            if c.isdigit() and 1 <= int(c) <= len(pdf_files):
                sel.append(pdf_files[int(c)-1])
        if sel: pdf_files = sel

    # Traitement
    declarations = []
    errors = []
    for pdf in pdf_files:
        try:
            d = process_file(pdf)
            if d: declarations.append(d)
            else: errors.append((pdf.name, "Extraction échouée"))
        except Exception as e:
            print(f"   ❌ ERREUR: {e}")
            errors.append((pdf.name, str(e)))

    if not declarations:
        print("\n❌ Aucune déclaration extraite.")
        if errors:
            print("\n📋 Erreurs:")
            for f, e in errors:
                print(f"   - {f}: {e}")
        input("Entrée..."); return

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    soc = declarations[0].raison_sociale or "SOCIETE"
    soc_clean = re.sub(r'[\\/*?:"<>|]', '_', soc)

    # 1. Canva
    canva_out = OUTPUT_DIR / f"TVA_Canva_{soc_clean}_{ts}.xlsx"
    fill_canva(declarations, canva_path, canva_out)

    # 2. Détail
    detail_out = OUTPUT_DIR / f"TVA_Detail_{soc_clean}_{ts}.xlsx"
    generate_detail(declarations, detail_out)

    # Résumé
    print(f"\n{'='*60}")
    print(f"📊 RÉSUMÉ — {soc}")
    print(f"{'='*60}")
    print(f"{'Mois':>12} {'CA HT':>14} {'TVA Coll.':>12} {'TVA Déd.':>12} {'Solde':>14}")
    print(f"{'-'*12} {'-'*14} {'-'*12} {'-'*12} {'-'*14}")

    for d in sorted(declarations, key=lambda x: (x.annee, x.mois_num)):
        ca = d.ligne_10 or 0
        coll = d.ligne_132 or 0
        ded = d.ligne_190 or 0
        solde = coll - ded
        label = f"DÛ {solde:,.2f}" if solde > 0 else f"CRÉDIT {abs(solde):,.2f}"
        print(f"{d.mois:>12} {ca:>14,.2f} {coll:>12,.2f} {ded:>12,.2f} {label:>14}")

    tot_coll = sum(d.ligne_132 or 0 for d in declarations)
    tot_ded = sum(d.ligne_190 or 0 for d in declarations)
    tot_solde = tot_coll - tot_ded
    print(f"{'-'*12} {'-'*14} {'-'*12} {'-'*12} {'-'*14}")
    print(f"{'TOTAL':>12} {'':>14} {tot_coll:>12,.2f} {tot_ded:>12,.2f} {'DÛ' if tot_solde>0 else 'CRÉDIT'} {abs(tot_solde):>10,.2f}")

    if errors:
        print(f"\n⚠️ {len(errors)} fichier(s) en erreur:")
        for f, e in errors:
            print(f"   - {f}: {e}")

    input("\n✅ Terminé! Entrée...")

if __name__ == "__main__":
    main()