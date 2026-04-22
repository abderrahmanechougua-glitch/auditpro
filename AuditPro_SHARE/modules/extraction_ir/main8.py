#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXTRACT-IR-SIMPLE - Extraction simple des Avis de Versement IR
Support: PDF natifs (pdfplumber) et PDF scannés (OCR)
Avec score de qualité pour les fichiers scannés
Seulement: Mois, Date de versement, Montant
"""

import os
import re
import sys
import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Optional

# ============================================================
# CONFIGURATION
# ============================================================

# Détection automatique Tesseract / Poppler via core/ocr_paths.py
import importlib.util as _ilu

def _load_ocr_paths():
    """Charge core/ocr_paths.py depuis la racine de l'application (avec cache sys.modules)."""
    import sys as _sys
    _mod_name = "core.ocr_paths"
    if _mod_name in _sys.modules:
        return _sys.modules[_mod_name]
    _here = Path(__file__).resolve()
    for _parent in _here.parents:
        _candidate = _parent / "core" / "ocr_paths.py"
        if _candidate.exists():
            _spec = _ilu.spec_from_file_location(_mod_name, str(_candidate))
            _mod = _ilu.module_from_spec(_spec)
            _sys.modules[_mod_name] = _mod
            _spec.loader.exec_module(_mod)
            return _mod
    return None

_ocr = _load_ocr_paths()
if _ocr:
    TESSERACT_PATH = _ocr.TESSERACT_PATH
    POPPLER_PATH   = _ocr.POPPLER_PATH
else:
    TESSERACT_PATH = None
    POPPLER_PATH   = None

BASE_DIR = Path.home() / "AuditPro_output" / "tva_ir"
IR_DIR = BASE_DIR / "IR"
OUTPUT_DIR = BASE_DIR / "output"

try:
    IR_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass

import pdfplumber  # inclus dans requirements.txt

try:
    import pytesseract
    from pdf2image import convert_from_path
    if TESSERACT_PATH:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
except ImportError:
    pytesseract = None          # Mode dégradé : OCR désactivé
    convert_from_path = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# DATACLASSES
# ============================================================

@dataclass
class DeclarationIR:
    """Structure simple pour IR"""
    fichier: str = ""
    mois: str = ""
    annee: int = 0
    montant: Optional[float] = None
    date_versement: str = ""
    type_document: str = ""
    type_pdf: str = ""  # "natif" ou "scanne"
    qualite_ocr: int = 0  # Score de 0 à 100 pour les fichiers scannés
    statut: str = "OK"

# ============================================================
# FONCTIONS
# ============================================================

def normalize_amount(amount_str: str) -> Optional[float]:
    """Convertit un montant en float"""
    if not amount_str:
        return None
    
    s = amount_str.strip()
    s = s.replace(" ", "").replace("\u00A0", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    
    try:
        return float(s)
    except:
        return None

def normalize_date(date_str: str) -> str:
    """Normalise une date au format JJ/MM/AAAA"""
    if not date_str:
        return ""
    
    s = date_str.strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        jj, mm, aaaa = m.groups()
        if len(aaaa) == 2:
            aaaa = "20" + aaaa
        return f"{int(jj):02d}/{int(mm):02d}/{aaaa}"
    
    return s

def get_month_number(month_name: str) -> int:
    """Convertit un nom de mois en numéro"""
    mois_fr = {
        'janvier': 1, 'février': 2, 'mars': 3, 'avril': 4, 'mai': 5, 'juin': 6,
        'juillet': 7, 'août': 8, 'septembre': 9, 'octobre': 10, 'novembre': 11, 'décembre': 12,
        'jan': 1, 'fév': 2, 'mar': 3, 'avr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'aoû': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'déc': 12
    }
    return mois_fr.get(month_name.lower(), 0)

mois_names = {
    1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril', 5: 'Mai', 6: 'Juin',
    7: 'Juillet', 8: 'Août', 9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
}

def extract_from_filename(filename: str) -> Tuple[Optional[int], Optional[int]]:
    """Extrait mois et année depuis le nom du fichier (ex: IR 12-2025.pdf)"""
    patterns = [
        r"IR\s+(\d{1,2})[-/](\d{4})",
        r"(\d{1,2})[-/](\d{4})",
    ]
    for pattern in patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match and match.group(1).isdigit():
            mois_num = int(match.group(1))
            annee = int(match.group(2))
            if 1 <= mois_num <= 12:
                return mois_num, annee
    return None, None

def is_pdf_native(pdf_path: Path) -> bool:
    """Détecte si un PDF est natif (texte extractible) ou scanné"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) > 0:
                text = pdf.pages[0].extract_text()
                if text and len(text.strip()) > 50:
                    return True
        return False
    except:
        return False

def extract_text_native(pdf_path: Path) -> Tuple[str, int]:
    """Extrait le texte d'un PDF natif avec pdfplumber, retourne (texte, qualite)"""
    try:
        all_text = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text.append(text)
        text = "\n".join(all_text)
        return text, 100  # PDF natif = qualité 100%
    except Exception as e:
        print(f"   ❌ Erreur extraction PDF natif: {e}")
        return "", 0

def calculate_ocr_quality(text: str, original_pdf_size: int) -> int:
    """
    Calcule un score de qualité OCR (0-100)
    Critères:
    - Longueur du texte extrait (30 points)
    - Présence de mots-clés importants (40 points)
    - Taille du fichier vs texte (10 points)
    - Détection de nombres et dates (20 points)
    """
    score = 0
    
    # 1. Longueur du texte (max 30 points)
    text_length = len(text)
    if text_length > 2000:
        score += 30
    elif text_length > 1000:
        score += 25
    elif text_length > 500:
        score += 20
    elif text_length > 200:
        score += 15
    elif text_length > 100:
        score += 10
    elif text_length > 50:
        score += 5
    
    # 2. Présence de mots-clés importants (max 40 points)
    mots_cles = {
        'impot': 5, 'revenu': 5, 'montant': 5, 'total': 5,
        'versement': 4, 'payer': 4, 'mois': 4, 'année': 4,
        'principal': 4
    }
    mots_trouves = 0
    for mot, points in mots_cles.items():
        if re.search(rf'\b{mot}\b', text, re.IGNORECASE):
            score += points
            mots_trouves += 1
    
    # Bonus si beaucoup de mots-clés trouvés
    if mots_trouves >= 5:
        score += 5
    
    # 3. Taille du fichier vs texte (max 10 points)
    # Plus le texte est long par rapport à la taille du fichier, meilleur est l'OCR
    if original_pdf_size > 0:
        ratio = text_length / original_pdf_size
        if ratio > 0.5:
            score += 10
        elif ratio > 0.3:
            score += 7
        elif ratio > 0.1:
            score += 4
    
    # 4. Détection de nombres et dates (max 20 points)
    # Compter les nombres trouvés
    nombres = re.findall(r'\d{1,3}(?:[\s.,]\d{3})*(?:,\d{2})?', text)
    if len(nombres) >= 5:
        score += 15
    elif len(nombres) >= 3:
        score += 10
    elif len(nombres) >= 1:
        score += 5
    
    # Détection de dates
    dates = re.findall(r'\d{1,2}/\d{1,2}/\d{2,4}', text)
    if dates:
        score += 5
    
    # Limiter à 100
    return min(score, 100)

def extract_text_scanned(pdf_path: Path) -> Tuple[str, int]:
    """Extrait le texte d'un PDF scanné avec OCR, retourne (texte, qualite)"""
    if pytesseract is None or convert_from_path is None:
        print("   ⚠️ OCR non disponible (Tesseract absent). PDF scanné ignoré.")
        return "", 0
    try:
        pdf_size = pdf_path.stat().st_size
        
        print(f"   🔄 Conversion PDF scanné en images...")
        images = convert_from_path(str(pdf_path), dpi=200, poppler_path=POPPLER_PATH or None)
        all_text = []
        
        # Configuration OCR pour meilleure qualité
        custom_config = r'--oem 3 --psm 6 -l fra+eng'
        
        for i, img in enumerate(images):
            print(f"      Page {i+1}/{len(images)} - OCR en cours...", end="\r")
            text = pytesseract.image_to_string(img, config=custom_config)
            all_text.append(text)
        
        print(f"   ✅ {len(images)} pages traitées" + " " * 30)
        text = "\n".join(all_text)
        
        # Calculer la qualité OCR
        qualite = calculate_ocr_quality(text, pdf_size)
        
        return text, qualite
    except Exception as e:
        print(f"   ❌ Erreur OCR: {e}")
        return "", 0

def detect_document_type(text: str) -> str:
    """Détecte le type de document: avis_versement ou accuse_depot"""
    if re.search(r"BORDERAU-AVIS\s+DE\s+VERSEMENT", text, re.IGNORECASE):
        return "avis_versement"
    if re.search(r"Accus[ée]\s+de\s+(d[ée]p[ôo]t|r[ée]ception)", text, re.IGNORECASE):
        return "accuse_depot"
    if re.search(r"Accus[ée]\s+de\s+r[ée]ception", text, re.IGNORECASE):
        return "accuse_depot"
    return "inconnu"

def extract_amount_from_text(text: str, doc_type: str) -> Optional[float]:
    """Extrait le montant IR à payer selon le type de document"""
    
    if doc_type == "avis_versement":
        # PRIORITÉ: "Montant total à payer"
        match = re.search(r"Montant\s+total\s+[àa]\s+payer\s*:\s*([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
        
        # Fallback: "Montant en principal"
        match = re.search(r"Montant\s+en\s+principal\s*:\s*([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
    
    elif doc_type == "accuse_depot":
        # PRIORITÉ: "Total"
        match = re.search(r"Total\s+([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
        
        # Fallback: "Montant en principal a payer"
        match = re.search(r"Montant\s+en\s+principal\s+a\s+payer\s+([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
    
    # Si type inconnu, essayer les deux
    else:
        # Essayer "Montant total à payer"
        match = re.search(r"Montant\s+total\s+[àa]\s+payer\s*:\s*([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
        
        # Sinon "Total"
        match = re.search(r"Total\s+([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
        
        # En dernier "Montant en principal"
        match = re.search(r"Montant\s+en\s+principal\s*:\s*([\d\s\.,]+)", text, re.IGNORECASE)
        if match:
            val = normalize_amount(match.group(1))
            if val:
                return val
    
    return None

def extract_month_from_text(text: str) -> Tuple[str, int]:
    """Extrait le mois depuis le texte"""
    # Format: "Mois: Novembre / Année: 2025"
    match = re.search(r"Mois\s*:\s*(\w+)\s*[/]\s*Année\s*:\s*(\d{4})", text, re.IGNORECASE)
    if match:
        mois_nom = match.group(1)
        annee = int(match.group(2))
        mois_num = get_month_number(mois_nom)
        if mois_num > 0:
            return mois_names[mois_num], annee
    
    # Format: "Mois Novembre / 2025"
    match = re.search(r"Mois\s+(\w+)\s*[/]\s*(\d{4})", text, re.IGNORECASE)
    if match:
        mois_nom = match.group(1)
        annee = int(match.group(2))
        mois_num = get_month_number(mois_nom)
        if mois_num > 0:
            return mois_names[mois_num], annee
    
    # Format simple: "Mois Novembre"
    match = re.search(r"Mois\s+(\w+)", text, re.IGNORECASE)
    if match:
        mois_nom = match.group(1)
        mois_num = get_month_number(mois_nom)
        if mois_num > 0:
            return mois_names[mois_num], 0
    
    return "", 0

def extract_year_from_text(text: str) -> int:
    """Extrait l'année depuis le texte"""
    match = re.search(r"Année\s+(\d{4})", text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return 0

def extract_date_from_text(text: str) -> str:
    """Extrait la date de versement"""
    # Format "Le 26/12/2025"
    match = re.search(r"Le\s+(\d{1,2}/\d{1,2}/\d{4})", text, re.IGNORECASE)
    if match:
        return normalize_date(match.group(1))
    
    # Format "26/06/25 17:39"
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})\s+\d{1,2}:\d{2}", text)
    if match:
        return normalize_date(match.group(1))
    
    # Format "Date et heure de l'opération 26/06/25 17:39"
    match = re.search(r"Date\s+et\s+heure\s+de\s+l['']opération\s+(\d{1,2}/\d{1,2}/\d{2,4})", text, re.IGNORECASE)
    if match:
        return normalize_date(match.group(1))
    
    # Format "Edité le 26/12/2025"
    match = re.search(r"Editié le\s+(\d{1,2}/\d{1,2}/\d{4})", text, re.IGNORECASE)
    if match:
        return normalize_date(match.group(1))
    
    return ""

def get_quality_color(quality: int) -> str:
    """Retourne la couleur en fonction de la qualité"""
    if quality >= 80:
        return "🟢"  # Excellente
    elif quality >= 60:
        return "🟡"  # Bonne
    elif quality >= 40:
        return "🟠"  # Moyenne
    else:
        return "🔴"  # Faible

def process_ir_file(pdf_path: Path) -> DeclarationIR:
    """Traite un fichier IR (natif ou scanné)"""
    print(f"\n📄 {pdf_path.name}")
    
    # 1. Extraire depuis le nom du fichier
    mois_num, annee = extract_from_filename(pdf_path.name)
    mois = mois_names[mois_num] if mois_num else ""
    
    # 2. Détecter si PDF natif ou scanné
    is_native = is_pdf_native(pdf_path)
    pdf_type = "natif" if is_native else "scanne"
    print(f"   📄 Type: PDF {pdf_type}")
    
    # 3. Extraire le texte selon le type
    if is_native:
        text, qualite = extract_text_native(pdf_path)
    else:
        text, qualite = extract_text_scanned(pdf_path)
    
    if not text:
        print(f"   ⚠️ Aucun texte extrait")
        return DeclarationIR(fichier=pdf_path.name, type_pdf=pdf_type, qualite_ocr=qualite, statut="ERREUR")
    
    # Afficher la qualité pour les fichiers scannés
    if not is_native:
        qualite_color = get_quality_color(qualite)
        print(f"   📊 Qualité OCR: {qualite}% {qualite_color}")
        
        # Afficher un aperçu pour les faibles qualités
        if qualite < 50:
            print(f"   🔍 Aperçu OCR (premiers 300 caractères):")
            print(f"   {text[:300].replace(chr(10), chr(10)+'   ')}")
            print(f"   {'-'*50}")
    
    # 4. Détecter le type de document
    doc_type = detect_document_type(text)
    print(f"   📋 Type: {'Avis de versement' if doc_type == 'avis_versement' else 'Accusé de dépôt' if doc_type == 'accuse_depot' else 'Inconnu'}")
    
    # 5. Compléter avec les infos du texte si nécessaire
    if not mois:
        mois, annee_texte = extract_month_from_text(text)
        if not annee and annee_texte:
            annee = annee_texte
    
    if not annee:
        annee = extract_year_from_text(text)
    
    # 6. Extraire le montant
    montant = extract_amount_from_text(text, doc_type)
    
    # 7. Extraire la date
    date_versement = extract_date_from_text(text)
    
    # 8. Afficher le résultat
    mois_complet = f"{mois} {annee}" if annee else mois
    print(f"   📅 Mois: {mois_complet if mois_complet else 'Non trouvé'}")
    print(f"   💰 Montant IR: {montant:,.2f} MAD" if montant else "   💰 Montant IR: NON TROUVÉ")
    print(f"   📆 Date versement: {date_versement}" if date_versement else "   📆 Date: NON TROUVÉE")
    
    # 9. Créer la déclaration
    decl = DeclarationIR(
        fichier=pdf_path.name,
        mois=mois,
        annee=annee,
        montant=montant,
        date_versement=date_versement,
        type_document=doc_type,
        type_pdf=pdf_type,
        qualite_ocr=qualite
    )
    
    # 10. Vérifier les anomalies
    if not mois:
        decl.statut = "ATTENTION"
    if not montant:
        decl.statut = "ERREUR"
    
    return decl

def process_all_ir_files(input_dir: Path) -> List[DeclarationIR]:
    """Traite tous les fichiers IR"""
    declarations = []
    pdf_files = sorted(list(input_dir.glob("*.pdf")))
    
    if not pdf_files:
        print(f"\n❌ Aucun PDF trouvé dans {input_dir}")
        return []
    
    print(f"\n📁 {len(pdf_files)} fichier(s) IR à traiter:")
    
    for pdf in pdf_files:
        decl = process_ir_file(pdf)
        declarations.append(decl)
    
    return declarations

def generate_excel(declarations: List[DeclarationIR], societe: str, output_dir: Path):
    """Génère l'Excel avec les 3 colonnes demandées + colonne qualité pour scannés"""
    wb = Workbook()
    ws = wb.active
    ws.title = "IR"
    
    # En-têtes
    headers = ["Mois", "Montant IR (MAD)", "Date de versement", "Qualité OCR (%)"]
    ws.append(headers)
    
    # Style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Remplir les données (tri par année et mois)
    def get_sort_key(d):
        mois_order = {
            'Janvier': 1, 'Février': 2, 'Mars': 3, 'Avril': 4, 'Mai': 5, 'Juin': 6,
            'Juillet': 7, 'Août': 8, 'Septembre': 9, 'Octobre': 10, 'Novembre': 11, 'Décembre': 12
        }
        return (d.annee, mois_order.get(d.mois, 99))
    
    declarations_sorted = sorted(declarations, key=get_sort_key)
    
    for decl in declarations_sorted:
        mois_complet = f"{decl.mois} {decl.annee}" if decl.annee else decl.mois
        qualite = decl.qualite_ocr if decl.type_pdf == "scanne" else ""
        
        row = [
            mois_complet,
            decl.montant if decl.montant else "",
            decl.date_versement,
            qualite
        ]
        ws.append(row)
        
        row_num = ws.max_row
        
        # Colorer selon la qualité OCR (pour les scannés)
        if decl.type_pdf == "scanne" and decl.qualite_ocr > 0:
            qualite_cell = ws.cell(row=row_num, column=4)
            if decl.qualite_ocr >= 80:
                qualite_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Vert
            elif decl.qualite_ocr >= 60:
                qualite_cell.fill = PatternFill("solid", fgColor="FFEB9C")  # Jaune
            elif decl.qualite_ocr >= 40:
                qualite_cell.fill = PatternFill("solid", fgColor="FCE4D6")  # Orange
            else:
                qualite_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Rouge
        
        # Colorer le montant si manquant
        if not decl.montant:
            ws.cell(row=row_num, column=2).fill = PatternFill("solid", fgColor="FFC7CE")
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Sauvegarder
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    societe_clean = re.sub(r'[\\/*?:"<>|]', "_", societe)
    output_path = output_dir / f"IR_{societe_clean}_{timestamp}.xlsx"
    wb.save(output_path)
    
    return output_path

def main():
    print("\n" + "=" * 60)
    print("  EXTRACT-IR - Extraction IR (Mois, Montant, Date)")
    print("  Support: PDF natifs + PDF scannés (avec score de qualité)")
    print("=" * 60)
    
    print(f"\n📁 Dossier IR: {IR_DIR}")
    print(f"📁 Dossier output: {OUTPUT_DIR}")
    
    societe = input("\n🏢 Nom de la société: ").strip().upper()
    if not societe:
        societe = "SOCIETE"
    
    declarations = process_all_ir_files(IR_DIR)
    
    if declarations:
        # Filtrer pour garder seulement ceux qui ont un montant
        valides = [d for d in declarations if d.montant]
        
        if valides:
            output_path = generate_excel(valides, societe, OUTPUT_DIR)
            print(f"\n✅ EXCEL GÉNÉRÉ: {output_path}")
            
            total_montant = sum(d.montant for d in valides if d.montant)
            print(f"\n📊 RÉSUMÉ:")
            print(f"   Fichiers traités: {len(declarations)}")
            print(f"   Factures trouvées: {len(valides)}")
            print(f"   Montant total IR: {total_montant:,.2f} MAD")
            
            # Statistiques par type de PDF
            natifs = [d for d in valides if d.type_pdf == "natif"]
            scannes = [d for d in valides if d.type_pdf == "scanne"]
            if natifs:
                print(f"   PDF natifs: {len(natifs)}")
            if scannes:
                print(f"   PDF scannés: {len(scannes)}")
                # Statistiques de qualité OCR
                qualite_moyenne = sum(d.qualite_ocr for d in scannes) / len(scannes)
                print(f"   Qualité OCR moyenne: {qualite_moyenne:.1f}%")
                faibles = [d for d in scannes if d.qualite_ocr < 50]
                if faibles:
                    print(f"   ⚠️ {len(faibles)} fichier(s) scanné(s) avec qualité < 50%")
            
            # Afficher les fichiers sans montant
            sans_montant = [d for d in declarations if not d.montant]
            if sans_montant:
                print(f"\n⚠️ {len(sans_montant)} fichier(s) sans montant:")
                for d in sans_montant:
                    print(f"   - {d.fichier} ({d.type_pdf})" + (f" - Qualité: {d.qualite_ocr}%" if d.type_pdf == "scanne" else ""))
        else:
            print("\n⚠️ Aucun montant IR trouvé dans les fichiers.")
    else:
        print("\n⚠️ Aucun fichier traité.")
    
    input("\n✅ Appuyez sur Entrée pour quitter...")

if __name__ == "__main__":
    main()