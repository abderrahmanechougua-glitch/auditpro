#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FACTEXT V3 — Extraction Précise de Factures Marocaines
Amélioré après analyse de 7 factures réelles :
  - AFRIQUIA (scan RICOH, TVA 10%)
  - BEST PROFIL x3 (CamScanner, format espace+point)
  - EDIC (scan, format marocain standard)
  - GestamProjects (scan)
  - FIBMED CONSULTING (scan)
  - Green Line International (format anglo-saxon MAD)
  - Engineering Conseil Solution (scan)

V3 architecture changes:
  1. Patterns centralized in patterns.py  (PatternRule objects — single source of truth)
  2. TVA validation delegated to validators.py
  3. extract_best_match() replaces first-match: collects ALL matches, scores with context
  4. get_context_score() — line-level neighborhood analysis
  5. Debug mode: JSON candidate dump per field  (FACTEXT_DEBUG=1 or DEBUG_MODE=True)
  6. ML fallback placeholder when confidence < 70
"""

import os
import re
import sys
import json
import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Optional

# ============================================================
# CONFIGURATION — Détection automatique Tesseract / Poppler
# ============================================================
from pathlib import Path as _Path
import importlib.util as _ilu, sys as _sys

def _load_ocr_paths():
    """Charge core/ocr_paths.py depuis la racine de l'application (avec cache sys.modules)."""
    _mod_name = "core.ocr_paths"
    if _mod_name in _sys.modules:
        return _sys.modules[_mod_name]
    _here = _Path(__file__).resolve()
    for _parent in _here.parents:
        _candidate = _parent / "core" / "ocr_paths.py"
        if _candidate.exists():
            _spec = _ilu.spec_from_file_location(_mod_name, str(_candidate))
            _mod = _ilu.module_from_spec(_spec)
            _sys.modules[_mod_name] = _mod
            _spec.loader.exec_module(_mod)
            return _mod
    return None

_ocr = _load_ocr_paths()
if _ocr:
    TESSERACT_PATH = _ocr.TESSERACT_PATH
    POPPLER_PATH   = _ocr.POPPLER_PATH
else:
    TESSERACT_PATH = None
    POPPLER_PATH   = None

BASE_DIR = Path.home() / "AuditPro_output" / "factX"

import pdfplumber

try:
    import pytesseract
    from pdf2image import convert_from_path
    if TESSERACT_PATH:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
except ImportError:
    pytesseract = None
    convert_from_path = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_DIR  = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
try:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass

# ============================================================
# LOCAL IMPORTS — patterns.py and validators.py
# Loaded via importlib so the file works regardless of CWD / sys.path
# ============================================================

_MODULE_DIR = Path(__file__).resolve().parent

def _import_local(name: str, filepath: Path):
    if name in _sys.modules:
        return _sys.modules[name]
    spec = _ilu.spec_from_file_location(name, str(filepath))
    mod  = _ilu.module_from_spec(spec)
    _sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod

_pat = _import_local("extraction_factures.patterns",  _MODULE_DIR / "patterns.py")
_val = _import_local("extraction_factures.validators", _MODULE_DIR / "validators.py")

# Re-export for convenience
PatternRule       = _pat.PatternRule
AMOUNT_RE         = _pat.AMOUNT_RE
HT_KEYWORDS       = _pat.HT_KEYWORDS
TVA_KEYWORDS      = _pat.TVA_KEYWORDS
TTC_KEYWORDS      = _pat.TTC_KEYWORDS
RULES_NUM_FACTURE = _pat.RULES_NUM_FACTURE
RULES_DATE        = _pat.RULES_DATE
RULES_HT          = _pat.RULES_HT
RULES_TVA_MONTANT = _pat.RULES_TVA_MONTANT
RULES_TVA_TAUX    = _pat.RULES_TVA_TAUX
RULES_TTC         = _pat.RULES_TTC
RULES_FOURNISSEUR = _pat.RULES_FOURNISSEUR
RULES_ICE         = _pat.RULES_ICE
RULES_IF          = _pat.RULES_IF
RULES_TP          = _pat.RULES_TP
RULES_RC          = _pat.RULES_RC
RULES_CNSS        = _pat.RULES_CNSS
RULES_DESCRIPTION = _pat.RULES_DESCRIPTION

TVAControl   = _val.TVAControl
validate_tva = _val.validate_tva

# ============================================================
# DEBUG MODE
# ============================================================

DEBUG_MODE: bool = os.environ.get("FACTEXT_DEBUG", "").strip() == "1"
_debug_candidates: dict = {}   # Populated per-invoice; written to JSON if DEBUG_MODE


# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class ExtractionResult:
    value: str = ""
    confidence: int = 0
    method: str = ""

@dataclass
class Invoice:
    fichier_original: str = ""
    page_num: int = 0
    num_facture:   ExtractionResult = field(default_factory=ExtractionResult)
    date_facture:  ExtractionResult = field(default_factory=ExtractionResult)
    montant_ht:    ExtractionResult = field(default_factory=ExtractionResult)
    montant_tva:   ExtractionResult = field(default_factory=ExtractionResult)
    montant_ttc:   ExtractionResult = field(default_factory=ExtractionResult)
    taux_tva:      ExtractionResult = field(default_factory=ExtractionResult)
    description:   ExtractionResult = field(default_factory=ExtractionResult)
    fournisseur:   ExtractionResult = field(default_factory=ExtractionResult)
    ice:           ExtractionResult = field(default_factory=ExtractionResult)
    if_fiscal:     ExtractionResult = field(default_factory=ExtractionResult)
    tp_patente:    ExtractionResult = field(default_factory=ExtractionResult)
    rc:            ExtractionResult = field(default_factory=ExtractionResult)
    cnss:          ExtractionResult = field(default_factory=ExtractionResult)
    devise:        str = "MAD"
    montant_lettres: str = ""
    tva_control:   TVAControl = field(default_factory=TVAControl)
    anomalies:     List[str]  = field(default_factory=list)
    score_global:  int = 0


# ============================================================
# NORMALISATION DES MONTANTS
# ============================================================

def normalize_amount(amount_str: str) -> Optional[float]:
    """
    Convertit un montant texte en float.
    Gère les 4 formats rencontrés dans les factures marocaines :
      1. Marocain standard : 175 000,00  (espace=milliers, virgule=décimal)
      2. Best Profil :       1 749 118.68 (espace=milliers, point=décimal)
      3. Anglo-saxon (MAD) : 25,000.00    (virgule=milliers, point=décimal)
      4. Simple :            24000         (pas de séparateur)
    """
    if not amount_str:
        return None

    s = amount_str.strip()

    # Corriger les confusions OCR les plus fréquentes dans les montants
    s = s.replace("O", "0").replace("o", "0")
    s = s.replace("I", "1").replace("l", "1").replace("|", "1")
    s = s.replace("S", "5")

    s = s.replace("\u00A0", " ")
    s = re.sub(r"[A-Za-zÀ-ÿ()]", "", s)
    s = re.sub(r"\s+", " ", s)
    s = s.strip()

    if not s:
        return None

    nb_points   = s.count(".")
    nb_virgules = s.count(",")
    nb_espaces  = len(re.findall(r" ", s))

    if nb_espaces >= 1 and nb_virgules == 1 and nb_points == 0:
        s = s.replace(" ", "").replace(",", ".")
    elif nb_espaces >= 1 and nb_points == 1 and nb_virgules == 0:
        s = s.replace(" ", "")
    elif nb_espaces >= 0 and nb_virgules == 1 and nb_points >= 1 and s.rfind(",") > s.rfind("."):
        s = s.replace(" ", "").replace(".", "").replace(",", ".")
    elif nb_virgules >= 1 and nb_points == 1 and s.rfind(".") > s.rfind(","):
        s = s.replace(" ", "").replace(",", "")
    elif nb_points >= 2 and nb_virgules == 0:
        s = s.replace(" ", "").replace(".", "")
    elif nb_virgules == 1 and nb_points == 0 and nb_espaces == 0:
        s = s.replace(",", ".")
    else:
        s = s.replace(" ", "")

    s = re.sub(r"[^\d.\-]", "", s)
    s = s.strip(".")

    try:
        val = float(s)
        return val if val >= 0 else None
    except (ValueError, TypeError):
        return None


# ============================================================
# NORMALISATION DES DATES
# ============================================================

def normalize_date(date_str: str) -> str:
    if not date_str:
        return ""
    s = date_str.strip()
    s = s.replace("'", "/").replace("’", "/")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9/\-.]", "", s)
    m = re.match(r"(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})", s)
    if m:
        jj, mm, aaaa = m.groups()
        if len(aaaa) == 2:
            aaaa = "20" + aaaa
        return f"{int(jj):02d}/{int(mm):02d}/{aaaa}"
    m = re.match(r"(\d{4})[\/\-\.](\d{1,2})[\/\-\.](\d{1,2})", s)
    if m:
        aaaa, mm, jj = m.groups()
        return f"{int(jj):02d}/{int(mm):02d}/{aaaa}"
    return s


# ============================================================
# MONTANT EN LETTRES → NOMBRE
# ============================================================

def parse_french_words_to_number(text: str) -> Optional[float]:
    """
    Convertit un montant en lettres françaises en nombre.
    """
    if not text:
        return None
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    # Common OCR slips on words frequently seen in amount-in-words lines.
    text = text.replace("tente", "trente")
    text = text.replace("mile", "mille")
    text = text.replace("millc", "mille")
    text = text.replace("dirharns", "dirhams")
    text = text.replace("dirhams", "dirhams")
    parts = re.split(r"\b(?:dirham|dirhams|dh)\b", text, maxsplit=1)
    main_text     = parts[0].strip()
    centimes_text = parts[1].strip() if len(parts) > 1 else ""
    centimes_text = re.sub(r"\b(?:centime|centimes|cts?)\b", "", centimes_text).strip()

    def words_to_int(txt: str) -> int:
        if not txt.strip():
            return 0
        UNITS = {
            'zéro': 0, 'zero': 0, 'un': 1, 'une': 1, 'deux': 2, 'trois': 3,
            'quatre': 4, 'cinq': 5, 'six': 6, 'sept': 7, 'huit': 8, 'neuf': 9,
            'dix': 10, 'onze': 11, 'douze': 12, 'treize': 13, 'quatorze': 14,
            'quinze': 15, 'seize': 16, 'dix-sept': 17, 'dix-huit': 18, 'dix-neuf': 19,
            'vingt': 20, 'trente': 30, 'quarante': 40, 'cinquante': 50, 'soixante': 60,
        }
        txt = txt.replace("et ", " ")
        txt = re.sub(r"[\-–]", " ", txt)
        tokens  = txt.split()
        numbers = []
        i = 0
        while i < len(tokens):
            w = tokens[i]
            if w == "quatre" and i + 1 < len(tokens) and tokens[i + 1] in ("vingt", "vingts"):
                base = 80
                i += 2
                if i < len(tokens) and tokens[i] == "dix":
                    base = 90
                    i += 1
                    if i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 10:
                        base += UNITS[tokens[i]]; i += 1
                elif i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 20:
                    base += UNITS[tokens[i]]; i += 1
                numbers.append(("unit", base))
            elif w == "soixante":
                base = 60
                i += 1
                if i < len(tokens) and tokens[i] == "dix":
                    base = 70
                    i += 1
                    if i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 10:
                        base += UNITS[tokens[i]]; i += 1
                elif i < len(tokens) and tokens[i] in UNITS and UNITS[tokens[i]] < 20:
                    base += UNITS[tokens[i]]; i += 1
                numbers.append(("unit", base))
            elif w in ("cent", "cents"):
                numbers.append(("mult", 100)); i += 1
            elif w == "mille":
                numbers.append(("mult", 1000)); i += 1
            elif w in ("million", "millions"):
                numbers.append(("mult", 1_000_000)); i += 1
            elif w in ("milliard", "milliards"):
                numbers.append(("mult", 1_000_000_000)); i += 1
            elif w in UNITS:
                numbers.append(("unit", UNITS[w])); i += 1
            elif w.isdigit():
                numbers.append(("unit", int(w))); i += 1
            else:
                i += 1

        total = 0
        current = 0
        for typ, val in numbers:
            if typ == "unit":
                current += val
            elif typ == "mult":
                if val == 100:
                    current = current * 100 if current > 0 else 100
                elif val == 1000:
                    current = current * 1000 if current > 0 else 1000
                    total += current; current = 0
                elif val >= 1_000_000:
                    current = current * val if current > 0 else val
                    total += current; current = 0
        total += current
        return total

    main_amount  = words_to_int(main_text)
    cents_amount = words_to_int(centimes_text)
    result = main_amount + cents_amount / 100.0
    return result if result > 0 else None


# ============================================================
# SCORING ENGINE
# ============================================================

def get_context_score(line: str, keyword_list: list) -> int:
    """
    Return a bonus score (0 or 10) based on whether the line contains
    any keyword from keyword_list (case-insensitive match).
    """
    line_lower = line.lower()
    for kw in keyword_list:
        if kw in line_lower:
            return 10
    return 0


def extract_best_match(
    text: str,
    rules,
    field_name: str = "",
    context_keywords: list = None,
) -> ExtractionResult:
    """
    Collect ALL regex matches for all rules, score each with context-awareness,
    and return the highest-scoring result.

    Scoring formula:
        score = rule.confidence
              + context_bonus  (+10 if a nearby line contains a relevant keyword)
              + amount_bonus   (+5 if the matched value is a valid numeric amount)
              (capped at 100)

    When DEBUG_MODE is True, all candidates are stored in _debug_candidates[field_name].
    """
    if context_keywords is None:
        context_keywords = []

    lines      = text.split('\n')
    candidates: List[ExtractionResult] = []

    for rule in rules:
        for m in re.finditer(rule.pattern, text, re.IGNORECASE | re.MULTILINE):
            raw = m.group(1)
            if raw is None:
                continue
            value = re.sub(r"[\|\s]+$", "", raw).strip()
            value = re.sub(r"\s+", " ", value).strip()   # collapse internal newlines
            if not value:
                continue

            line_idx = text[: m.start()].count('\n')

            # Context bonus: scan ±2 surrounding lines
            context_bonus = 0
            if context_keywords:
                for ctx_line in lines[max(0, line_idx - 2): line_idx + 3]:
                    context_bonus = max(
                        context_bonus,
                        get_context_score(ctx_line, context_keywords),
                    )

            amount_bonus = 5 if normalize_amount(value) is not None else 0
            score = min(rule.confidence + context_bonus + amount_bonus, 100)

            candidates.append(
                ExtractionResult(value=value, confidence=score, method=rule.pattern[:40])
            )

    if DEBUG_MODE and field_name:
        _debug_candidates[field_name] = [
            {"value": c.value, "confidence": c.confidence, "method": c.method}
            for c in candidates
        ]

    if not candidates:
        return ExtractionResult()
    return max(candidates, key=lambda r: r.confidence)


def _ml_fallback_extract(text: str, field_name: str) -> ExtractionResult:
    """
    Placeholder for ML-based extraction when regex confidence < 70.
    TODO: plug in a trained model (e.g., LayoutLM, Donut, fine-tuned DistilBERT).
    Returns empty ExtractionResult until a model is integrated.
    """
    return ExtractionResult()


def _extract_field(
    text: str,
    rules,
    field_name: str,
    context_keywords: list = None,
) -> ExtractionResult:
    """Run extract_best_match, then fall back to ML placeholder if confidence < 70."""
    result = extract_best_match(text, rules, field_name, context_keywords)
    if result.confidence < 70:
        ml = _ml_fallback_extract(text, field_name)
        if ml.value and ml.confidence > result.confidence:
            return ml
    return result


def _normalize_ocr_text(text: str) -> str:
    """
    Lightweight OCR cleanup used only as a fallback extraction pass.
    The goal is to improve noisy scans without overfitting to one supplier template.
    """
    s = text or ""
    s = s.replace("\u00A0", " ")
    s = s.replace("\u2009", " ")
    s = s.replace("\u202f", " ")
    # Common OCR punctuation noise around labels/totals.
    s = re.sub(r"[|]{2,}", "|", s)
    s = re.sub(r"[_]{2,}", " ", s)
    # Label normalization (broad, but safe for extraction contexts).
    s = re.sub(r"(?i)\bT\s*\.?\s*V\s*\.?\s*A\b", "TVA", s)
    s = re.sub(r"(?i)\bT\s*\.?\s*T\s*\.?\s*C\b", "TTC", s)
    s = re.sub(r"(?i)\bF\s*\.?\s*T\s*\.?\s*C\b", "TTC", s)
    s = re.sub(r"(?i)\bH\s*\.?\s*T\b", "HT", s)
    s = re.sub(r"(?i)\bFACTURE\s+N\s*:?", "FACTURE N° ", s)
    # Normalize spacing around separators to help regex matching.
    s = re.sub(r"\s*:\s*", ": ", s)
    s = re.sub(r"[ \t]+", " ", s)
    return s


def _pick_better(primary: ExtractionResult, fallback: ExtractionResult) -> ExtractionResult:
    """Pick the best extraction result while favoring a non-empty value."""
    if fallback.value and (not primary.value or fallback.confidence > primary.confidence):
        return fallback
    return primary


# ============================================================
# FOURNISSEUR — special multi-strategy extraction
# ============================================================

def extract_fournisseur(text: str) -> ExtractionResult:
    """Extraction du fournisseur avec logique spéciale."""
    result = extract_best_match(text, RULES_FOURNISSEUR, "fournisseur")
    if result.value:
        return result

    # Fallback: look for an ALL-CAPS line in the first 10 lines
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    STOP_WORDS = {"FACTURE", "INVOICE", "DATE", "TOTAL", "N°", "CLIENT", "DESTINATAIRE",
                  "LIVRÉ", "FACTURÉ", "TVA", "HT", "TTC", "COMPTABILISÉ", "COMPTABILISE",
                  "PAGE", "DESIGNATION", "MONTANT", "PROJET", "OBJET", "PAYÉ", "PAYE"}

    for line in lines[:10]:
        if len(line) < 4 or len(line) > 80:
            continue
        upper_ratio = sum(1 for c in line if c.isupper()) / max(len(line), 1)
        if upper_ratio < 0.5:
            continue
        words = line.upper().split()
        if any(w in STOP_WORDS for w in words):
            continue
        if re.match(r"^\d+$", line):
            continue
        return ExtractionResult(value=line[:60], confidence=50, method="header_line")

    return ExtractionResult()


# ============================================================
# MONTANT EN LETTRES → TTC (cross-check)
# ============================================================

def extract_ttc_with_words(text: str) -> Tuple[Optional[float], str]:
    """Extrait le TTC depuis le montant en lettres."""
    m = re.search(
        r"Arr[êeé]t[ée]e?\s+(?:la\s+pr[ée]sente\s+facture\s+)?[àa]\s+la\s+(?:somme|valeur)\s+de\s*:?\s*(.+?)(?:\.|\n|$)",
        text, re.IGNORECASE
    )
    if m:
        words_text = m.group(1).strip()
        num = parse_french_words_to_number(words_text)
        if num and num > 0:
            return num, words_text

    m = re.search(
        r"Arr[êeé]t[ée]e?\s+[àa]\s+la\s+valeur\s+de\s*:?\s*(.+?)(?:DH|DIRHAMS?|MAD|\.|\n|$)",
        text, re.IGNORECASE
    )
    if m:
        words_text = m.group(1).strip()
        num = parse_french_words_to_number(words_text)
        if num and num > 0:
            return num, words_text

    # OCR fallback: amount-in-words often starts on next line after "somme de"
    m = re.search(
        r"somme\s+de\s*:?\s*\n?\s*([A-Za-zÀ-ÿ\-\s]{8,160})(?:DH|DIRHAMS?|MAD|\.|\n|$)",
        text,
        re.IGNORECASE,
    )
    if m:
        words_text = re.sub(r"\s+", " ", m.group(1)).strip()
        num = parse_french_words_to_number(words_text)
        if num and num > 0:
            return num, words_text

    return None, ""


# ============================================================
# CONTEXTUAL AMOUNTS — table block detection
# ============================================================

def extract_amounts_context(text: str) -> dict:
    """
    Look for a structured HT / TVA / TTC table block and return
    raw string values if found.
    """
    amounts = {}
    table_pattern = re.compile(
        rf"(?:Total|Montant|Sous[- ]?total)\s+H\.?T\.?\s*[:\s]*{AMOUNT_RE}"
        r".*?"
        rf"T\.?V\.?A\.?\s*(?:\(?\s*\d+\s*%?\s*\)?)?\s*[:\s]*{AMOUNT_RE}"
        r".*?"
        rf"(?:Total|Montant)\s+(?:G[ée]n[ée]ral|T\.?T\.?C\.?)\s*[:\s]*{AMOUNT_RE}",
        re.IGNORECASE | re.DOTALL
    )
    m = table_pattern.search(text)
    if m:
        amounts["ht_table"]  = m.group(1).strip()
        amounts["tva_table"] = m.group(2).strip()
        amounts["ttc_table"] = m.group(3).strip()
        return amounts

    amount_pattern = re.compile(r"\d[\d\s']{0,15}[.,]\d{2,5}")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]

    for idx, line in enumerate(lines):
        lower = line.lower()
        if not line:
            continue

        if re.search(r"(?:montant\s+total\s+h\.?t\.?|montant\s+h\.?t\.?|total\s+h\.?t\.?|sous[- ]?total)", lower):
            line_amounts = amount_pattern.findall(line)
            if line_amounts:
                amounts["ht_table"] = line_amounts[-1].strip()
                for look_ahead in lines[idx + 1: idx + 5]:
                    if not look_ahead:
                        continue
                    ahead_lower = look_ahead.lower()
                    ahead_amounts = amount_pattern.findall(look_ahead)
                    if not ahead_amounts:
                        continue
                    if "%" in look_ahead and "tva" in ahead_lower and "tva_table" not in amounts:
                        amounts["tva_table"] = ahead_amounts[-1].strip()
                        continue
                    if "ttc_table" in amounts:
                        continue
                    if re.fullmatch(r"(?:[0-9\s'.,]+)(?:\s*(?:dh|mad))?", ahead_lower):
                        amounts["ttc_table"] = ahead_amounts[-1].strip()
                        continue
                    if any(token in ahead_lower for token in ("total ttc", "net a payer", "net apayer", "total général", "total general")):
                        amounts["ttc_table"] = ahead_amounts[-1].strip()
                if amounts:
                    return amounts

        if ("total ht" in lower or "montant" in lower) and ("total ttc" in lower or "net a payer" in lower or "net apayer" in lower):
            for summary_line in lines[idx + 1: idx + 3]:
                if not re.search(r"\d{1,2}\s*%", summary_line):
                    continue
                percent_match = re.search(r"\d{1,2}\s*%", summary_line)
                matches = list(amount_pattern.finditer(summary_line))
                after_percent = [m.group(0).strip() for m in matches if m.start() > percent_match.end()]
                if len(after_percent) >= 2:
                    amounts["tva_table"] = after_percent[0]
                    amounts["ht_table"] = after_percent[1]
                    ttc_candidates = [normalize_amount(value) for value in after_percent[2:]]
                    ttc_candidates = [value for value in ttc_candidates if value is not None]
                    if ttc_candidates:
                        amounts["ttc_table"] = f"{max(ttc_candidates):.2f}"
                    return amounts
    return amounts


def infer_ht_from_line_items(text: str) -> Optional[ExtractionResult]:
    """
    Infer HT from OCR table rows that contain a VAT rate and a trailing line total.

    Typical examples:
      - product qty unit_price 20% total
      - fuel volume unit_price 10% total DH

    This is a fallback only, used when labeled HT extraction failed.
    """
    amount_pattern = re.compile(r"\d[\d\s']{0,15}[.,]\d{2,5}")
    total_candidates = []

    for line in text.splitlines():
        clean = re.sub(r"\s+", " ", line).strip()
        if not clean:
            continue

        lower = clean.lower()
        if "tva" in lower or "total ttc" in lower or "net a payer" in lower:
            continue
        if not re.search(r"\d{1,2}\s*%", clean):
            continue

        amounts = amount_pattern.findall(clean)
        if len(amounts) < 2:
            continue

        inferred_total = normalize_amount(amounts[-1])
        if inferred_total is None or inferred_total <= 0:
            continue

        total_candidates.append(inferred_total)

    if not total_candidates:
        return None

    inferred_ht = round(sum(total_candidates), 2)
    return ExtractionResult(value=f"{inferred_ht:.2f}", confidence=76, method="line_items_sum")


# ============================================================
# MAIN EXTRACTION
# ============================================================

def extract_all_fields(text: str, extract_juridical: bool = True) -> Invoice:
    """Extraction complète de tous les champs d'une facture."""
    inv = Invoice()
    global _debug_candidates
    _debug_candidates = {}

    ocr_text = _normalize_ocr_text(text)

    # === NUMÉRO DE FACTURE ===
    inv.num_facture = _extract_field(text, RULES_NUM_FACTURE, "num_facture")
    inv.num_facture = _pick_better(
        inv.num_facture,
        _extract_field(ocr_text, RULES_NUM_FACTURE, "num_facture_ocr"),
    )
    if inv.num_facture.value:
        inv.num_facture.value = re.sub(
            r"\s+(?:ICE|I\.?F\.?|RC|TP|PATENTE|CNSS|DATE|CLIENT|COMMANDE|BON\s+DE\s+COMMANDE|LIVRE\s+A)\b.*$",
            "",
            inv.num_facture.value,
            flags=re.IGNORECASE,
        ).strip(" .-|_:")

    # === DATE ===
    date_result = _extract_field(text, RULES_DATE, "date_facture")
    date_result = _pick_better(
        date_result,
        _extract_field(ocr_text, RULES_DATE, "date_facture_ocr"),
    )
    if date_result.value:
        date_result.value = normalize_date(date_result.value)
    inv.date_facture = date_result

    # === MONTANTS — table block has priority ===
    table_amounts = extract_amounts_context(text)
    table_amounts_ocr = extract_amounts_context(ocr_text)

    # HT
    if "ht_table" in table_amounts:
        inv.montant_ht = ExtractionResult(value=table_amounts["ht_table"], confidence=95, method="table_block")
    elif "ht_table" in table_amounts_ocr:
        inv.montant_ht = ExtractionResult(value=table_amounts_ocr["ht_table"], confidence=92, method="table_block_ocr")
    else:
        inv.montant_ht = _extract_field(text, RULES_HT, "montant_ht", HT_KEYWORDS)
        inv.montant_ht = _pick_better(
            inv.montant_ht,
            _extract_field(ocr_text, RULES_HT, "montant_ht_ocr", HT_KEYWORDS),
        )
        inferred_ht = infer_ht_from_line_items(ocr_text)
        if inferred_ht:
            inv.montant_ht = _pick_better(inv.montant_ht, inferred_ht)

    # TVA montant
    if "tva_table" in table_amounts:
        inv.montant_tva = ExtractionResult(value=table_amounts["tva_table"], confidence=95, method="table_block")
    elif "tva_table" in table_amounts_ocr:
        inv.montant_tva = ExtractionResult(value=table_amounts_ocr["tva_table"], confidence=92, method="table_block_ocr")
    else:
        inv.montant_tva = _extract_field(text, RULES_TVA_MONTANT, "montant_tva", TVA_KEYWORDS)
        inv.montant_tva = _pick_better(
            inv.montant_tva,
            _extract_field(ocr_text, RULES_TVA_MONTANT, "montant_tva_ocr", TVA_KEYWORDS),
        )

    # TVA taux
    inv.taux_tva = _extract_field(text, RULES_TVA_TAUX, "taux_tva", TVA_KEYWORDS)
    inv.taux_tva = _pick_better(
        inv.taux_tva,
        _extract_field(ocr_text, RULES_TVA_TAUX, "taux_tva_ocr", TVA_KEYWORDS),
    )

    # TTC
    if "ttc_table" in table_amounts:
        inv.montant_ttc = ExtractionResult(value=table_amounts["ttc_table"], confidence=95, method="table_block")
    elif "ttc_table" in table_amounts_ocr:
        inv.montant_ttc = ExtractionResult(value=table_amounts_ocr["ttc_table"], confidence=92, method="table_block_ocr")
    else:
        inv.montant_ttc = _extract_field(text, RULES_TTC, "montant_ttc", TTC_KEYWORDS)
        inv.montant_ttc = _pick_better(
            inv.montant_ttc,
            _extract_field(ocr_text, RULES_TTC, "montant_ttc_ocr", TTC_KEYWORDS),
        )

    # If TTC likely came from a words-based OCR match, parse it into a numeric TTC.
    ttc_looks_words = bool(re.search(r"[A-Za-zÀ-ÿ]", inv.montant_ttc.value or ""))
    if inv.montant_ttc.method.startswith("Arr[") or ttc_looks_words:
        words_text = inv.montant_ttc.value
        num = parse_french_words_to_number(words_text)
        ht_guess = normalize_amount(inv.montant_ht.value)
        # Accept words-based TTC only when plausible against HT (if HT exists).
        if num and (ht_guess is None or num >= ht_guess):
            inv.montant_lettres = words_text
            inv.montant_ttc = ExtractionResult(value=f"{num:.2f}", confidence=95, method="words_to_number")
        else:
            inv.montant_lettres = words_text
            inv.montant_ttc = _extract_field(text, RULES_TTC[1:], "montant_ttc_fallback", TTC_KEYWORDS)

    # Cross-check via words amount
    ttc_words, words_text = extract_ttc_with_words(text)
    if ttc_words and not inv.montant_lettres:
        inv.montant_lettres = words_text
    if ttc_words and not normalize_amount(inv.montant_ttc.value):
        inv.montant_ttc = ExtractionResult(value=f"{ttc_words:.2f}", confidence=78, method="words_fallback")

    # === DESCRIPTION ===
    inv.description = _extract_field(text, RULES_DESCRIPTION, "description")

    # === FOURNISSEUR ===
    inv.fournisseur = extract_fournisseur(text)

    # === DONNÉES JURIDIQUES ===
    if extract_juridical:
        inv.ice        = _extract_field(text, RULES_ICE,  "ice")
        inv.if_fiscal  = _extract_field(text, RULES_IF,   "if_fiscal")
        inv.tp_patente = _extract_field(text, RULES_TP,   "tp_patente")
        inv.rc         = _extract_field(text, RULES_RC,   "rc")
        inv.cnss       = _extract_field(text, RULES_CNSS, "cnss")

    # === CONTRÔLE TVA — delegated to validators.py ===
    ht_val   = normalize_amount(inv.montant_ht.value)
    ttc_val  = normalize_amount(inv.montant_ttc.value)
    tva_val  = normalize_amount(inv.montant_tva.value)
    taux_str = inv.taux_tva.value.replace(",", ".") if inv.taux_tva.value else None
    taux_doc = float(taux_str) if taux_str else None

    inv.tva_control = validate_tva(ht_val, tva_val, ttc_val, taux_doc, ttc_words)

    if not inv.montant_tva.value and inv.tva_control.tva_calculee is not None:
        inv.montant_tva = ExtractionResult(
            value=f"{inv.tva_control.tva_calculee:.2f}",
            confidence=65,
            method="calculated_from_ht_ttc",
        )

    # If TTC was computed by validate_tva, expose it on the invoice
    if inv.tva_control.statut == "TTC_CALCULE" and inv.tva_control.ttc:
        inv.montant_ttc = ExtractionResult(
            value=f"{inv.tva_control.ttc:.2f}", confidence=70, method="calculated"
        )

    # Cross-check anomaly (words vs numeric)
    if ttc_words and ttc_val and abs(ttc_words - ttc_val) > 1.0:
        inv.anomalies.append(f"Lettres ({ttc_words:.2f}) ≠ TTC ({ttc_val:.2f})")

    # === ANOMALIES ===
    if not inv.num_facture.value:
        inv.anomalies.append("N° facture manquant")
    if not inv.date_facture.value:
        inv.anomalies.append("Date manquante")
    if not inv.montant_ttc.value:
        inv.anomalies.append("TTC manquant")
    if not inv.montant_ht.value:
        inv.anomalies.append("HT manquant")
    if not inv.fournisseur.value:
        inv.anomalies.append("Fournisseur non identifié")
    if extract_juridical and not inv.ice.value:
        inv.anomalies.append("ICE manquant")

    # === SCORE GLOBAL ===
    fields_to_score = [inv.num_facture, inv.date_facture, inv.montant_ht, inv.montant_tva, inv.montant_ttc]
    scores = [f.confidence for f in fields_to_score if f.value]
    inv.score_global = sum(scores) // max(len(scores), 1) if scores else 0

    # === DEBUG OUTPUT ===
    if DEBUG_MODE:
        _save_debug_candidates(inv)

    return inv


def _save_debug_candidates(inv: Invoice):
    """Write per-field extraction candidates to a JSON file in OUTPUT_DIR."""
    try:
        Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
        fname    = f"DEBUG_CANDIDATES_{inv.fichier_original or 'unknown'}_p{inv.page_num}.json"
        out_path = Path(OUTPUT_DIR) / fname
        payload  = {
            "file": inv.fichier_original,
            "page": inv.page_num,
            "selected": {
                "num_facture":  inv.num_facture.value,
                "date_facture": inv.date_facture.value,
                "montant_ht":   inv.montant_ht.value,
                "montant_tva":  inv.montant_tva.value,
                "montant_ttc":  inv.montant_ttc.value,
            },
            "candidates": _debug_candidates,
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ============================================================
# PDF PROCESSING
# ============================================================

def get_pdf_text(pdf_path: Path) -> Tuple[int, List[str]]:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        if any(len(t.strip()) > 50 for t in pages_text):
            print(f"   ✅ {len(pages_text)} page(s) extraites (PDF natif)")
            return len(pages_text), pages_text
    except Exception as e:
        print(f"   ⚠️ pdfplumber: {e}")

    if pytesseract is None or convert_from_path is None:
        print("   ⚠️ OCR non disponible (Tesseract/pdf2image absent). PDF scanné ignoré.")
        return 0, []
    try:
        if TESSERACT_PATH:
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        images     = convert_from_path(str(pdf_path), dpi=250, poppler_path=POPPLER_PATH or None)
        pages_text = []
        for i, img in enumerate(images):
            print(f"      OCR page {i+1}/{len(images)}...", end="\r")
            text = pytesseract.image_to_string(img, lang="fra+eng", config="--psm 6 --oem 3")
            pages_text.append(text)
        print(f"   ✅ {len(images)} page(s) traitées (OCR)" + " " * 20)
        return len(images), pages_text
    except Exception as e:
        print(f"   ❌ Erreur OCR: {e}")
        return 0, []


def detect_invoice_boundaries(pages_text: List[str]) -> List[Tuple[int, int]]:
    n = len(pages_text)
    if n == 0:
        return []
    if n == 1:
        return [(0, 0)]

    starts = []
    for i in range(n):
        text = pages_text[i]
        has_facture    = bool(re.search(r"\b(FACTURE|Invoice)\b", text, re.I))
        has_num        = bool(re.search(r"(?:N[°º]\s*(?:FACTURE|Client)|NUM[ÉE]RO|Facture\s+N)", text, re.I))
        has_new_total  = bool(re.search(r"(?:Total\s+H\.?T|Montant\s+H\.?T)", text, re.I))

        if i == 0:
            starts.append(i)
        elif has_facture and has_num:
            starts.append(i)
        elif has_facture and has_new_total and i > 0:
            prev_has_ttc = bool(re.search(r"(?:T\.?T\.?C|Net\s+[àa]\s+payer|Total\s+G)", pages_text[i - 1], re.I))
            if prev_has_ttc:
                starts.append(i)

    if len(starts) <= 1:
        all_have_structure = all(
            re.search(r"(?:N[°º]\s*FACTURE|Total|TTC|Montant)", pages_text[i], re.I)
            for i in range(n)
        )
        if all_have_structure and n > 1:
            return [(i, i) for i in range(n)]
        else:
            return [(0, n - 1)]

    boundaries = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else n - 1
        boundaries.append((start, end))
    return boundaries


def process_pdf(pdf_path: Path, extract_juridical: bool) -> List[Invoice]:
    print(f"\n📄 {pdf_path.name}")
    pages_count, pages_text = get_pdf_text(pdf_path)
    if pages_count == 0:
        return []

    try:
        Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
        debug_path = Path(OUTPUT_DIR) / f"DEBUG_OCR_{pdf_path.stem}.txt"
        with open(debug_path, "w", encoding="utf-8") as f:
            for i, t in enumerate(pages_text):
                f.write(f"\n{'='*60}\nPAGE {i+1}\n{'='*60}\n{t}\n")
    except Exception:
        pass

    boundaries = detect_invoice_boundaries(pages_text)
    print(f"   🔍 {len(boundaries)} facture(s) détectée(s)")

    invoices = []
    for idx, (start, end) in enumerate(boundaries):
        combined_text = "\n\n".join(pages_text[start: end + 1])

        inv = extract_all_fields(combined_text, extract_juridical)
        inv.fichier_original = pdf_path.name
        inv.page_num         = start + 1

        stem_supplier = re.sub(r"[_\-]+", " ", pdf_path.stem).strip()
        current_supplier = re.sub(r"[^a-z0-9]", "", (inv.fournisseur.value or "").lower())
        filename_supplier = re.sub(r"[^a-z0-9]", "", stem_supplier.lower())
        if stem_supplier and (
            not inv.fournisseur.value
            or "facture" in (inv.fournisseur.value or "").lower()
            or (current_supplier and filename_supplier.endswith(current_supplier) and len(filename_supplier) > len(current_supplier))
        ):
            inv.fournisseur = ExtractionResult(value=stem_supplier, confidence=48, method="filename")

        num    = inv.num_facture.value or "?"
        ttc    = inv.montant_ttc.value or "?"
        date   = inv.date_facture.value or "?"
        four   = inv.fournisseur.value or "?"
        status = inv.tva_control.statut
        print(f"      #{idx+1} | P.{start+1} | {four[:20]} | N°{num} | TTC:{ttc} | {date} | TVA:{status}")

        if inv.anomalies:
            print(f"         ⚠️ {', '.join(inv.anomalies)}")

        invoices.append(inv)

    return invoices


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_excel(invoices: List[Invoice], output_path: Path, extract_juridical: bool):
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"

    headers = [
        "Fichier", "Page", "Fournisseur", "N° Facture", "Date",
        "Montant HT", "Taux TVA", "Montant TVA", "Montant TTC",
        "TVA Calculée", "Taux Calculé", "Statut TVA", "Message TVA",
        "Montant Lettres", "Description",
        "Score %", "Anomalies"
    ]
    if extract_juridical:
        headers.extend(["ICE", "IF", "TP/Patente", "RC", "CNSS"])

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    FILL_OK    = PatternFill("solid", fgColor="C6EFCE")
    FILL_WARN  = PatternFill("solid", fgColor="FFEB9C")
    FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")
    FILL_GREY  = PatternFill("solid", fgColor="D9D9D9")
    FONT_DATA  = Font(name="Arial", size=9)

    for inv in invoices:
        row_data = [
            inv.fichier_original,
            inv.page_num,
            inv.fournisseur.value,
            inv.num_facture.value,
            inv.date_facture.value,
            normalize_amount(inv.montant_ht.value)  if inv.montant_ht.value  else None,
            inv.taux_tva.value + "%" if inv.taux_tva.value else "",
            normalize_amount(inv.montant_tva.value) if inv.montant_tva.value else None,
            normalize_amount(inv.montant_ttc.value) if inv.montant_ttc.value else None,
            inv.tva_control.tva_calculee,
            f"{inv.tva_control.taux:.2f}%" if inv.tva_control.taux else "",
            inv.tva_control.statut,
            inv.tva_control.message,
            inv.montant_lettres[:80] if inv.montant_lettres else "",
            inv.description.value[:100] if inv.description.value else "",
            f"{inv.score_global}%",
            "; ".join(inv.anomalies) if inv.anomalies else "OK"
        ]
        if extract_juridical:
            row_data.extend([
                inv.ice.value, inv.if_fiscal.value, inv.tp_patente.value,
                inv.rc.value,  inv.cnss.value
            ])

        ws.append(row_data)
        row_num = ws.max_row

        for col_idx in [6, 8, 9, 10]:
            cell = ws.cell(row=row_num, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

        status_cell = ws.cell(row=row_num, column=12)
        if inv.tva_control.statut == "OK":
            status_cell.fill = FILL_OK
        elif inv.tva_control.statut in ("TVA_MANQUANTE", "TTC_CALCULE"):
            status_cell.fill = FILL_WARN
        elif "ANOMALIE" in inv.tva_control.statut:
            status_cell.fill = FILL_ERROR
        else:
            status_cell.fill = FILL_GREY

        anomaly_cell = ws.cell(row=row_num, column=17)
        if inv.anomalies:
            anomaly_cell.fill = FILL_WARN
            anomaly_cell.font = Font(name="Arial", size=9, color="9C5700")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            if not cell.font or cell.font == Font():
                cell.font = FONT_DATA
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\n✅ EXCEL → {output_path}")


# ============================================================
# INTERACTIVE MENU
# ============================================================

def show_menu():
    print("\n" + "=" * 70)
    print("  FACTEXT V3 — Extraction Précise de Factures Marocaines")
    print("  Supporte : AFRIQUIA, BEST PROFIL, EDIC, GestamProjects,")
    print("             FIBMED, Green Line, Engineering Conseil, etc.")
    print("=" * 70)
    print("\n📊 Données toujours extraites :")
    print("   ✓ N° facture, Date, Fournisseur, Description")
    print("   ✓ HT, TVA (taux + montant), TTC")
    print("   ✓ Contrôle TVA (HT+TVA=TTC, taux cohérent)")
    print("   ✓ Montant en lettres (cross-check)")
    print("\n⚖️ Données juridiques (optionnelles) :")
    print("   • ICE, IF, TP/Patente, RC, CNSS")
    if DEBUG_MODE:
        print("\n🔍 Mode DEBUG actif — candidats JSON dans output/")

    while True:
        choix = input("\nExtraire les données juridiques ? (O/N) [O]: ").strip().upper()
        if choix in ('O', 'N', ''):
            return choix != 'N'
        print("❌ Répondez O ou N")


def main():
    extract_juridical = show_menu()

    pdf_files = sorted(INPUT_DIR.glob("*.pdf"))
    if not pdf_files:
        print(f"\n❌ Aucun PDF dans {INPUT_DIR}")
        print(f"   Placez vos factures PDF dans : {INPUT_DIR}")
        input("\nAppuyez sur Entrée...")
        return

    print(f"\n📁 {len(pdf_files)} fichier(s) PDF :")
    for i, pdf in enumerate(pdf_files, 1):
        size_mb = pdf.stat().st_size / 1024 / 1024
        print(f"   {i}. {pdf.name} ({size_mb:.1f} MB)")

    print(f"\nOptions : T=Tous, ou numéros séparés par virgule (ex: 1,3)")
    choix = input("Votre choix [T]: ").strip().upper()

    if choix and choix != 'T':
        selected = []
        for c in choix.split(','):
            c = c.strip()
            if c.isdigit() and 1 <= int(c) <= len(pdf_files):
                selected.append(pdf_files[int(c) - 1])
        if selected:
            pdf_files = selected

    all_invoices = []
    for pdf in pdf_files:
        invoices = process_pdf(pdf, extract_juridical)
        all_invoices.extend(invoices)

    if all_invoices:
        timestamp  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix     = "Complet" if extract_juridical else "Base"
        output_path = OUTPUT_DIR / f"Factures_{suffix}_{timestamp}.xlsx"
        generate_excel(all_invoices, output_path, extract_juridical)

        print(f"\n{'='*50}")
        print(f"📊 RÉSUMÉ")
        print(f"{'='*50}")
        print(f"   Factures traitées : {len(all_invoices)}")
        print(f"   Avec N° facture   : {sum(1 for i in all_invoices if i.num_facture.value)}/{len(all_invoices)}")
        print(f"   Avec Date         : {sum(1 for i in all_invoices if i.date_facture.value)}/{len(all_invoices)}")
        print(f"   Avec HT           : {sum(1 for i in all_invoices if i.montant_ht.value)}/{len(all_invoices)}")
        print(f"   Avec TTC          : {sum(1 for i in all_invoices if i.montant_ttc.value)}/{len(all_invoices)}")
        print(f"   Avec Fournisseur  : {sum(1 for i in all_invoices if i.fournisseur.value)}/{len(all_invoices)}")
        print(f"   TVA OK            : {sum(1 for i in all_invoices if i.tva_control.statut == 'OK')}/{len(all_invoices)}")
        print(f"   TVA ANOMALIE      : {sum(1 for i in all_invoices if 'ANOMALIE' in i.tva_control.statut)}/{len(all_invoices)}")
        print(f"   Score moyen       : {sum(i.score_global for i in all_invoices) // len(all_invoices)}%")
        print(f"\n   Fichier Excel     : {output_path.name}")
        print(f"   Debug OCR         : {OUTPUT_DIR}/DEBUG_OCR_*.txt")
        if DEBUG_MODE:
            print(f"   Debug candidats   : {OUTPUT_DIR}/DEBUG_CANDIDATES_*.json")
    else:
        print("\n❌ Aucune facture extraite.")

    input("\n✅ Terminé ! Appuyez sur Entrée...")


if __name__ == "__main__":
    main()
