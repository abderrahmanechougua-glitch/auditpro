#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXTRACT-CNSS V2 — Extraction des Bordereaux CNSS

Structure du bordereau CNSS (2 pages par fichier) :
  Page 1 — Régime Général :
    C1: Allocations Familiales  | Masse salariale | 6.40%  | Montant AF
    C2: Prestations Sociales    | Masse salariale | 13.46% | Montant PS
    C3: Total cotisations       |                          | AF + PS
    C4: Pénalités               |                          | 0.00
    C8: Taxe Formation Pro      | Masse salariale | 1.60%  | Montant TFP
    C9: Pénalités TFP           |                          | 0.00
    C10: Montant global         |                          | Total page 1

  Page 2 — AMO :
    C1: Participation AMO       | Masse salariale | 1.85%  | Montant
    C2: Cotisation AMO          | Masse salariale | 4.52%  | Montant
    C3: Total AMO               |                          | Part + Cot
    C4: Pénalités AMO           |                          | 0.00
    C10: Montant global AMO     |                          | Total page 2
"""

import os, re, sys, datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

# ============================================================
# CONFIGURATION
# ============================================================
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH = r"C:\poppler\poppler-25.12.0\Library\bin"
BASE_DIR = Path.home() / "AuditPro_output" / "cnss"
CNSS_DIR = BASE_DIR / "CNSS"
OUTPUT_DIR = BASE_DIR / "output"
try:
    CNSS_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass
os.environ["PATH"] = os.environ["PATH"] + os.pathsep + r"C:\Program Files\Tesseract-OCR"

def install_pkg(p):
    os.system(f"{sys.executable} -m pip install {p}")

try: import pdfplumber
except ImportError: install_pkg("pdfplumber"); import pdfplumber

try:
    import pytesseract; from pdf2image import convert_from_path
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
except ImportError:
    install_pkg("pytesseract"); install_pkg("pdf2image")
    import pytesseract; from pdf2image import convert_from_path
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    install_pkg("openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

MOIS = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
         7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}

# ============================================================
# PARSE AMOUNT
# ============================================================
def parse_amount(s: str) -> Optional[float]:
    if not s: return None
    s = s.strip().replace('\u00A0','').replace(' ','')
    # Anglo-saxon: 86,111.10
    if re.match(r'^\d{1,3}(,\d{3})*\.\d{2}$', s):
        return float(s.replace(',',''))
    # Européen: 86 111,10
    s2 = s.replace(',','.')
    try: return float(s2)
    except: pass
    s3 = re.sub(r'[^\d.\-]','',s)
    try: return float(s3)
    except: return None

AMOUNT_RE = re.compile(r'(\d{1,3}(?:,\d{3})*\.\d{2})')

def find_amounts(text: str) -> List[float]:
    return [parse_amount(m) for m in AMOUNT_RE.findall(text) if parse_amount(m) is not None]

# ============================================================
# DATACLASS
# ============================================================
@dataclass
class BordereauCNSS:
    fichier: str = ""
    mois_num: int = 0
    annee: int = 0
    raison_sociale: str = ""
    n_affilie: str = ""

    # Page 1 — Régime Général
    masse_salariale_af: float = 0.0   # Masse salariale AF (peut différer de PS à cause du plafond)
    masse_salariale_ps: float = 0.0   # Masse salariale PS (plafonnée différemment)
    af_taux: float = 0.0          # 6.40%
    af_montant: float = 0.0       # Allocations Familiales
    ps_taux: float = 0.0          # 13.46%
    ps_montant: float = 0.0       # Prestations Sociales
    total_cotisations: float = 0.0  # C3
    penalites_cotis: float = 0.0    # C4
    tfp_masse: float = 0.0
    tfp_taux: float = 0.0          # 1.60%
    tfp_montant: float = 0.0       # Taxe Formation Pro
    penalites_tfp: float = 0.0     # C9
    montant_global_rg: float = 0.0  # C10

    # Page 2 — AMO
    masse_salariale_amo: float = 0.0
    participation_amo_taux: float = 0.0  # 1.85%
    participation_amo: float = 0.0
    cotisation_amo_taux: float = 0.0     # 4.52%
    cotisation_amo: float = 0.0
    total_amo: float = 0.0
    penalites_amo: float = 0.0
    montant_global_amo: float = 0.0      # C10

    # Totaux
    montant_total: float = 0.0  # RG + AMO

    anomalies: List[str] = field(default_factory=list)


# ============================================================
# EXTRACTION
# ============================================================

def extract_cnss_line(text: str, line_code: int) -> List[float]:
    """Extrait les montants d'une ligne codée C (1, 2, 3, 4, 8, 9, 10)."""
    lines = text.split('\n')
    for i, line in enumerate(lines):
        # Chercher la ligne qui commence par le code numéro
        # Format: "1 ﺔّﻴﻠﺋﺎﻌﻟﺍ ﺢﻨﻤﻟﺍ 86,111.10 6,40% 5,511.11"
        # ou:     "10 ﺀﺍﺩﻸﻟ ﻲﻟﺎﻤﺟﻹﺍ ﻎﻠﺒﻤﻟﺍ 18,479.44"
        m = re.match(rf'^\s*{line_code}\b', line)
        if not m:
            continue

        # Combiner avec la ligne suivante (wrapping possible)
        combined = line
        if i + 1 < len(lines) and not re.match(r'^\s*\d{1,2}\b', lines[i+1]):
            combined += " " + lines[i+1]

        amounts = find_amounts(combined)
        return amounts

    return []


def extract_page_rg(text: str, bord: BordereauCNSS):
    """Extrait les données de la page Régime Général."""

    # Période : "11 2025" sur la ligne avant "VERSEMENT DU MOIS DE"
    m = re.search(r'(\d{1,2})\s+(\d{4})\s*\n\s*VERSEMENT', text, re.I)
    if not m:
        m = re.search(r'VERSEMENT\s+DU\s+MOIS\s+DE\s*\n?\s*(\d{1,2})\s+(\d{4})', text, re.I)
    if not m:
        # Chercher juste "MM YYYY" dans le texte
        m = re.search(r'(\d{1,2})\s+(20\d{2})', text)
    if m:
        mois = int(m.group(1))
        annee = int(m.group(2))
        if 1 <= mois <= 12:
            bord.mois_num = mois
            bord.annee = annee

    # Raison sociale : format CNSS = "Raison sociale:" sur une ligne, nom sur une autre
    # Ou bien le nom apparaît après le N° d'affilié : "5162424 ISSAL MADINA"
    m = re.search(r'\b(\d{7})\s+([A-Z][A-Z\s]{3,}?)(?:\n|$)', text)
    if m:
        name = m.group(2).strip()
        if name and name not in ("REFERENCE STRUCTUREE", "N D AFFILIE"):
            bord.raison_sociale = name[:50]
    if not bord.raison_sociale:
        m = re.search(r'Raison\s+sociale\s*:\s*\n?\s*([A-Z][A-Z\s]{3,})', text, re.I)
        if m:
            bord.raison_sociale = m.group(1).strip()[:50]

    # N° affilié
    m = re.search(r'(\d{7})', text)
    if m:
        bord.n_affilie = m.group(1)

    # C1: Allocations Familiales
    amounts = extract_cnss_line(text, 1)
    if len(amounts) >= 2:
        bord.masse_salariale_af = amounts[0]
        bord.af_montant = amounts[-1]
        for line in text.split('\n'):
            if re.match(r'^\s*1\b', line) and '6' in line:
                tm = re.search(r'(\d+[.,]\d+)%', line)
                if tm:
                    bord.af_taux = float(tm.group(1).replace(',','.'))
                break

    # C2: Prestations Sociales
    amounts = extract_cnss_line(text, 2)
    if len(amounts) >= 2:
        bord.masse_salariale_ps = amounts[0]
        bord.ps_montant = amounts[-1]

    # C3: Total cotisations
    amounts = extract_cnss_line(text, 3)
    if amounts:
        bord.total_cotisations = amounts[-1]

    # C4: Pénalités
    amounts = extract_cnss_line(text, 4)
    if amounts:
        bord.penalites_cotis = amounts[-1]

    # C8: TFP — chercher la ligne avec "1,60%" ou "1.60%"
    # Le C8 n'est pas toujours bien parsé car la ligne 8 est mélangée
    # Chercher le montant TFP directement
    m = re.search(r'(\d[\d,]*\.\d{2})\s+1[.,]60%\s+(\d[\d,]*\.\d{2})', text)
    if m:
        bord.tfp_masse = parse_amount(m.group(1)) or 0
        bord.tfp_montant = parse_amount(m.group(2)) or 0
        bord.tfp_taux = 1.60

    # C9: Pénalités TFP
    amounts = extract_cnss_line(text, 9)
    if amounts:
        bord.penalites_tfp = amounts[-1]

    # C10: Montant global RG
    amounts = extract_cnss_line(text, 10)
    if amounts:
        bord.montant_global_rg = amounts[-1]


def extract_page_amo(text: str, bord: BordereauCNSS):
    """Extrait les données de la page AMO."""

    # C1: Participation AMO
    amounts = extract_cnss_line(text, 1)
    if len(amounts) >= 2:
        bord.masse_salariale_amo = amounts[0]
        bord.participation_amo = amounts[-1]
        for line in text.split('\n'):
            if re.match(r'^\s*1\b', line):
                tm = re.search(r'(\d+[.,]\d+)%', line)
                if tm:
                    bord.participation_amo_taux = float(tm.group(1).replace(',','.'))
                break

    # C2: Cotisation AMO
    amounts = extract_cnss_line(text, 2)
    if len(amounts) >= 2:
        bord.cotisation_amo = amounts[-1]
        for line in text.split('\n'):
            if re.match(r'^\s*2\b', line):
                tm = re.search(r'(\d+[.,]\d+)%', line)
                if tm:
                    bord.cotisation_amo_taux = float(tm.group(1).replace(',','.'))
                break

    # C3: Total AMO
    amounts = extract_cnss_line(text, 3)
    if amounts:
        bord.total_amo = amounts[-1]

    # C4: Pénalités AMO
    amounts = extract_cnss_line(text, 4)
    if amounts:
        bord.penalites_amo = amounts[-1]

    # C10: Montant global AMO
    amounts = extract_cnss_line(text, 10)
    if amounts:
        bord.montant_global_amo = amounts[-1]


def process_cnss_file(pdf_path: Path) -> Optional[BordereauCNSS]:
    print(f"\n📄 {pdf_path.name}")
    bord = BordereauCNSS(fichier=pdf_path.name)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
    except:
        # Fallback OCR
        try:
            images = convert_from_path(str(pdf_path), dpi=250, poppler_path=POPPLER_PATH)
            pages = [pytesseract.image_to_string(img, lang="fra+eng", config="--psm 6") for img in images]
        except Exception as e:
            print(f"   ❌ Erreur: {e}")
            return None

    if not pages:
        print(f"   ⚠️ Aucune page")
        return None

    # Debug
    debug = OUTPUT_DIR / f"DEBUG_CNSS_{pdf_path.stem}.txt"
    with open(debug, "w", encoding="utf-8") as f:
        for i, t in enumerate(pages):
            f.write(f"\n{'='*60}\nPAGE {i+1}\n{'='*60}\n{t}\n")

    # Page 1: Régime Général (contient "Allocations Familliales")
    # Page 2: AMO (contient "Cotisation AMO")
    for text in pages:
        if "Allocations" in text or "Régime Général" in text or "formation professionnelle" in text:
            extract_page_rg(text, bord)
        if "AMO" in text or "Assurance Maladie" in text:
            extract_page_amo(text, bord)

    # Total
    bord.montant_total = bord.montant_global_rg + bord.montant_global_amo

    # Cross-checks
    if bord.af_montant > 0 and bord.ps_montant > 0:
        calc_total = bord.af_montant + bord.ps_montant
        if bord.total_cotisations > 0:
            ecart = abs(calc_total - bord.total_cotisations)
            if ecart > 0.02:
                bord.anomalies.append(f"AF+PS={calc_total:.2f} ≠ C3={bord.total_cotisations:.2f}")

    if bord.masse_salariale_af > 0 and bord.af_montant > 0:
        taux_calc = bord.af_montant / bord.masse_salariale_af * 100
        if abs(taux_calc - 6.40) > 0.1:
            bord.anomalies.append(f"Taux AF calculé={taux_calc:.2f}% ≠ 6.40%")

    if not bord.mois_num:
        # Fallback: extraire du nom de fichier
        m = re.search(r'(\d{1,2})[-_](\d{2,4})', pdf_path.stem)
        if m:
            bord.mois_num = int(m.group(1))
            a = m.group(2)
            bord.annee = int(a) if len(a)==4 else 2000+int(a)
        if not bord.mois_num:
            bord.anomalies.append("Mois non trouvé")

    # Affichage
    mois_label = MOIS.get(bord.mois_num, "?")
    print(f"   📅 {mois_label} {bord.annee} | {bord.raison_sociale}")
    print(f"   Masse salariale AF : {bord.masse_salariale_af:>12,.2f}")
    print(f"   Masse salariale PS : {bord.masse_salariale_ps:>12,.2f}")
    print(f"   AF (6.40%)         : {bord.af_montant:>12,.2f}")
    print(f"   PS (13.46%)        : {bord.ps_montant:>12,.2f}")
    print(f"   TFP (1.60%)        : {bord.tfp_montant:>12,.2f}")
    print(f"   Total RG           : {bord.montant_global_rg:>12,.2f}")
    print(f"   Participation AMO  : {bord.participation_amo:>12,.2f}")
    print(f"   Cotisation AMO     : {bord.cotisation_amo:>12,.2f}")
    print(f"   Total AMO          : {bord.montant_global_amo:>12,.2f}")
    print(f"   TOTAL GLOBAL       : {bord.montant_total:>12,.2f}")
    if bord.anomalies:
        for a in bord.anomalies: print(f"   ⚠️ {a}")

    return bord


# ============================================================
# GÉNÉRATION EXCEL
# ============================================================

def generate_excel(declarations: List[BordereauCNSS], societe: str, output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "CNSS"

    headers = [
        "Mois", "Masse Salariale AF", "Masse Salariale PS",
        "Alloc. Familiales (6.40%)", "Prestations Sociales (13.46%)",
        "Total Cotisations", "Pénalités Cotis.",
        "TFP (1.60%)", "Pénalités TFP",
        "Total Régime Général",
        "Participation AMO (1.85%)", "Cotisation AMO (4.52%)",
        "Total AMO", "Pénalités AMO",
        "Total Global (RG+AMO)",
        "Anomalies"
    ]
    ws.append(headers)

    # Style header
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name="Arial", color="FFFFFF", bold=True, size=9)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Index par mois
    data_by_month = {}
    for d in declarations:
        if d.mois_num:
            data_by_month[d.mois_num] = d

    # 12 mois
    for m in range(1, 13):
        row_num = m + 1  # row 2-13
        if m in data_by_month:
            d = data_by_month[m]
            ws.append([
                f"{MOIS[m]} {d.annee}",
                d.masse_salariale_af, d.masse_salariale_ps,
                d.af_montant, d.ps_montant,
                d.total_cotisations, d.penalites_cotis,
                d.tfp_montant, d.penalites_tfp,
                d.montant_global_rg,
                d.participation_amo, d.cotisation_amo,
                d.total_amo, d.penalites_amo,
                d.montant_total,
                "; ".join(d.anomalies) if d.anomalies else "OK"
            ])
        else:
            ws.append([MOIS[m]] + [""] * 15)
            # Colorer en jaune les mois manquants
            rn = ws.max_row
            for c in range(2, 16):
                ws.cell(row=rn, column=c).fill = PatternFill("solid", fgColor="FFEB9C")

        # Formatage nombres
        rn = ws.max_row
        for c in range(2, 16):
            cell = ws.cell(row=rn, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            cell.font = Font(name="Arial", size=9)

    # Total avec formules
    total_row = 14
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", size=9, bold=True)
    for c in range(2, 16):
        cl = get_column_letter(c)
        ws.cell(row=total_row, column=c).value = f"=SUM({cl}2:{cl}13)"
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(name="Arial", size=9, bold=True)
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")

    # Largeurs
    widths = {1:16, 2:18, 3:18, 4:22, 5:24, 6:18, 7:16, 8:16, 9:14, 10:20,
              11:20, 12:20, 13:14, 14:14, 15:20, 16:30}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    soc = re.sub(r'[\\/*?:"<>|]', '_', societe)
    out = output_dir / f"CNSS_{soc}_{ts}.xlsx"
    wb.save(out)
    return out


# ============================================================
# MENU
# ============================================================

def main():
    print("\n" + "="*70)
    print("  EXTRACT-CNSS V2 — Bordereaux CNSS (RG + AMO)")
    print("  AF, PS, TFP, AMO par mois")
    print("="*70)
    print(f"\n📁 CNSS   : {CNSS_DIR}")
    print(f"📁 Output : {OUTPUT_DIR}")

    pdf_files = sorted(CNSS_DIR.glob("*.pdf"))
    if not pdf_files:
        print(f"\n❌ Aucun PDF dans {CNSS_DIR}")
        input("Entrée..."); return

    print(f"\n📄 {len(pdf_files)} fichier(s):")
    for i, f in enumerate(pdf_files, 1):
        print(f"   {i}. {f.name}")

    societe = input("\n🏢 Société: ").strip().upper() or "SOCIETE"

    choix = input(f"\nSélection (T=Tous, ou ex: 1,3) [T]: ").strip().upper()
    if choix and choix != 'T':
        sel = []
        for c in choix.split(','):
            c = c.strip()
            if c.isdigit() and 1 <= int(c) <= len(pdf_files):
                sel.append(pdf_files[int(c)-1])
        if sel: pdf_files = sel

    declarations = []
    for pdf in pdf_files:
        d = process_cnss_file(pdf)
        if d: declarations.append(d)

    if not declarations:
        print("\n❌ Aucune donnée extraite.")
        input("Entrée..."); return

    out = generate_excel(declarations, societe, OUTPUT_DIR)
    print(f"\n✅ Excel → {out}")

    # Résumé
    print(f"\n{'='*60}")
    print(f"📊 RÉSUMÉ — {societe}")
    print(f"{'='*60}")
    print(f"{'Mois':>12} {'Masse AF':>14} {'Masse PS':>14} {'AF':>12} {'PS':>12} {'TFP':>10} {'AMO':>10} {'TOTAL':>12}")
    print(f"{'-'*12} {'-'*14} {'-'*14} {'-'*12} {'-'*12} {'-'*10} {'-'*10} {'-'*12}")
    for d in sorted(declarations, key=lambda x: (x.annee, x.mois_num)):
        print(f"{MOIS.get(d.mois_num,'?'):>12} {d.masse_salariale_af:>14,.2f} {d.masse_salariale_ps:>14,.2f} {d.af_montant:>12,.2f} "
              f"{d.ps_montant:>12,.2f} {d.tfp_montant:>10,.2f} {d.montant_global_amo:>10,.2f} {d.montant_total:>12,.2f}")

    tot_af = sum(d.af_montant for d in declarations)
    tot_ps = sum(d.ps_montant for d in declarations)
    tot_tfp = sum(d.tfp_montant for d in declarations)
    tot_amo = sum(d.montant_global_amo for d in declarations)
    tot_all = sum(d.montant_total for d in declarations)
    print(f"{'-'*12} {'-'*14} {'-'*12} {'-'*12} {'-'*10} {'-'*10} {'-'*12}")
    print(f"{'TOTAL':>12} {'':>14} {tot_af:>12,.2f} {tot_ps:>12,.2f} {tot_tfp:>10,.2f} {tot_amo:>10,.2f} {tot_all:>12,.2f}")

    input("\n✅ Terminé! Entrée...")

if __name__ == "__main__":
    main()