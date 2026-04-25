"""
Lecture et validation de la feuille BG — Revue Analytique BG-GL.
"""
from __future__ import annotations

import re
from typing import Optional

import openpyxl
from openpyxl import Workbook

from .models import BGRow
from .account_grouper import AccountNormalizer

MANDATORY_COLUMNS = {"Compte", "Intitule", "Solde N", "Solde N-1"}

_COLUMN_ALIASES: dict[str, list[str]] = {
    "Compte": [
        "compte", "num compte", "numero compte", "n° compte", "no compte",
        "ledger account", "account", "account number", "numero", "num",
    ],
    "Intitule": [
        "intitule", "intitulé", "libelle", "libellé", "designation", "désignation",
        "name", "account name", "account name in english", "libellé compte",
        "nom", "description",
    ],
    "Solde N": [
        "solde n", "solden", "solde_n", "solde exercice", "solde",
        "closing balance", "solde cloture", "solde clôture",
        "solde final", "solde débiteur", "solde créditeur",
        "solde n-0", "31/12", "31/12/2025", "31/12/2026", "31/12/2027", "31/12/2028",
    ],
    "Solde N-1": [
        "solde n-1", "solden-1", "solde_n-1", "solde n1", "solde precedent", "solde précédent",
        "opening balance", "solde ouverture",
        "solde n-2", "solde initial", "solde debut", "31/12/2024", "31/12/2023", "31/12/2022",
    ],
    "Ref": ["ref", "réf", "reference", "référence", "réf"],
}

_HEADER_SCAN_MAX_ROWS = 20
_HEADER_SCAN_MAX_COLS = 80

_DATE_HEADER_RE = re.compile(r"^\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}$")

_CLASSIFICATION_RE = re.compile(r"^[0-9A-Za-z\-]+$")


class BGValidationError(Exception):
    """Erreur fatale — arrêt de la génération."""


def _norm_header(h: str) -> str:
    """Normalise un en-tête pour la comparaison (minuscule, strip, sans accent)."""
    import unicodedata
    s = str(h).strip().lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def _is_date_header(h: str) -> bool:
    """True si l'en-tête ressemble à une date (format JJ/MM/AAAA)."""
    return bool(_DATE_HEADER_RE.match(str(h).strip()))


def _is_formula(v) -> bool:
    """True si la valeur est une formule Excel (commence par =)."""
    return isinstance(v, str) and v.strip().startswith("=")


def _is_account_code(v) -> bool:
    """True si la valeur ressemble à un numéro de compte (6+ chiffres numériques)."""
    s = str(v).strip()
    if not s:
        return False
    if not s.isdigit():
        return False
    if len(s) >= 6:
        return True
    return False


def _classify_by_data_pattern(
    ws,
    header_row_idx: int,
    col_idx: int,
    max_sample_rows: int = 10,
) -> Optional[str]:
    """
    Examine les données sous l'en-tête pour inférer le type de colonne.
    Retourne 'Compte', 'Intitule', 'Solde N', 'Solde N-1' ou None.
    """
    if ws is None:
        return None
    sample_vals: list = []
    for r in range(header_row_idx + 1, min(header_row_idx + 1 + max_sample_rows, ws.max_row + 1)):
        v = ws.cell(row=r, column=col_idx).value
        sample_vals.append(v)

    non_none = [v for v in sample_vals if v is not None and str(v).strip() != ""]
    if not non_none:
        return None

    num_count = sum(1 for v in non_none if _is_numeric_like(v))
    str_count = sum(1 for v in non_none if _is_string_like(v))
    account_count = sum(1 for v in non_none if _is_account_code(v))

    if account_count >= len(non_none) * 0.6:
        return "Compte"

    if num_count >= len(non_none) * 0.7:
        return "Solde N"

    if str_count >= len(non_none) * 0.7:
        first_str = str(non_none[0])[:20]
        if _CLASSIFICATION_RE.match(first_str) and first_str.isalnum():
            return "Compte"
        return "Intitule"

    return None


def _best_numeric_cols(
    ws,
    header_row_idx: int,
    norm_headers: list[str],
    max_sample_rows: int = 15,
) -> tuple[int, int]:
    """
    Retourne (col_solde_n1, col_solde_n) comme indices 0-based.
    Pour les fichiers D/C : utilise la dernière paire D/C avec données.
    Pour les autres fichiers : colonnes avec le plus de données numériques.
    """
    c_cols: list[int] = []
    d_cols: list[int] = []
    for idx, h in enumerate(norm_headers):
        if h == "c":
            c_cols.append(idx)
        elif h == "d":
            d_cols.append(idx)

    if c_cols:
        best_c = c_cols[-1]
        best_c_count = 0
        for r in range(header_row_idx + 1, min(header_row_idx + 1 + max_sample_rows, ws.max_row + 1)):
            v = ws.cell(row=r, column=best_c + 1).value
            if v is not None and str(v).strip() != "" and not _is_formula(v):
                best_c_count += 1

        if best_c_count > 0:
            paired_d_idx = best_c - 1
            if paired_d_idx in d_cols:
                c_n = best_c
                c_n1 = paired_d_idx
            elif len(c_cols) >= 2:
                c_n = c_cols[-1]
                c_n1 = c_cols[-2]
            else:
                c_n = c_cols[-1]
                c_n1 = -1
            return c_n1, c_n

    # Identifier les colonnes de classification (C1, C2, C3, C4, C5) à ignorer
    classification_cols = set()
    for idx, h in enumerate(norm_headers):
        if h in ("c1", "c2", "c3", "c4", "c5"):
            classification_cols.add(idx)

    # Chercher d'abord les colonnes avec en-têtes de date (31/12/YYYY, JJ/MM/AA, etc.)
    date_header_cols: list[int] = []
    for idx, h in enumerate(norm_headers):
        if _is_date_header(h):
            date_header_cols.append(idx)

    # Si colonnes de date trouvées, les utiliser (inverser: N-1 puis N)
    if len(date_header_cols) >= 2:
        return date_header_cols[-1], date_header_cols[-2]
    if len(date_header_cols) == 1:
        return -1, date_header_cols[0]

    all_numeric: dict[int, int] = {}
    for col_idx_0, h in enumerate(norm_headers):
        if _is_formula(h):
            continue
        # Ignorer les colonnes de classification
        if col_idx_0 in classification_cols:
            continue
        if col_idx_0 in all_numeric.values():
            continue
        inferred = _classify_by_data_pattern(ws, header_row_idx, col_idx_0 + 1, max_sample_rows)
        if inferred == "Solde N":
            non_none = 0
            for r in range(header_row_idx + 1, min(header_row_idx + 1 + max_sample_rows, ws.max_row + 1)):
                v = ws.cell(row=r, column=col_idx_0 + 1).value
                if v is not None and str(v).strip() != "" and not _is_formula(v):
                    non_none += 1
            all_numeric[col_idx_0] = non_none

    if not all_numeric:
        return -1, -1
    sorted_cols = sorted(all_numeric.keys(), key=lambda k: all_numeric[k], reverse=True)
    if len(sorted_cols) >= 2:
        return sorted_cols[1], sorted_cols[0]
    if len(sorted_cols) == 1:
        return -1, sorted_cols[0]
    return -1, -1


def _is_numeric_like(v) -> bool:
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def _is_string_like(v) -> bool:
    s = str(v).strip()
    if not s:
        return False
    try:
        float(s)
        return False
    except (ValueError, TypeError):
        return True


def _map_columns(
    raw_headers: list[str],
    require_mandatory: bool = True,
    ws=None,
    header_row_idx: int = 1,
) -> dict[str, int]:
    """
    Retourne un mapping {clé_métier: index_colonne_0_based}.
    Stratégie en 4 passes :
    1. Détection de colonnes de date (31/12/YYYY) pour Solde N et N-1
    2. Alias nominaux (Compte, Solde N, etc.)
    3. Heuristique : colonnes avec le plus de données numériques (pour D/C, dates)
    4. Inférence par pattern de données
    """
    norm_headers = [_norm_header(h) for h in raw_headers]
    mapping: dict[str, int] = {}

    # PASS 1: Détection de colonnes de date (prioritaire)
    date_header_cols: list[int] = []
    for idx, h in enumerate(norm_headers):
        if _is_date_header(h):
            date_header_cols.append(idx)
    if len(date_header_cols) >= 2:
        mapping["Solde N-1"] = date_header_cols[-1]
        mapping["Solde N"] = date_header_cols[-2]
    elif len(date_header_cols) == 1:
        mapping["Solde N"] = date_header_cols[0]

    # PASS 2: Alias nominaux (Compte, Solde N, etc.) - mais ne surcharge pas les colonnes de date
    for key, aliases in _COLUMN_ALIASES.items():
        if key in mapping:
            continue
        for alias in aliases:
            if alias in norm_headers:
                mapping[key] = norm_headers.index(alias)
                break

    missing = MANDATORY_COLUMNS - set(mapping.keys())
    if missing and ws is not None:
        remaining_missing = [k for k in missing if k not in mapping]

        col_n1, col_n = _best_numeric_cols(ws, header_row_idx, norm_headers, 15)
        if col_n >= 0 and "Solde N" in remaining_missing:
            mapping["Solde N"] = col_n
            remaining_missing.remove("Solde N")
        if col_n1 >= 0 and "Solde N-1" in remaining_missing:
            mapping["Solde N-1"] = col_n1
            remaining_missing.remove("Solde N-1")

        for col_idx_0, h in enumerate(norm_headers):
            if col_idx_0 in mapping.values():
                continue
            inferred = _classify_by_data_pattern(
                ws, header_row_idx, col_idx_0 + 1, max_sample_rows=15
            )
            if inferred and inferred in remaining_missing:
                mapping[inferred] = col_idx_0
                remaining_missing.remove(inferred)
                if not remaining_missing:
                    break

    missing = MANDATORY_COLUMNS - set(mapping.keys())
    if require_mandatory and missing:
        raise BGValidationError(
            f"Colonnes obligatoires manquantes dans la feuille BG : {', '.join(sorted(missing))}. "
            f"En-têtes trouvés : {raw_headers}"
        )
    return mapping


class BGReader:
    """Lit la feuille BG, valide sa structure et produit une liste de BGRow."""

    def __init__(self, filepath: str, sheet_name: str | None = None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.workbook: Workbook | None = None
        self.invalid_rows_count: int = 0
        self.invalid_rows_examples: list[str] = []
        self.warnings: list[str] = []
        # Exposés après read_rows() pour FormulaBuilder
        self.col_map: dict[str, int] = {}   # {clé_métier: col_index_1based}
        self.last_data_row: int = 1
        self.has_ref_col: bool = False
        self.header_row_idx: int = 1  # 1-based
        self._resolved_ws = None
        self._resolved_col_map_0: dict[str, int] = {}
        self._is_resolved: bool = False

    def _open(self) -> None:
        if self.workbook is None:
            try:
                self.workbook = openpyxl.load_workbook(self.filepath, data_only=False)
            except Exception as exc:
                raise BGValidationError(f"Impossible d'ouvrir le classeur '{self.filepath}' : {exc}") from exc

    def _open_data_only(self) -> Workbook:
        try:
            return openpyxl.load_workbook(self.filepath, data_only=True)
        except Exception as exc:
            raise BGValidationError(f"Impossible d'ouvrir le classeur en mode données : {exc}") from exc

    def _sheet_priority(self) -> list[str]:
        names = list(self.workbook.sheetnames)
        if self.sheet_name and self.sheet_name in names:
            return [self.sheet_name] + [n for n in names if n != self.sheet_name]
        return names

    @staticmethod
    def _row_score(headers: list[str], mapping: dict[str, int]) -> float:
        norm = [_norm_header(h) for h in headers]
        score = float(len(mapping))
        if any("compte" in h or "account" in h or "ledger" in h for h in norm):
            score += 3.0
        if any("solde" in h or "balance" in h or "debit" in h or "credit" in h for h in norm):
            score += 2.0
        if any("intitule" in h or "name" in h or "libelle" in h or "libellé" in h for h in norm):
            score += 2.0
        empty_count = sum(1 for h in headers if not h.strip())
        score -= empty_count * 0.1
        return score

    def _detect_sheet_and_header(self):
        best = None
        best_score = -1.0
        for sheet_name in self._sheet_priority():
            ws = self.workbook[sheet_name]
            max_row = min(ws.max_row or 1, _HEADER_SCAN_MAX_ROWS)
            max_col = min(ws.max_column or 1, _HEADER_SCAN_MAX_COLS)

            for row_idx in range(1, max_row + 1):
                row_vals = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
                headers = [str(v) if v is not None else "" for v in row_vals]

                if not any(h.strip() for h in headers):
                    continue

                mapping = _map_columns(headers, require_mandatory=False, ws=ws, header_row_idx=row_idx)
                missing = MANDATORY_COLUMNS - set(mapping.keys())
                if missing:
                    continue

                score = self._row_score(headers, mapping)
                candidate = (score, ws, row_idx, mapping)

                if row_idx < max_row:
                    next_row_is_data = False
                    for col_idx in range(1, max_col + 1):
                        first_cell_val = ws.cell(row=row_idx + 1, column=col_idx).value
                        if first_cell_val is not None:
                            next_is_formula = _is_formula(first_cell_val)
                            next_is_empty = not str(first_cell_val).strip()
                            if next_is_formula or next_is_empty:
                                continue
                            next_row_is_data = True
                            break
                    if not next_row_is_data:
                        score -= 5.0
                        candidate = (score, ws, row_idx, mapping)

                if score > best_score:
                    best_score = score
                    best = candidate

        if best is None:
            raise BGValidationError(
                "Impossible de détecter automatiquement la feuille BG et ses en-têtes obligatoires "
                f"({', '.join(sorted(MANDATORY_COLUMNS))}). "
                "Vérifiez que le fichier contient bien les colonnes Compte, Intitule, Solde N et Solde N-1."
            )

        _, ws, row_idx, mapping = best
        return ws, row_idx, mapping

    def _resolve_structure(self) -> None:
        if self._is_resolved:
            return

        self._open()
        ws, header_row_idx, col_map = self._detect_sheet_and_header()

        self._resolved_ws = ws
        self._resolved_col_map_0 = col_map
        self._is_resolved = True

        self.sheet_name = ws.title
        self.header_row_idx = header_row_idx
        self.col_map = {k: v + 1 for k, v in col_map.items()}
        self.has_ref_col = "Ref" in col_map

        self._data_wb = self._open_data_only()
        self._data_ws = self._data_wb[ws.title] if ws.title in self._data_wb.sheetnames else None

        if not self.has_ref_col:
            msg = (
                "Colonne 'Ref' absente de la feuille BG — la colonne Ref des feuilles analytiques "
                "sera alimentée par une formule vide (=\"\")."
            )
            if msg not in self.warnings:
                self.warnings.append(msg)

    def validate_only(self) -> None:
        """Ouvre le classeur et vérifie la structure BG sans lire les données."""
        self._resolve_structure()

    def read_rows(self) -> list[BGRow]:
        """Lit toutes les lignes BG valides. Les lignes invalides sont comptées mais ignorées."""
        self._resolve_structure()
        data_ws = self._data_ws if self._data_ws is not None else self._resolved_ws
        col_map = self._resolved_col_map_0

        normalizer = AccountNormalizer()
        result: list[BGRow] = []
        self.invalid_rows_count = 0
        self.invalid_rows_examples = []

        has_ref_col = self.has_ref_col

        last_data_row = self.header_row_idx
        for physical_row_idx, row in enumerate(
            data_ws.iter_rows(min_row=self.header_row_idx + 1, max_row=data_ws.max_row, values_only=True),
            start=self.header_row_idx + 1,
        ):
            raw_compte = row[col_map["Compte"]]
            if raw_compte is None:
                continue

            if _is_formula(raw_compte):
                continue

            norm = normalizer.normalize(raw_compte)
            classification = normalizer.classify(norm)

            if classification == "invalid":
                self.invalid_rows_count += 1
                if len(self.invalid_rows_examples) < 10:
                    self.invalid_rows_examples.append(
                        f"Ligne {physical_row_idx}: compte='{raw_compte}' → normalisé='{norm}'"
                    )
                continue

            try:
                raw_solde_n = row[col_map["Solde N"]]
                raw_solde_n1 = row[col_map["Solde N-1"]]
                if isinstance(raw_solde_n, str):
                    raw_solde_n = raw_solde_n.strip().replace(",", ".")
                    if raw_solde_n in ("", "-", "–", "--"):
                        raw_solde_n = "0"
                if isinstance(raw_solde_n1, str):
                    raw_solde_n1 = raw_solde_n1.strip().replace(",", ".")
                    if raw_solde_n1 in ("", "-", "–", "--"):
                        raw_solde_n1 = "0"
                solde_n = float(raw_solde_n or 0)
                solde_n1 = float(raw_solde_n1 or 0)
            except (ValueError, TypeError):
                solde_n = 0.0
                solde_n1 = 0.0

            intitule = str(row[col_map["Intitule"]] or "")
            ref = str(row[col_map["Ref"]] or "") if has_ref_col else ""

            result.append(BGRow(
                row_index=physical_row_idx,
                compte_raw=raw_compte,
                compte_norm=norm,
                intitule=intitule,
                solde_n=solde_n,
                solde_n1=solde_n1,
                ref=ref,
            ))
            last_data_row = physical_row_idx

        self.last_data_row = last_data_row
        return result
