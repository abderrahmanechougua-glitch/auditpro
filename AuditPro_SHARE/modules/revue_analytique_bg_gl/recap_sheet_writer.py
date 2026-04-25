"""
Génération des feuilles récapitulatives par poste (3 chiffres) — Revue Analytique BG-GL.

Regroupe les tableaux analytiques (4 chiffres) par poste (3 chiffres) dans des feuilles dédiées.
Exemple: La feuille "612" regroupe tous les comptes 6121, 6122, 6123, 6124...
"""
from __future__ import annotations

from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .poste_labels import get_poste_label


class RecapitulativeSheetWriter:
    """
    Génère les feuilles récapitulatives par poste (3 chiffres).

    Chaque feuille regroupe verticalement tous les tableaux analytiques (4 chiffres)
    appartenant au même poste (3 premiers chiffres).

    Exemple pour la feuille "612":
    - Titre: "Feuille récapitulative 612" (Gras, Arial 12, fond jaune)
    - 3 lignes vides
    - Tableau 6121 + commentaire
    - 3 lignes vides
    - Tableau 6122 + commentaire
    - 3 lignes vides
    - Tableau 6123 + commentaire
    - ...
    """

    def __init__(
        self,
        wb: Workbook,
        aggregates: list,
    ):
        """
        Initialise le writer pour feuilles récapitulatives.

        wb: Workbook openpyxl (doit déjà contenir les feuilles 4 chiffres)
        aggregates: Liste des AggregateAccount (pour identifier les comptes)
        """
        self.wb = wb
        self.aggregates = aggregates

        # Style pour le titre
        self._title_font = Font(bold=True, name="Arial", size=12)
        self._title_fill = PatternFill("solid", fgColor="FFFF00")  # Jaune

    def _get_postes_ordered(self) -> dict[str, list]:
        """
        Identifie les postes (3 chiffres) dans l'ordre d'apparition.

        Retourne: {
            "612": [AggregateAccount(code_4="6121"), AggregateAccount(code_4="6122"), ...],
            "613": [AggregateAccount(code_4="6131"), ...],
            ...
        }
        """
        postes = {}
        for agg in self.aggregates:
            poste_3 = agg.code_4[:3]  # "6121" → "612"
            if poste_3 not in postes:
                postes[poste_3] = []
            postes[poste_3].append(agg)

        return postes

    def _copy_sheet_content(
        self,
        source_ws,
        dest_ws,
        start_row: int,
    ) -> int:
        """
        Copie le contenu complet d'une feuille source vers une feuille destination.

        Copie: valeurs, formules, styles (font, fill, border, number_format, alignment).
        Exclut la dernière ligne (commentaire analytique).

        source_ws: Feuille source (ex: feuille "6121")
        dest_ws: Feuille destination (ex: feuille "612")
        start_row: Numéro de ligne où commencer la copie dans la destination

        Retourne: Nombre de lignes copiées
        """
        rows_copied = 0
        max_row = source_ws.max_row

        for row_idx, row in enumerate(source_ws.iter_rows(values_only=False), start=1):
            # Exclure la dernière ligne (commentaire analytique)
            if row_idx == max_row:
                break

            dest_row_idx = start_row + row_idx - 1
            rows_copied = row_idx

            for col_idx, cell in enumerate(row, start=1):
                dest_cell = dest_ws.cell(row=dest_row_idx, column=col_idx)

                # Copier la valeur
                dest_cell.value = cell.value

                # Copier les styles
                if cell.font:
                    dest_cell.font = copy(cell.font)
                if cell.fill:
                    dest_cell.fill = copy(cell.fill)
                if cell.border:
                    dest_cell.border = copy(cell.border)
                if cell.alignment:
                    dest_cell.alignment = copy(cell.alignment)

                dest_cell.number_format = cell.number_format

                # Copier la largeur de colonne (une seule fois)
                if row_idx == 1:
                    col_letter = get_column_letter(col_idx)
                    if source_ws.column_dimensions[col_letter].width:
                        dest_ws.column_dimensions[col_letter].width = (
                            source_ws.column_dimensions[col_letter].width
                        )

        return rows_copied

    def _add_recap_title(self, ws, poste_code: str) -> None:
        """
        Ajoute le titre de la feuille récapitulative.

        Format: Colonne C, ligne 2, "Feuille récapitulative 612 - Achats consommés..."
        Style: Gras, Arial 12, fond jaune
        """
        poste_label = get_poste_label(poste_code)
        if poste_label:
            title = f"Feuille récapitulative {poste_code} - {poste_label}"
        else:
            title = f"Feuille récapitulative {poste_code}"

        title_cell = ws.cell(row=2, column=3)  # Colonne C
        title_cell.value = title
        title_cell.font = self._title_font
        title_cell.fill = self._title_fill

    def generate_all(self) -> None:
        """
        Génère toutes les feuilles récapitulatives.

        Crée une feuille par poste (3 chiffres) avec tous les comptes (4 chiffres)
        appartenant à ce poste empilés verticalement.
        """
        postes = self._get_postes_ordered()

        for poste_code, aggs_list in postes.items():
            # Créer la feuille pour ce poste
            ws = self.wb.create_sheet(title=poste_code)

            # Ajouter le titre
            self._add_recap_title(ws, poste_code)

            # Commencer après le titre + 3 lignes vides
            current_row = 6  # Ligne 2 (titre) + 3 lignes vides + 1

            # Empiler verticalement chaque compte 4 chiffres du poste
            for idx, agg in enumerate(aggs_list):
                # Récupérer la feuille source (créée dans write_sheet)
                if agg.code_4 not in self.wb.sheetnames:
                    # Si la feuille n'existe pas, la créer vide avec un message
                    source_ws = self.wb.create_sheet(title=agg.code_4)
                    source_ws.cell(1, 1).value = f"Compte {agg.code_4} non trouvé"
                else:
                    source_ws = self.wb[agg.code_4]

                # Copier le contenu complet de la feuille 4 chiffres
                rows_copied = self._copy_sheet_content(source_ws, ws, current_row)

                # Avancer pour le prochain compte (nombre de lignes copiées + 3 lignes vides)
                current_row += rows_copied + 3
