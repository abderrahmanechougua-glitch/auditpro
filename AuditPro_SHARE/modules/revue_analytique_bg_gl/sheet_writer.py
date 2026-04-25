"""
Écriture des feuilles analytiques dans le classeur — Revue Analytique BG-GL.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import AggregateAccount
from .formula_builder import (
    FormulaBuilder,
    COL_COMPTE, COL_INTITULE, COL_REF, COL_SOLDE_N, COL_SOLDE_N1, COL_VARIATION, COL_VARIATION_PCT,
)
from .account_labels import get_account_label
from .comment_writer import CommentWriter

# Ligne "En KMAD" + En-têtes + Données
KMAD_ROW = 1
HEADER_ROW = 2
FIRST_DATA_ROW = 3

# Style "En KMAD" — Arial 9
_KMAD_FONT = Font(bold=True, italic=True, name="Arial", size=9)

# Style en-têtes (Violet) — Arial 9 blanc bold
_HEADER_FILL = PatternFill("solid", fgColor="7030A0")  # Violet
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9, name="Arial")

# Style texte normal — Arial 9
_DATA_FONT = Font(name="Arial", size=9)

# Style ligne de total — Arial 9
_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=9, name="Arial")
_TOTAL_FILL = PatternFill("solid", fgColor="7030A0")  # Violet
_NUM_FMT = '#,##0;-#,##0;"-"'

# Bordures
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class SheetWriter:
    """
    Écrit (ou recrée) les feuilles analytiques dans le classeur openpyxl.

    Nouvelle mise en page:
    - Ligne 1: "En KMAD" (bold italic Arial 9)
    - Ligne 2: En-têtes (violet)
    - Ligne 3+: Données

    replace_existing : si True, supprime la feuille existante avant création.
    """

    def __init__(
        self,
        wb: Workbook,
        formula_builder: FormulaBuilder | None = None,
        replace_existing: bool = True,
        bg_rows: list | None = None,
        exercice: str = "2025",
    ):
        self.wb = wb
        self.formula_builder = formula_builder
        self.replace_existing = replace_existing
        self.bg_rows = bg_rows or []
        self.exercice = exercice
        self._comment_writer = CommentWriter()

    def _delete_existing(self, sheet_name: str) -> None:
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]

    def _lookup_in_bg(self, code_8_norm: str) -> tuple:
        """Lookup direct dans les données BG. Retourne (intitule, solde_n, solde_n1) ou ('', 0, 0)."""
        for row in self.bg_rows:
            if row.compte_norm == code_8_norm:
                return (row.intitule, row.solde_n, row.solde_n1)
        return ("", 0, 0)

    def _write_kmad_header(self, ws) -> None:
        """Écrit "En KMAD" en colonne B ligne 1 avec formatage bold italic Arial 9 (sans bordures)."""
        cell = ws.cell(row=KMAD_ROW, column=2, value="En KMAD")
        cell.font = _KMAD_FONT

    def _write_headers(self, ws) -> None:
        """Écrit les en-têtes du tableau à la ligne 2 avec colonne Variation et bordures."""
        try:
            year_n = int(self.exercice)
            year_n1 = year_n - 1
        except (ValueError, TypeError):
            year_n = "N"
            year_n1 = "N-1"

        headers = (
            "Compte",
            "Intitulé",
            "Réf",
            f"31/12/{year_n}",
            f"31/12/{year_n1}",
            "Variation",
            "Variation %",
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
    def create_sheet(self, aggregate: AggregateAccount):
        """T014 — Crée la feuille analytique. Délègue à write_sheet si formula_builder disponible."""
        if self.formula_builder is not None:
            self.write_sheet(aggregate)
            return

        sheet_name = aggregate.sheet_name
        if self.replace_existing:
            self._delete_existing(sheet_name)

        ws = self.wb.create_sheet(title=sheet_name)
        self._write_kmad_header(ws)
        self._write_headers(ws)
        self._set_column_widths(ws)

    def write_sheet(self, aggregate: AggregateAccount) -> None:
        """
        T020 + T023 — Feuille complète :
        - Ligne 1: "En KMAD" en colonne B (bold italic Arial 9)
        - Ligne 2: En-têtes (fond violet, blanc, Arial 9)
        - Lignes 3..N: Sous-comptes avec Variation
        - Ligne N+1: Agrégat 4 chiffres avec intitulé (fond violet, Arial 9)
        - Ligne N+2: Commentaire analytique
        """
        sheet_name = aggregate.sheet_name
        if self.replace_existing:
            self._delete_existing(sheet_name)

        ws = self.wb.create_sheet(title=sheet_name)
        self._write_kmad_header(ws)
        self._write_headers(ws)

        # ── Lignes sous-comptes ──────────────────────────────────────────
        data_start = FIRST_DATA_ROW
        current_row = data_start
        total_solde_n = 0
        total_solde_n1 = 0

        for i, sub in enumerate(aggregate.sub_accounts):
            intitule, solde_n, solde_n1 = self._lookup_in_bg(sub.code_8)
            total_solde_n += solde_n
            total_solde_n1 += solde_n1

            c_compte = ws.cell(row=current_row, column=COL_COMPTE, value=sub.code_8)
            c_intitule = ws.cell(row=current_row, column=COL_INTITULE, value=intitule)
            c_ref = ws.cell(row=current_row, column=COL_REF, value="")
            c_n = ws.cell(row=current_row, column=COL_SOLDE_N, value=solde_n)
            c_n1 = ws.cell(row=current_row, column=COL_SOLDE_N1, value=solde_n1)

            # Colonne Variation = Solde N - Solde N-1
            solde_n_col = get_column_letter(COL_SOLDE_N)
            solde_n1_col = get_column_letter(COL_SOLDE_N1)

            c_variation = ws.cell(row=current_row, column=COL_VARIATION)
            c_variation.value = f"={solde_n_col}{current_row}-{solde_n1_col}{current_row}"
            c_variation.font = _DATA_FONT
            c_variation.alignment = Alignment(horizontal="right")
            c_variation.number_format = _NUM_FMT
            c_variation.border = _THIN_BORDER

            # Colonne Variation % = (Solde N - Solde N-1) / ABS(Solde N-1) * 100
            c_variation_pct = ws.cell(row=current_row, column=COL_VARIATION_PCT)
            c_variation_pct.value = f"=IF({solde_n1_col}{current_row}=0,0,({solde_n_col}{current_row}-{solde_n1_col}{current_row})/ABS({solde_n1_col}{current_row})*100)"
            c_variation_pct.font = _DATA_FONT
            c_variation_pct.alignment = Alignment(horizontal="right")
            c_variation_pct.number_format = '0.0"%"'
            c_variation_pct.border = _THIN_BORDER

            # Formatage Arial 9 et bordures pour les données numériques
            c_compte.font = _DATA_FONT
            c_compte.border = _THIN_BORDER
            c_intitule.font = _DATA_FONT
            c_intitule.border = _THIN_BORDER
            c_ref.font = _DATA_FONT
            c_ref.border = _THIN_BORDER

            # Formatage pour Solde N et N-1
            for c in (c_n, c_n1):
                c.font = _DATA_FONT
                c.alignment = Alignment(horizontal="right")
                c.number_format = _NUM_FMT
                c.border = _THIN_BORDER

            current_row += 1

        last_data_row = current_row - 1

        # ── Ligne agrégat (fond violet, Arial 9) ────────────────────────
        total_row_idx = current_row
        solde_n_col = get_column_letter(COL_SOLDE_N)
        solde_n1_col = get_column_letter(COL_SOLDE_N1)
        variation_col = get_column_letter(COL_VARIATION)
        variation_pct_col = get_column_letter(COL_VARIATION_PCT)

        total_n_formula = f"=SUM({solde_n_col}{data_start}:{solde_n_col}{last_data_row})"
        total_n1_formula = f"=SUM({solde_n1_col}{data_start}:{solde_n1_col}{last_data_row})"
        total_variation_formula = f"=SUM({variation_col}{data_start}:{variation_col}{last_data_row})"
        total_variation_pct_formula = f"=IF({solde_n1_col}{total_row_idx}=0,0,({solde_n_col}{total_row_idx}-{solde_n1_col}{total_row_idx})/ABS({solde_n1_col}{total_row_idx})*100)"

        # Récupérer l'intitulé du compte 4 chiffres depuis le mapping
        account_label = get_account_label(aggregate.code_4)
        display_label = account_label if account_label else (aggregate.label or "")

        cells_total = [
            ws.cell(row=total_row_idx, column=COL_COMPTE, value=aggregate.code_4),
            ws.cell(row=total_row_idx, column=COL_INTITULE, value=display_label),
            ws.cell(row=total_row_idx, column=COL_REF, value=""),
        ]

        for cell in cells_total:
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL
            cell.border = _THIN_BORDER

        # Ajouter les cellules de formules avec formatage
        for col, formula in [
            (COL_SOLDE_N, total_n_formula),
            (COL_SOLDE_N1, total_n1_formula),
            (COL_VARIATION, total_variation_formula),
            (COL_VARIATION_PCT, total_variation_pct_formula),
        ]:
            c = ws.cell(row=total_row_idx, column=col)
            c.value = formula
            c.font = _TOTAL_FONT
            c.fill = _TOTAL_FILL
            c.border = _THIN_BORDER
            c.alignment = Alignment(horizontal="right")
            # Format pourcentage pour Variation %
            c.number_format = '0.0"%"' if col == COL_VARIATION_PCT else _NUM_FMT

        current_row += 1

        # ── Commentaire analytique (T023 / FR-008, FR-009) ───────────────
        comment = self._comment_writer.build_comment(
            display_label, total_solde_n, total_solde_n1, self.exercice
        )
        comment_cell = ws.cell(row=current_row, column=COL_COMPTE, value=comment.text_rendered)
        comment_cell.font = _DATA_FONT
        comment_cell.border = _THIN_BORDER
        comment_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Fusionner les colonnes du commentaire pour meilleure lisibilité
        ws.merge_cells(f"A{current_row}:G{current_row}")

        self._set_column_widths(ws)
    def _set_column_widths(self, ws) -> None:
        widths = {
            COL_COMPTE: 14,
            COL_INTITULE: 44,
            COL_REF: 8,
            COL_SOLDE_N: 16,
            COL_SOLDE_N1: 16,
            COL_VARIATION: 14,
            COL_VARIATION_PCT: 13,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
